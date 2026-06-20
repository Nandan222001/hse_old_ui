"""
Org Admin — Data Management controller.
Handles: import history, validation logs, API integrations,
         full xlsx import, document upload/list/delete,
         and the per-sheet ingestion endpoint used by Excel/CSV tab.
"""

import os
import datetime
import logging
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import Response
from sqlalchemy.orm import Session

from app.config.database import get_db
from app.core.dependencies import CurrentUser, get_current_user
from app.models.data_import import DataImport
from app.models.validation_log import ValidationLog
from app.models.api_integration import ApiIntegration
from app.models.document import Document

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/org-admin/data-management", tags=["Data Management"])

UPLOADS_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "uploads", "documents")
os.makedirs(UPLOADS_DIR, exist_ok=True)


# ── helpers ───────────────────────────────────────────────────────────────────

def _uploading_user(request: Request) -> str:
    return (
        request.headers.get("X-User-Email")
        or request.headers.get("X-User-Name")
        or "Admin"
    )


def _size_str(n_bytes: int) -> str:
    kb = n_bytes / 1024
    if kb < 1024:
        return f"{kb:.1f} KB"
    return f"{kb / 1024:.1f} MB"


def _import_to_dict(row: DataImport) -> dict:
    return {
        "id": row.id,
        "file_name": row.file_name,
        "import_type": row.import_type,
        "data_type": row.data_type,
        "records_total": row.records_total,
        "records_success": row.records_success,
        "records_failed": row.records_failed,
        "status": row.status,
        "uploaded_by": row.uploaded_by,
        "created_at": row.created_at.isoformat() if row.created_at else "",
    }


def _vlog_to_dict(row: ValidationLog) -> dict:
    return {
        "id": row.id,
        "file_name": row.file_name,
        "rule": row.rule,
        "status": row.status,
        "records_affected": row.records_affected,
        "message": row.message,
        "timestamp": row.timestamp or (row.created_at.isoformat() if row.created_at else ""),
    }


def _integration_to_dict(row: ApiIntegration) -> dict:
    return {
        "id": row.id,
        "name": row.name,
        "type": row.type,
        "endpoint_url": row.endpoint_url,
        "auth_type": row.auth_type,
        "is_active": row.is_active,
        "sync_frequency": row.sync_frequency,
        "description": row.description,
        "last_sync": row.last_sync,
        "created_at": row.created_at.isoformat() if row.created_at else "",
    }


def _doc_to_dict(row: Document) -> dict:
    return {
        "id": str(row.id),
        "file_name": row.file_name,
        "file_type": row.file_type,
        "category": row.category,
        "record_type": row.record_type,
        "size": row.size,
        "uploaded_by": row.uploaded_by,
        "created_at": row.created_at.isoformat() if row.created_at else "",
    }


# ── Import History ─────────────────────────────────────────────────────────────

@router.get("/imports")
def list_imports(db: Session = Depends(get_db)) -> dict:
    rows = db.query(DataImport).order_by(DataImport.id.desc()).limit(200).all()
    return {"data": [_import_to_dict(r) for r in rows]}


@router.post("/imports")
async def create_import(request: Request, db: Session = Depends(get_db)) -> dict:
    try:
        body = await request.json()
    except Exception:
        body = {}
    payload = body.get("data", body)
    count = int(payload.get("records_estimated") or payload.get("records_total") or 0)
    row = DataImport(
        file_name=payload.get("file_name", "unknown"),
        import_type=payload.get("import_type", "excel"),
        data_type=payload.get("data_type", "Unknown"),
        records_total=count,
        records_success=count,
        records_failed=0,
        status=payload.get("status", "success"),
        uploaded_by=_uploading_user(request),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"data": _import_to_dict(row)}


# ── Validation Logs ────────────────────────────────────────────────────────────

@router.get("/validation-logs")
def list_validation_logs(db: Session = Depends(get_db)) -> dict:
    rows = db.query(ValidationLog).order_by(ValidationLog.id.desc()).limit(500).all()
    return {"data": [_vlog_to_dict(r) for r in rows]}


# ── API Integrations ───────────────────────────────────────────────────────────

@router.get("/integrations")
def list_integrations(db: Session = Depends(get_db)) -> dict:
    rows = db.query(ApiIntegration).filter(ApiIntegration.is_active == True).all()
    return {"data": [_integration_to_dict(r) for r in rows]}


@router.post("/integrations")
async def create_integration(request: Request, db: Session = Depends(get_db)) -> dict:
    try:
        body = await request.json()
    except Exception:
        body = {}
    payload = body.get("data", body)
    row = ApiIntegration(
        name=payload.get("name", "Integration"),
        type=payload.get("type", "custom"),
        endpoint_url=payload.get("endpoint_url"),
        auth_type=payload.get("auth_type", "api_key"),
        is_active=bool(payload.get("is_active", True)),
        sync_frequency=payload.get("sync_frequency", "realtime"),
        description=payload.get("description"),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"data": _integration_to_dict(row)}


@router.delete("/integrations/{integration_id}")
def delete_integration(integration_id: int, db: Session = Depends(get_db)) -> dict:
    row = db.query(ApiIntegration).filter(ApiIntegration.id == integration_id).first()
    if row:
        db.delete(row)
        db.commit()
    return {"deleted": True}


# ── Documents ──────────────────────────────────────────────────────────────────

@router.get("/documents")
def list_documents(db: Session = Depends(get_db)) -> dict:
    rows = db.query(Document).order_by(Document.id.desc()).all()
    return {"data": {"items": [_doc_to_dict(r) for r in rows]}}


@router.post("/documents/upload")
async def upload_document(
    request: Request,
    file: UploadFile = File(...),
    category: str = Form("pdf"),
    record_type: str = Form(""),
    db: Session = Depends(get_db),
) -> dict:
    content = await file.read()
    size = _size_str(len(content))
    ext = (file.filename or "file").rsplit(".", 1)[-1].upper() if "." in (file.filename or "") else "FILE"

    # Save file to disk
    safe_name = (file.filename or "upload").replace(" ", "_")
    dest = os.path.join(UPLOADS_DIR, f"{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}_{safe_name}")
    try:
        with open(dest, "wb") as fh:
            fh.write(content)
    except Exception as exc:
        logger.warning("Could not save document to disk: %s", exc)
        dest = ""

    row = Document(
        file_name=file.filename or "unnamed",
        file_type=ext,
        category=category,
        record_type=record_type or None,
        size=size,
        uploaded_by=_uploading_user(request),
        file_path=dest,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"data": _doc_to_dict(row)}


@router.delete("/documents/{doc_id}")
def delete_document(doc_id: int, db: Session = Depends(get_db)) -> dict:
    row = db.query(Document).filter(Document.id == doc_id).first()
    if row:
        if row.file_path and os.path.exists(row.file_path):
            try:
                os.remove(row.file_path)
            except Exception:
                pass
        db.delete(row)
        db.commit()
    return {"deleted": True}


# ── Full Import (all 17 sheets) ────────────────────────────────────────────────

# Agent assignment for display in FullImportCard
_SHEET_AGENT: dict[str, str] = {
    "Organisation":     "MasterDataAgent",
    "Hazard_Categories":"SafetyOpsAgent",
    "Hazards":          "SafetyOpsAgent",
    "Roles":            "MasterDataAgent",
    "Sites":            "MasterDataAgent",
    "Permit_Types":     "ComplianceAgent",
    "Training_Programs":"PeopleAgent",
    "Policies":         "ComplianceAgent",
    "Departments":      "MasterDataAgent",
    "Working_Stations": "MasterDataAgent",
    "Employees":        "PeopleAgent",
    "dept_managers":    "MasterDataAgent",
    "Permits_To_Work":  "ComplianceAgent",
    "Incidents":        "SafetyOpsAgent",
    "Near_Misses":      "SafetyOpsAgent",
    "Safety_Walks":     "SafetyOpsAgent",
    "CAPA_Actions":     "ComplianceAgent",
    "Shift_Schedule":   "MasterDataAgent",
}


@router.post("/full-import")
async def full_import(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> dict:
    from app.services.excel_import_service import SHEET_STEPS, capture_tenant_table_max_ids, link_new_rows_to_org
    from sqlalchemy import text as _text

    content = await file.read()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot open Excel file: {exc}")

    # Resolve org_id at request time from the database; JWT org_id may be stale
    # immediately after organisation setup.
    org_id = current_user.org_id
    if not org_id or org_id <= 0:
        raise HTTPException(status_code=400, detail="Organisation setup is not complete. Complete setup before importing data.")
    before_ids = capture_tenant_table_max_ids(db)

    per_sheet = []
    total_processed = 0
    total_failed = 0

    try:
        for sheet_label, table_key, fn in SHEET_STEPS:
            processed = 0
            failed = 0
            errors: list[str] = []
            agent = _SHEET_AGENT.get(sheet_label, "MasterDataAgent")
            try:
                # SET inside the transaction so it stays on the same connection as INSERTs
                db.execute(_text("SET FOREIGN_KEY_CHECKS=0"))
                processed = fn(db, wb)
                db.commit()
            except Exception as exc:
                logger.error("full-import: error in %s: %s", sheet_label, exc, exc_info=True)
                errors.append(str(exc))
                failed = 1
                db.rollback()
            per_sheet.append({
                "sheet": sheet_label,
                "agent": agent,
                "processed": processed,
                "failed": failed,
                "errors": errors,
            })
            total_processed += processed
            total_failed += failed
    finally:
        try:
            db.execute(_text("SET FOREIGN_KEY_CHECKS=1"))
            db.commit()
        except Exception:
            pass

    # Stamp organisation_id only on rows inserted by this upload.
    if org_id and org_id > 0:
        link_new_rows_to_org(db, org_id, before_ids)
        db.commit()

    # Log the import
    import_row = DataImport(
        file_name=file.filename or "full_import.xlsx",
        import_type="excel",
        data_type="Full Import (all sheets)",
        records_total=total_processed + total_failed,
        records_success=total_processed,
        records_failed=total_failed,
        status="success" if total_failed == 0 else "partial",
        uploaded_by=_uploading_user(request),
    )
    db.add(import_row)
    db.commit()

    return {
        "total_processed": total_processed,
        "total_failed": total_failed,
        "per_sheet": per_sheet,
    }


# ── Full Template Download ─────────────────────────────────────────────────────

_TEMPLATE_SHEETS: list[tuple[str, list[str], list[list[Any]]]] = [
    ("Organisation", [
        "Org_ID", "Organisation_Name", "Country", "Industry_Sector",
        "Number_of_Employees", "Headquarters_Location", "Parent_Company",
        "ISO_45001_Status", "Regulatory_Authority", "Establishment_Date",
    ], [["ORG001", "Example Corp", "UK", "Manufacturing", 500, "London", "", "Certified", "HSE", "2010-01-01"]]),
    ("Hazard_Categories", ["Category_ID", "Category_Name", "Description"],
     [["HC001", "Chemical", "Chemical hazards"]]),
    ("Hazards", ["Hazard_ID", "Category_ID", "Hazard_Name", "Severity", "Probability"],
     [["HAZ001", "HC001", "Chemical Spill", "Serious", "Possible"]]),
    ("Roles", ["Role_ID", "Role_Name", "Job_Category", "Authority_Level", "Permit_Authority", "Safety_Signatory"],
     [["ROLE001", "HSE Manager", "Management", "Senior", "Yes", "Yes"]]),
    ("Sites", [
        "Site_ID", "Site_Name", "Address", "Postcode", "City", "Type",
        "Operational_Status", "Number_of_Working_Stations", "Employee_Count",
        "Primary_Products", "Hazard_Classification",
    ], [["SITE001", "Main Site", "123 Industrial Ave", "CF31 3TR", "Bridgend",
         "Manufacturing", "Active", 20, 150, "Steel Parts", "High Risk"]]),
    ("Permit_Types", ["Permit_Type_ID", "Permit_Type_Name", "Risk_Level",
                      "Validity_Period_Hours", "Concurrent_Limit"],
     [["PT001", "Hot Work", "High", 8, 3]]),
    ("Training_Programs", ["Training_ID", "Training_Name", "Duration_Hours",
                            "Frequency", "Certification", "Expiry_Months"],
     [["TRN001", "Fire Safety", 4, "Annual", "Fire Safety Certificate", 12]]),
    ("Policies", ["Policy_ID", "Policy_Name", "Category", "Issue_Date", "Owner", "Status"],
     [["POL001", "Health & Safety Policy", "Safety", "2024-01-01", "HSE Manager", "Active"]]),
    ("Departments", ["Dept_ID", "Site_ID", "Department_Name", "Manager_ID", "Number_of_Teams"],
     [["DEPT001", "SITE001", "Operations", "EMP001", 3]]),
    ("Working_Stations", [
        "Station_ID", "Station_Name", "Site_ID", "Department", "Zone_Classification",
        "Primary_Hazard_ID", "Staffing_Requirement", "Equipment_List",
        "Permit_Types_Required", "Access_Restrictions",
    ], [["STN001", "Assembly Line A", "SITE001", "DEPT001", "Zone 1",
         "HAZ001", 5, "Conveyor Belt", "Hot Work", "Authorised Personnel"]]),
    ("Employees", [
        "Employee_ID", "Full_Name", "Date_of_Birth", "Gender", "Employment_Type",
        "Employment_Start_Date", "Current_Role_ID", "Department_ID",
        "Shift_Pattern", "Manager_ID", "Induction_Date", "Active_Status",
    ], [["EMP001", "Jane Smith", "1985-03-15", "F", "Permanent",
         "2020-01-10", "ROLE001", "DEPT001", "Days", "", "2020-01-10", "Active"]]),
    ("Permits_To_Work", [
        "Permit_ID", "Permit_Type_ID", "Date_Issued", "Time_Issued",
        "Location_Station_ID", "Work_Description", "Duration_Requested_Hours",
        "Issued_By", "Approved_By", "Validity_Start", "Validity_End",
        "Work_Start_Actual", "Work_End_Actual", "Number_of_Workers",
        "Status", "Deviation_Reported", "Incident_Occurred",
    ], [["PTW001", "PT001", "2024-03-01", "08:00", "STN001",
         "Welding on roof", 4, "EMP001", "EMP001",
         "2024-03-01 08:00", "2024-03-01 16:00",
         "2024-03-01 08:00", "2024-03-01 12:00", 3, "active", "No", "No"]]),
    ("Incidents", [
        "Incident_ID", "Report_Date", "Incident_DateTime", "Location_Station_ID",
        "Incident_Type", "Severity", "Number_Persons_Involved", "Description",
        "Immediate_Cause", "Root_Cause", "Hazard_ID", "Permit_Active",
        "Control_Failure", "Reported_By", "Investigation_Status",
        "CAPA_Generated", "Days_Away", "Root_Cause_Category",
    ], [["INC001", "2024-03-10", "2024-03-10 09:30", "STN001",
         "incident_report", "medium", 1, "Slip on wet floor",
         "Wet surface", "Housekeeping", "HAZ001", "No", "Yes",
         "EMP001", "Open", "No", 0, "Environment"]]),
    ("Near_Misses", [
        "Near_Miss_ID", "Report_Date", "Event_DateTime", "Location_Station",
        "Description", "Potential_Consequence", "Hazard_Involved",
        "Underlying_Cause", "Control_Failure", "Reported_By", "CAPA_Escalation",
    ], [["NM001", "2024-03-09", "2024-03-09 09:39", "STN001",
         "Near-miss description", "Injury", "HAZ001",
         "Procedure Gap", "No", "EMP001", "Yes"]]),
    ("Safety_Walks", [
        "Walk_ID", "Inspection_DateTime", "Location_Station_ID", "Inspector_ID",
        "Inspection_Type", "Issues_Found", "Critical_Issues",
        "Housekeeping_Rating", "Compliance_Rating", "Follow_Up_Required",
    ], [["SW001", "2024-03-15 10:00", "STN001", "EMP001",
         "Routine", 2, 0, "Good", "Satisfactory", "No"]]),
    ("CAPA_Actions", [
        "Action_ID", "Incident_ID", "Action_Type", "Description",
        "Root_Cause_Addressed", "Responsible_Person", "Due_Date",
        "Status", "Effectiveness_Rating",
    ], [["CAPA001", "INC001", "Corrective", "Fix machine guard",
         "Training", "EMP001", "2024-05-18", "Completed", 4]]),
    ("Shift_Schedule", [
        "Schedule_ID", "Employee_ID", "Shift_Date", "Shift_Type",
        "Shift_Start", "Shift_End", "Actual_Hours_Worked",
        "Station_Assigned", "Supervisor",
    ], [["SCH001", "EMP001", "2024-06-01", "Morning",
         "06:00", "14:00", 8.0, "STN001", "EMP001"]]),
]


@router.get("/full-template")
def download_full_template() -> Response:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for sheet_name, headers, sample_rows in _TEMPLATE_SHEETS:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        for row in sample_rows:
            ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=HSE_Full_Import_Template.xlsx"},
    )


# ── Single-sheet / Shift-schedule ingestion ────────────────────────────────────

@router.post("/import/file")
async def import_single_file(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> dict:
    """
    Generic single-sheet import used by the Excel tab for Shift Schedule.
    Returns {summary: {total_inserted}, sheets: [{status, header_errors, row_errors}]}.
    """
    from app.services.excel_import_service import _insert_shift_schedule, capture_tenant_table_max_ids, link_new_rows_to_org

    content = await file.read()
    filename = file.filename or "upload"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    header_errors: list[str] = []
    row_errors: list[dict] = []
    count = 0
    status = "ok"
    if not current_user.org_id or current_user.org_id <= 0:
        raise HTTPException(status_code=400, detail="Organisation setup is not complete. Complete setup before importing data.")
    before_ids = capture_tenant_table_max_ids(db)

    try:
        if ext == "csv":
            import csv, io
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Shift_Schedule"
            reader = csv.reader(io.StringIO(content.decode("utf-8-sig", errors="replace")))
            for row in reader:
                ws.append(row)
        else:
            wb = openpyxl.load_workbook(BytesIO(content), read_only=False, data_only=True)
            if "Shift_Schedule" not in wb.sheetnames and wb.sheetnames:
                wb.active.title = "Shift_Schedule"

        count = _insert_shift_schedule(db, wb)
        if current_user.org_id and current_user.org_id > 0:
            link_new_rows_to_org(db, current_user.org_id, before_ids)
        db.commit()
    except Exception as exc:
        db.rollback()
        header_errors.append(str(exc))
        status = "rejected"
        count = 0

    # Log the import
    import_row = DataImport(
        file_name=filename,
        import_type="excel",
        data_type="Shift Schedule",
        records_total=count,
        records_success=count,
        records_failed=len(row_errors),
        status="success" if count > 0 else "failed",
        uploaded_by=_uploading_user(request),
    )
    db.add(import_row)
    try:
        db.commit()
    except Exception:
        db.rollback()

    return {
        "summary": {"total_inserted": count},
        "sheets": [{
            "sheet": "Shift_Schedule",
            "status": status,
            "header_errors": header_errors,
            "row_errors": row_errors,
        }],
    }
