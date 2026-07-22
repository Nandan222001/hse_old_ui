"""
Stub endpoints for features that have frontend UI but no backend model yet.
These return empty/default responses so pages render gracefully rather than
showing API errors.
"""
from io import BytesIO
from typing import Any, List, Dict, Tuple, Set, Optional
from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import JSONResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.config.database import get_db
from app.core.dependencies import get_current_user, CurrentUser
from app.models.organisation_invite import OrganisationInvite
from app.models.data_import import DataImport
from app.models.validation_log import ValidationLog

router = APIRouter(tags=["Stubs"])


# ── Compliance ────────────────────────────────────────────────────────────────

@router.get("/compliance-standards")
def list_compliance_standards() -> list:
    return []


@router.get("/audit-trail")
def get_audit_trail() -> list:
    return []


# ── Onboarding / Access Profile ───────────────────────────────────────────────

@router.get("/onboarding/access-profile")
def get_onboarding_access_profile(email: str = "", org_code: str = "") -> dict:
    return {"found": False, "approved": False}


@router.get("/onboarding/requests")
def list_onboarding_requests() -> list:
    return []


@router.post("/onboarding/requests")
def create_onboarding_request(payload: Any = None) -> dict:
    return {"uuid": None, "status": "pending"}


@router.delete("/onboarding/requests/{uuid}")
def delete_onboarding_request(uuid: str) -> dict:
    return {"deleted": True}


@router.get("/onboarding/requests/{uuid}")
def get_onboarding_request(uuid: str) -> dict:
    return {"uuid": uuid, "status": "pending"}


@router.patch("/onboarding/requests/{uuid}")
def update_onboarding_request(uuid: str, payload: Any = None) -> dict:
    return {"uuid": uuid, "status": "updated"}


@router.get("/onboarding/layer-options")
def get_onboarding_layer_options() -> dict:
    return {"layers": []}


@router.get("/onboarding/processing-queue")
def get_onboarding_processing_queue() -> list:
    return []


@router.post("/onboarding/theta-auth/login")
def theta_auth_login(payload: Any = None) -> JSONResponse:
    return JSONResponse(status_code=401, content={"detail": "Theta auth not configured"})


@router.post("/onboarding/password-reset/theta/request")
def theta_password_reset_request(payload: Any = None) -> dict:
    return {"sent": False}


@router.post("/onboarding/password-reset/theta/confirm")
def theta_password_reset_confirm(payload: Any = None) -> dict:
    return {"confirmed": False}


@router.post("/onboarding/password-reset/theta/direct")
def theta_password_reset_direct(payload: Any = None) -> dict:
    return {"reset": False}


@router.get("/onboarding/access-requests")
def list_onboarding_access_requests() -> list:
    return []


@router.patch("/onboarding/access-requests/{uuid}")
def update_onboarding_access_request(uuid: str, payload: Any = None) -> dict:
    return {"uuid": uuid, "updated": True}


@router.post("/onboarding")
def create_onboarding(payload: Any = None) -> dict:
    return {"created": False}


# ── AI Chat ───────────────────────────────────────────────────────────────────
# Moved to app/controllers/ai.py — real Claude integration


@router.get("/org-setup/progress")
def org_setup_progress(db: Session = Depends(get_db)) -> dict:
    from app.models.site import Site
    from app.models.organisation import Organisation
    steps_done = []
    # Use _wizard_org_id (set after step 1 is saved) to scope DB queries.
    # Before step 1 is saved, _wizard_org_id is None and the wizard is blank.
    if _wizard_step1:
        steps_done.append(1)
    elif _wizard_org_id and db.query(Organisation).filter(Organisation.id == _wizard_org_id).first():
        steps_done.append(1)
    if _wizard_step2:
        steps_done.append(2)
    if _wizard_org_id and db.query(Site).filter(Site.organisation_id == _wizard_org_id).first():
        steps_done.append(3)
    if _wizard_users:
        steps_done.append(4)
    if _wizard_step5:
        steps_done.append(5)
    if _wizard_documents or _wizard_imports:
        steps_done.append(6)
    if _wizard_step7:
        steps_done.append(7)
    total = 8
    return {
        "steps_completed": steps_done,
        "steps_total": total,
        "percent": round(len(steps_done) / total * 100),
        "activated": False,
    }

@router.get("/org-setup/step1")
def org_setup_step1_get(db: Session = Depends(get_db)) -> dict:
    if _wizard_step1:
        return _wizard_step1
    # Only read org from DB if step 1 was already saved this session.
    if not _wizard_org_id:
        return {}
    from app.models.organisation import Organisation
    org = db.query(Organisation).filter(Organisation.id == _wizard_org_id).first()
    if not org:
        return {}
    return {
        "organisationId":       str(org.id),
        "organisationName":     org.organisation_name or "",
        "country":              org.country or "",
        "industrySector":       org.industry_sector or "",
        "numberOfEmployees":    org.number_of_employees,
        "headquartersLocation": org.headquarters_location or "",
        "parentCompany":        org.parent_company or "",
        "iso45001Status":       org.iso_45001_status or "",
        "regulatoryAuthority":  org.regulatory_authority or "",
        "establishmentDate":    org.establishment_date.isoformat() if org.establishment_date else "",
    }

@router.post("/org-setup/step1")
def org_setup_step1_post(request: Request, payload: dict, db: Session = Depends(get_db)) -> dict:
    global _wizard_step1, _wizard_org_id
    d = payload.get("data", payload)
    _wizard_step1 = dict(d)
    from app.models.organisation import Organisation
    from datetime import date as _date
    est_date = None
    est = str(d.get("establishmentDate") or "")[:10]
    if est:
        try:
            est_date = _date.fromisoformat(est)
        except Exception:
            pass
    ne = d.get("numberOfEmployees")
    try:
        num_employees = int(ne) if ne else None
    except Exception:
        num_employees = None
    if _wizard_org_id:
        # Same wizard session — update this org's details (user editing step 1)
        org = db.query(Organisation).filter(Organisation.id == _wizard_org_id).first()
        if org:
            name = (d.get("organisationName") or "").strip()
            if name:
                org.organisation_name = name
            org.country = (d.get("country") or "").strip() or None
            org.industry_sector = (d.get("industrySector") or "").strip() or None
            org.number_of_employees = num_employees
            org.headquarters_location = (d.get("headquartersLocation") or "").strip() or None
            org.parent_company = (d.get("parentCompany") or "").strip() or None
            org.iso_45001_status = (d.get("iso45001Status") or "").strip() or None
            org.regulatory_authority = (d.get("regulatoryAuthority") or "").strip() or None
            if est_date:
                org.establishment_date = est_date
        else:
            # _wizard_org_id stale (e.g. after restart) — create fresh
            org = Organisation(
                organisation_name=(d.get("organisationName") or "New Organisation").strip(),
                country=(d.get("country") or "").strip() or None,
                industry_sector=(d.get("industrySector") or "").strip() or None,
                number_of_employees=num_employees,
                headquarters_location=(d.get("headquartersLocation") or "").strip() or None,
                parent_company=(d.get("parentCompany") or "").strip() or None,
                iso_45001_status=(d.get("iso45001Status") or "").strip() or None,
                regulatory_authority=(d.get("regulatoryAuthority") or "").strip() or None,
                establishment_date=est_date,
            )
            db.add(org)
    else:
        # New wizard session — ALWAYS create a brand-new org so each org admin
        # gets their own separate Organisation record, never overwriting another org.
        org = Organisation(
            organisation_name=(d.get("organisationName") or "New Organisation").strip(),
            country=(d.get("country") or "").strip() or None,
            industry_sector=(d.get("industrySector") or "").strip() or None,
            number_of_employees=num_employees,
            headquarters_location=(d.get("headquartersLocation") or "").strip() or None,
            parent_company=(d.get("parentCompany") or "").strip() or None,
            iso_45001_status=(d.get("iso45001Status") or "").strip() or None,
            regulatory_authority=(d.get("regulatoryAuthority") or "").strip() or None,
            establishment_date=est_date,
        )
        db.add(org)
    db.commit()
    db.refresh(org)
    _wizard_org_id = org.id  # scope all subsequent wizard DB queries to this org

    # Also link the admin user to this org right away (so dashboard shows data before activate)
    from app.services.auth_service import decode_access_token
    from app.models.user import User
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        token_data = decode_access_token(auth_header.removeprefix("Bearer ").strip())
        if token_data:
            admin_email = (token_data.get("email") or "").strip().lower()
            admin_user = db.query(User).filter(User.email == admin_email).first()
            if admin_user and admin_user.organisation_id is None:
                admin_user.organisation_id = org.id
                db.commit()

    return {"saved": True}

@router.post("/org-setup/step1/parse-excel")
async def org_setup_parse_excel(file: UploadFile = File(...)) -> dict:
    """
    Parse an uploaded Organisation Excel/CSV file and return the first
    organisation row's fields so the wizard can pre-populate the form.

    Expected sheet name: "Organisation"
    Column order (0-based, row 1 = headers, row 2+ = data):
      0  org_id | 1  org_name | 2  country | 3  industry_sector
      4  num_employees | 5  hq_location | 6  parent_company
      7  iso_45001_status | 8  regulatory_authority | 9  establishment_date

    CSV fallback: same column order, first row headers.
    """
    if not file.filename:
        return {"_error": "No file received"}

    fname = file.filename.lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".xls") or fname.endswith(".csv")):
        return {"_error": "Only .xlsx, .xls, or .csv files are accepted"}

    content = await file.read()
    if not content:
        return {"_error": "Uploaded file is empty"}

    try:
        if fname.endswith(".csv"):
            import csv, io
            reader = csv.reader(io.StringIO(content.decode("utf-8-sig", errors="replace")))
            rows = list(reader)
            data_rows = rows[1:] if len(rows) > 1 else []
        else:
            try:
                import openpyxl
            except ImportError:
                return {"_error": "openpyxl is not installed — cannot parse .xlsx files"}

            wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

            # Accept "Organisation" sheet or fall back to the first sheet
            ws = None
            for candidate in ("Organisation", "organisation", "Organization", "organization"):
                if candidate in wb.sheetnames:
                    ws = wb[candidate]
                    break
            if ws is None:
                ws = wb.worksheets[0]

            all_rows = list(ws.iter_rows(values_only=True))
            data_rows = [list(r) for r in all_rows[1:]] if len(all_rows) > 1 else []

        # Find first non-empty data row
        row = None
        for r in data_rows:
            if any(c for c in r if c is not None and str(c).strip()):
                row = list(r)
                break

        if not row:
            return {"_error": "No data rows found in the file. Check the sheet name and format."}

        def _cell(idx: int) -> str:
            val = row[idx] if idx < len(row) else None
            return str(val).strip() if val is not None and str(val).strip() else ""

        return {
            "organisationId":       _cell(0),
            "organisationName":     _cell(1),
            "country":              _cell(2),
            "industrySector":       _cell(3),
            "numberOfEmployees":    _cell(4),
            "headquartersLocation": _cell(5),
            "parentCompany":        _cell(6),
            "iso45001Status":       _cell(7),
            "regulatoryAuthority":  _cell(8),
            "establishmentDate":    _cell(9),
        }

    except Exception as exc:
        return {"_error": f"Could not parse file: {exc}"}

@router.post("/org-setup/step1/api-connect")
def org_setup_api_connect(payload: Any = None) -> dict:
    return {"data": {}}

@router.get("/org-setup/step1/template")
def org_setup_step1_template() -> dict:
    return {}

@router.get("/org-setup/step2")
def org_setup_step2_get(db: Session = Depends(get_db)) -> dict:
    from app.models.organisation import Organisation
    org = db.query(Organisation).order_by(Organisation.id.desc()).first()
    if org and org.compliance_config:
        return org.compliance_config
    return _wizard_step2

@router.post("/org-setup/step2")
def org_setup_step2_post(payload: dict, db: Session = Depends(get_db)) -> dict:
    global _wizard_step2
    from app.models.organisation import Organisation
    d = payload.get("data", payload)
    _wizard_step2 = dict(d)
    org = db.query(Organisation).order_by(Organisation.id.desc()).first()
    if org:
        org.compliance_config = dict(d)
        db.commit()
    return {"saved": True}

# ── Wizard in-memory state ────────────────────────────────────────────────────
# All state persists until server restart (acceptable for a one-time setup flow).
_wizard_users: List[dict] = []
_wizard_user_id = 0
_wizard_step1: dict = {}
_wizard_step2: dict = {}
_wizard_step5: dict = {}
_wizard_step7: dict = {}
_wizard_documents: List[dict] = []
_wizard_doc_id: int = 0
_wizard_imports: List[dict] = []
_wizard_import_id: int = 0
# Tracks the org being configured in the current wizard session.
# Set when step 1 is saved; cleared on reset. Used by step 3/4 to scope DB queries.
_wizard_org_id: Optional[int] = None


def _reset_wizard_state() -> None:
    global _wizard_users, _wizard_user_id, _wizard_step1, _wizard_step2
    global _wizard_step5, _wizard_step7, _wizard_documents, _wizard_doc_id
    global _wizard_imports, _wizard_import_id, _wizard_org_id
    _wizard_users = []
    _wizard_user_id = 0
    _wizard_step1 = {}
    _wizard_step2 = {}
    _wizard_step5 = {}
    _wizard_step7 = {}
    _wizard_documents = []
    _wizard_doc_id = 0
    _wizard_imports = []
    _wizard_import_id = 0
    _wizard_org_id = None


@router.post("/org-setup/reset")
def org_setup_reset() -> dict:
    """Clear wizard in-memory state so a fresh setup can begin."""
    _reset_wizard_state()
    return {"reset": True}


def _next_user_id() -> str:
    global _wizard_user_id
    _wizard_user_id += 1
    return f"wiz-{_wizard_user_id}"


def _site_to_dict(s) -> dict:
    return {
        "id":                   str(s.id),
        "name":                 s.site_name or "",
        "type":                 s.type or "",
        "address":              s.address or "",
        "postcode":             s.postcode or "",
        "city":                 s.city or "",
        "operationalStatus":    s.operational_status or "",
        "workingStations":      s.number_of_working_stations,
        "capacity":             s.capacity,
        "primaryProducts":      s.primary_products or "",
        "hazardClassification": s.hazard_classification or "",
    }


def _parse_sites_file(content: bytes, filename: str) -> List[dict]:
    """Return a list of site dicts parsed from an xlsx/xls/csv file."""
    from io import BytesIO
    fname = filename.lower()
    rows: List[list] = []

    if fname.endswith(".csv"):
        import csv, io
        reader = csv.reader(io.StringIO(content.decode("utf-8-sig", errors="replace")))
        all_rows = list(reader)
        rows = [list(r) for r in all_rows[1:]] if len(all_rows) > 1 else []
    else:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        # Prefer "Sites" sheet; fall back to first sheet
        ws = wb["Sites"] if "Sites" in wb.sheetnames else wb.worksheets[0]
        all_rows = list(ws.iter_rows(values_only=True))
        rows = [list(r) for r in all_rows[1:]] if len(all_rows) > 1 else []

    def _s(val) -> str:
        return str(val).strip() if val is not None and str(val).strip() else ""

    def _i(val):
        try: return int(val)
        except Exception: return None

    sites = []
    for r in rows:
        if not any(c for c in r if c is not None and str(c).strip()):
            continue
        # Full workbook format: col 0 = site_id prefix, col 1 = name … col 10 = hazard
        # Simple wizard template: col 0 = name, col 1 = type, col 2 = city, …
        # Detect by checking if col 0 looks like a prefixed ID (e.g. "SITE001")
        col0 = _s(r[0] if len(r) > 0 else "")
        if col0.upper().startswith("SITE") or (col0 and col0[0].isalpha() and any(c.isdigit() for c in col0)):
            # Full workbook format (offset by 1)
            sites.append({
                "name":                 _s(r[1] if len(r) > 1 else None),
                "type":                 _s(r[5] if len(r) > 5 else None),
                "address":              _s(r[2] if len(r) > 2 else None),
                "postcode":             _s(r[3] if len(r) > 3 else None),
                "city":                 _s(r[4] if len(r) > 4 else None),
                "operational_status":   _s(r[6] if len(r) > 6 else None),
                "working_stations":     _i(r[7] if len(r) > 7 else None),
                "capacity":             _i(r[8] if len(r) > 8 else None),
                "primary_products":     _s(r[9] if len(r) > 9 else None),
                "hazard_classification":_s(r[10] if len(r) > 10 else None),
            })
        else:
            # Simple wizard template: name | type | address | city | operational_status
            sites.append({
                "name":                 _s(r[0] if len(r) > 0 else None),
                "type":                 _s(r[1] if len(r) > 1 else None),
                "address":              _s(r[2] if len(r) > 2 else None),
                "postcode":             _s(r[3] if len(r) > 3 else None),
                "city":                 _s(r[4] if len(r) > 4 else None),
                "operational_status":   _s(r[5] if len(r) > 5 else None),
                "working_stations":     _i(r[6] if len(r) > 6 else None),
                "capacity":             _i(r[7] if len(r) > 7 else None),
                "primary_products":     _s(r[8] if len(r) > 8 else None),
                "hazard_classification":_s(r[9] if len(r) > 9 else None),
            })
    return [s for s in sites if s["name"]]


@router.get("/org-setup/step3/sites")
def org_setup_step3_sites(db: Session = Depends(get_db)) -> list:
    from app.models.site import Site
    if not _wizard_org_id:
        return []
    sites = db.query(Site).filter(Site.organisation_id == _wizard_org_id).order_by(Site.id.asc()).all()
    return [_site_to_dict(s) for s in sites]


@router.post("/org-setup/step3/site")
def org_setup_step3_create_site(payload: dict, db: Session = Depends(get_db)) -> dict:
    from app.models.site import Site
    from app.models.organisation import Organisation
    d = payload.get("data", payload)
    org_id = _wizard_org_id
    if not org_id:
        org = db.query(Organisation).order_by(Organisation.id.desc()).first()
        org_id = org.id if org else None
    site = Site(
        site_name=            (d.get("name") or "").strip(),
        type=                 (d.get("type") or "").strip() or None,
        address=              (d.get("address") or "").strip() or None,
        postcode=             (d.get("postcode") or "").strip() or None,
        city=                 (d.get("city") or "").strip() or None,
        operational_status=   (d.get("operationalStatus") or "").strip() or None,
        number_of_working_stations=d.get("workingStations") or None,
        capacity=             d.get("capacity") or None,
        primary_products=     (d.get("primaryProducts") or "").strip() or None,
        hazard_classification=(d.get("hazardClassification") or "").strip() or None,
        organisation_id=      org_id,
    )
    db.add(site)
    db.commit()
    db.refresh(site)
    return _site_to_dict(site)


@router.post("/org-setup/step3/bulk")
async def org_setup_step3_bulk(file: UploadFile = File(...), db: Session = Depends(get_db)) -> dict:
    from app.models.site import Site
    if not file.filename:
        return {"count": 0, "error": "No file received"}

    fname = file.filename.lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".xls") or fname.endswith(".csv")):
        return {"count": 0, "error": "Only .xlsx, .xls, or .csv files are accepted"}

    content = await file.read()
    if not content:
        return {"count": 0, "error": "Uploaded file is empty"}

    try:
        site_dicts = _parse_sites_file(content, file.filename)
    except Exception as exc:
        return {"count": 0, "error": f"Could not parse file: {exc}"}

    if not site_dicts:
        return {"count": 0, "error": "No site rows found in the file. Check the sheet name and column format."}

    from app.models.organisation import Organisation
    org_id = _wizard_org_id
    if not org_id:
        org = db.query(Organisation).order_by(Organisation.id.desc()).first()
        org_id = org.id if org else None

    count = 0
    for d in site_dicts:
        site = Site(
            site_name=            d["name"],
            type=                 d["type"] or None,
            address=              d["address"] or None,
            postcode=             d["postcode"] or None,
            city=                 d["city"] or None,
            operational_status=   d["operational_status"] or None,
            number_of_working_stations=d["working_stations"],
            capacity=             d["capacity"],
            primary_products=     d["primary_products"] or None,
            hazard_classification=d["hazard_classification"] or None,
            organisation_id=      org_id,
        )
        db.add(site)
        count += 1

    db.commit()
    return {"count": count}


@router.get("/org-setup/step3/template")
def org_setup_step3_template():
    """Return a downloadable CSV template for the sites bulk upload."""
    from fastapi.responses import Response
    header = "Site Name,Type,Address,Postcode,City,Operational Status,Working Stations,Capacity,Primary Products,Hazard Classification\n"
    example = "Main Plant,Site,123 Industrial Road,SW1A 1AA,London,Active,50,500,Steel Components,High Risk\n"
    return Response(
        content=(header + example).encode(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=sites_template.csv"},
    )


# ── Step 4 — Users ────────────────────────────────────────────────────────────

def _parse_users_file(content: bytes, filename: str) -> List[dict]:
    """Return a list of user dicts parsed from xlsx/csv."""
    fname = filename.lower()
    rows: List[list] = []

    if fname.endswith(".csv"):
        import csv, io
        reader = csv.reader(io.StringIO(content.decode("utf-8-sig", errors="replace")))
        all_rows = list(reader)
        rows = [list(r) for r in all_rows[1:]] if len(all_rows) > 1 else []
    else:
        from io import BytesIO
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb["Users"] if "Users" in wb.sheetnames else wb.worksheets[0]
        all_rows = list(ws.iter_rows(values_only=True))
        rows = [list(r) for r in all_rows[1:]] if len(all_rows) > 1 else []

    def _s(val) -> str:
        return str(val).strip() if val is not None and str(val).strip() else ""

    users = []
    for r in rows:
        if not any(c for c in r if c is not None and str(c).strip()):
            continue
        # Simple format: Name | Email | Role | Department
        users.append({
            "name":       _s(r[0] if len(r) > 0 else None),
            "email":      _s(r[1] if len(r) > 1 else None),
            "role":       _s(r[2] if len(r) > 2 else None),
            "department": _s(r[3] if len(r) > 3 else None),
        })
    return [u for u in users if u["name"] or u["email"]]



# Wizard role label → app_role.name mapping
_WIZARD_ROLE_MAP = {
    "site hse manager": "safety_manager",
    "hse manager":      "safety_manager",
    "safety manager":   "safety_manager",
    "supervisor":       "supervisor",
    "worker":           "operator",
    "operator":         "operator",
    "auditor":          "viewer",
    "viewer":           "viewer",
}


def _step4_org_id(request: Request, db: Session) -> Optional[int]:
    """Resolve the wizard organisation from persisted user state.

    The module-level wizard id is only a fallback: it is not shared between
    backend workers and is lost on restart.
    """
    from app.models.user import User
    from app.services.auth_service import decode_access_token

    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        token_data = decode_access_token(auth_header.removeprefix("Bearer ").strip())
        if token_data:
            user = None
            try:
                user_id = int(token_data.get("sub", 0))
            except (TypeError, ValueError):
                user_id = 0
            if user_id:
                user = db.query(User).filter(User.id == user_id).first()
            if user is None:
                email = (token_data.get("email") or "").strip().lower()
                if email:
                    user = db.query(User).filter(User.email == email).first()
            if user and user.organisation_id:
                return user.organisation_id
    return _wizard_org_id


def _create_wizard_user(d: dict, db: Session, org_id: Optional[int] = None) -> dict:
    """Create a real DB User from wizard Step-4 data. Returns serialisable dict."""
    import re, secrets, string, bcrypt
    from app.models.user import User
    from app.models.app_role import AppRole
    from app.models.organisation import Organisation

    name  = (d.get("name") or "").strip()
    email = (d.get("email") or "").strip().lower()
    if not email:
        return {}

    # If user already exists, re-link to this wizard's org and return (don't create duplicate)
    existing = db.query(User).filter(User.email == email).first()
    if existing:
        if org_id and existing.organisation_id != org_id:
            existing.organisation_id = org_id
            db.commit()
            db.refresh(existing)
        role_obj = db.query(AppRole).filter(AppRole.id == existing.app_role_id).first() if existing.app_role_id else None
        return {
            "id": str(existing.id), "name": existing.full_name or existing.username,
            "email": existing.email, "role": role_obj.label if role_obj else d.get("role", ""),
            "department": d.get("department", ""), "status": "active" if existing.is_active else "inactive",
        }

    # Map role label → app_role
    role_label = (d.get("role") or "").strip().lower()
    role_name  = _WIZARD_ROLE_MAP.get(role_label, "operator")
    app_role   = db.query(AppRole).filter(AppRole.name == role_name).first()
    if not app_role:
        app_role = db.query(AppRole).filter(AppRole.name == "operator").first()

    # Use the active wizard org; fall back to latest org only if wizard not started yet
    org = None
    if org_id:
        org = db.query(Organisation).filter(Organisation.id == org_id).first()
    if org is None:
        org = db.query(Organisation).order_by(Organisation.id.desc()).first()
    org_id = org.id if org else None

    # Unique username
    base = re.sub(r"[^a-z0-9_]", "_", email.split("@")[0].lower())
    username, suffix = base, 1
    while db.query(User).filter(User.username == username).first():
        username = f"{base}_{suffix}"; suffix += 1

    # Temp password
    alphabet     = string.ascii_letters + string.digits + "!@#$%"
    temp_password = "".join(secrets.choice(alphabet) for _ in range(12))
    pw_hash       = bcrypt.hashpw(temp_password.encode(), bcrypt.gensalt()).decode()

    user = User(
        username=username, full_name=name, email=email,
        password_hash=pw_hash, app_role_id=app_role.id,
        organisation_id=org_id, is_active=True,
    )
    db.add(user)
    db.commit()
    db.refresh(user)

    # Send invite email with credentials
    try:
        from app.services.email_service import send_email, _build_invite_html
        from app.config.settings import get_settings
        settings = get_settings()
        org_name = org.organisation_name if org else "your organisation"
        html = _build_invite_html(
            admin_name=name, organisation_name=org_name,
            email=email, temp_password=temp_password,
            login_url=f"{settings.frontend_url}/auth/login",
        )
        send_email(to_email=email, subject=f"You've been invited to {org_name} — HSE Intelligence",
                   html_content=html, to_name=name)
    except Exception:
        pass

    return {
        "id":         str(user.id),
        "name":       name,
        "email":      email,
        "role":       d.get("role", role_name),
        "department": d.get("department", ""),
        "status":     "active",
    }


@router.get("/org-setup/step4/users")
def org_setup_step4_users(request: Request, db: Session = Depends(get_db)) -> list:
    from app.models.user import User
    from app.models.app_role import AppRole
    from app.models.organisation import Organisation

    org_id = _step4_org_id(request, db)
    if not org_id:
        return list(_wizard_users)

    org = db.query(Organisation).filter(Organisation.id == org_id).first()
    if not org:
        return list(_wizard_users)

    role_map = {r.id: r for r in db.query(AppRole).all()}
    users = db.query(User).filter(
        User.organisation_id == org.id,
        User.app_role_id != None,
    ).all()

    # Exclude superadmin and admin role
    result = []
    for u in users:
        role = role_map.get(u.app_role_id)
        if role and role.name in ("superadmin", "admin"):
            continue
        result.append({
            "id":         str(u.id),
            "name":       u.full_name or u.username,
            "email":      u.email,
            "role":       role.label if role else "",
            "department": "",
            "status":     "active" if u.is_active else "inactive",
        })
    return result


@router.post("/org-setup/step4/user")
def org_setup_step4_create_user(request: Request, payload: dict, db: Session = Depends(get_db)) -> dict:
    d = payload.get("data", payload)
    result = _create_wizard_user(d, db, _step4_org_id(request, db))
    if not result:
        return {"id": _next_user_id(), "name": d.get("name",""), "email": d.get("email",""),
                "role": d.get("role",""), "department": d.get("department",""), "status": "active"}
    return result


@router.post("/org-setup/step4/bulk")
async def org_setup_step4_bulk(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)) -> dict:
    if not file.filename:
        return {"count": 0}

    fname = file.filename.lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".xls") or fname.endswith(".csv")):
        return {"count": 0, "error": "Only .xlsx, .xls, or .csv files are accepted"}

    content = await file.read()
    if not content:
        return {"count": 0, "error": "Uploaded file is empty"}

    try:
        user_dicts = _parse_users_file(content, file.filename)
    except Exception as exc:
        return {"count": 0, "error": f"Could not parse file: {exc}"}

    if not user_dicts:
        return {"count": 0, "error": "No user rows found. Expected columns: Name, Email, Role, Department"}

    imported_users = []
    org_id = _step4_org_id(request, db)
    for d in user_dicts:
        result = _create_wizard_user(d, db, org_id)
        if result:
            imported_users.append(result)

    return {"count": len(imported_users), "users": imported_users}


@router.post("/org-setup/step4/hrms-import")
def org_setup_hrms_import(payload: Any = None) -> dict:
    return {"count": 0, "error": "HRMS integration is not configured yet"}


@router.get("/org-setup/step4/template")
def org_setup_step4_template():
    """Return a downloadable CSV template for the users bulk upload."""
    from fastapi.responses import Response
    header = "Name,Email,Role,Department\n"
    example = "Jessica Hernandez,jessica@company.com,Supervisor,Heavy Assembly\n"
    return Response(
        content=(header + example).encode(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=users_template.csv"},
    )

@router.get("/org-setup/step5")
def org_setup_step5_get() -> dict:
    return _wizard_step5

@router.post("/org-setup/step5")
def org_setup_step5_post(payload: dict) -> dict:
    global _wizard_step5
    _wizard_step5 = payload.get("data", payload)
    return {"saved": True}


@router.get("/org-setup/step6/documents")
def org_setup_step6_documents() -> list:
    return list(_wizard_documents)


@router.post("/org-setup/step6/upload")
async def org_setup_step6_upload(
    file: UploadFile = File(...),
    name: str = Form(""),
    doc_type: str = Form("Other"),
) -> dict:
    import datetime
    global _wizard_doc_id
    _wizard_doc_id += 1
    content = await file.read()
    size_kb = round(len(content) / 1024, 1)
    size_str = f"{size_kb} KB" if size_kb < 1024 else f"{round(size_kb/1024, 1)} MB"
    doc = {
        "id": f"doc-{_wizard_doc_id}",
        "name": name or (file.filename or "unnamed"),
        "type": doc_type or "Other",
        "uploadedAt": datetime.datetime.now().isoformat(),
        "size": size_str,
    }
    _wizard_documents.append(doc)
    return doc


@router.get("/org-setup/step6a/imports")
def org_setup_step6a_imports() -> list:
    return list(_wizard_imports)


@router.post("/org-setup/step6a/import")
def org_setup_step6a_import(payload: dict) -> dict:
    import datetime
    global _wizard_import_id
    _wizard_import_id += 1
    d = payload.get("data", payload)
    imp = {
        "id": f"imp-{_wizard_import_id}",
        "dataType": d.get("dataType", ""),
        "method": d.get("method", "manual"),
        "importedAt": datetime.datetime.now().isoformat(),
        "records": int(d.get("records", 1)),
    }
    _wizard_imports.append(imp)
    return imp


# Module → (expected sheet name, insert function) mapping
_MODULE_SHEET = {
    "employees":            "Employees",
    "departments":          "Departments",
    "working_stations":     "Working_Stations",
    "roles":                "Roles",
    "policies":             "Policies",
    "permit_types":         "Permit_Types",
    "hazard_categories":    "Hazard_Categories",
    "hazards":              "Hazards",
    "training_programs":    "Training_Programs",
    "permits_to_work":      "Permits_To_Work",
    "incidents":            "Incidents",
    "near_misses":          "Near_Misses",
    "safety_walks":         "Safety_Walks",
    "capa_actions":         "CAPA_Actions",
    "shift_schedule":       "Shift_Schedule",
    # compliance_standards maps to policies table (Category/Jurisdiction = ISO/OSHA)
    "compliance_standards": "Policies",
}


@router.post("/org-setup/onboarding-bulk")
async def org_setup_onboarding_bulk(
    module: str = "",
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
) -> dict:
    import datetime
    from app.services.excel_import_service import (
        _insert_employees, _insert_departments, _insert_working_stations,
        _insert_roles, _insert_policies, _insert_permit_types,
        _insert_hazard_categories, _insert_hazards, _insert_training_programs,
        _insert_permits_to_work, _insert_incidents, _insert_near_misses,
        _insert_safety_walks, _insert_capa_actions, _insert_shift_schedule,
        capture_tenant_table_max_ids, link_new_rows_to_org,
    )
    _INSERT_FNS = {
        "employees": _insert_employees,
        "departments": _insert_departments,
        "working_stations": _insert_working_stations,
        "roles": _insert_roles,
        "policies": _insert_policies,
        "permit_types": _insert_permit_types,
        "hazard_categories": _insert_hazard_categories,
        "hazards": _insert_hazards,
        "training_programs": _insert_training_programs,
        "permits_to_work": _insert_permits_to_work,
        "incidents": _insert_incidents,
        "near_misses": _insert_near_misses,
        "safety_walks": _insert_safety_walks,
        "capa_actions": _insert_capa_actions,
        "shift_schedule": _insert_shift_schedule,
        # compliance_standards imports into policies table
        "compliance_standards": _insert_policies,
    }
    if module not in _INSERT_FNS:
        return {"count": 0, "errors": [f"Unknown module: {module!r}"]}
    if not current_user.org_id or current_user.org_id <= 0:
        return {"count": 0, "errors": ["Organisation setup is not complete. Complete setup before importing data."]}

    content = await file.read()
    fname = (file.filename or "").lower()
    sheet_name = _MODULE_SHEET[module]
    before_ids = capture_tenant_table_max_ids(db)

    try:
        import openpyxl
        if fname.endswith(".csv"):
            import csv, io as _io
            reader = csv.reader(_io.StringIO(content.decode("utf-8-sig", errors="replace")))
            all_rows = list(reader)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            for row in all_rows:
                ws.append(row)
        else:
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
            if sheet_name not in wb.sheetnames and wb.worksheets:
                wb.worksheets[0].title = sheet_name

        db.execute(text("SET FOREIGN_KEY_CHECKS=0"))
        count = _INSERT_FNS[module](db, wb)
        if current_user.org_id and current_user.org_id > 0:
            link_new_rows_to_org(db, current_user.org_id, before_ids)
        db.commit()
    except Exception as exc:
        db.rollback()
        return {"count": 0, "errors": [str(exc)]}
    finally:
        try:
            db.execute(text("SET FOREIGN_KEY_CHECKS=1"))
            db.commit()
        except Exception:
            pass

    global _wizard_import_id
    _wizard_import_id += 1
    _wizard_imports.append({
        "id": f"imp-{_wizard_import_id}",
        "dataType": module,
        "method": "bulk",
        "importedAt": datetime.datetime.now().isoformat(),
        "records": count,
    })

    label = module.replace("_", " ").title()
    import_row = DataImport(
        organisation_id=current_user.org_id,
        file_name=file.filename or f"{module}.xlsx",
        import_type="excel",
        data_type=label,
        records_total=count,
        records_success=count,
        records_failed=0,
        status="success" if count > 0 else "failed",
        uploaded_by=current_user.email or "Admin",
    )
    db.add(import_row)
    vlog = ValidationLog(
        organisation_id=current_user.org_id,
        file_name=file.filename or f"{module}.xlsx",
        rule="Required fields present",
        status="pass" if count > 0 else "fail",
        records_affected=count,
        message=f"{count} record{'s' if count != 1 else ''} imported successfully" if count > 0 else "No records imported",
        timestamp=datetime.datetime.now().isoformat(),
    )
    db.add(vlog)
    try:
        db.commit()
    except Exception:
        db.rollback()

    return {"count": count, "errors": []}


@router.post("/org-setup/fix-org-data")
def fix_org_data(db: Session = Depends(get_db), current_user: CurrentUser = Depends(get_current_user)) -> dict:
    """Global NULL-row repair is disabled to prevent seed/test data leakage."""
    return {
        "fixed": 0,
        "org_id": current_user.org_id,
        "message": "Global repair disabled. Upload data again; new imports are tenant-scoped automatically.",
    }


@router.get("/org-setup/step7")
def org_setup_step7_get() -> dict:
    return _wizard_step7

@router.post("/org-setup/step7")
def org_setup_step7_post(payload: dict) -> dict:
    global _wizard_step7
    _wizard_step7 = payload.get("data", payload)
    return {"saved": True}

@router.post("/org-setup/activate")
def org_setup_activate(request: Request, payload: Any = None, db: Session = Depends(get_db)) -> dict:
    from app.models.organisation import Organisation
    from app.models.user import User
    from app.services.auth_service import decode_access_token

    # Decode admin email from JWT
    email = ""
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header.removeprefix("Bearer ").strip()
        token_data = decode_access_token(token)
        if token_data:
            email = (token_data.get("email") or "").strip().lower()

    # Find the org that was created/identified in step 1 of this wizard session
    org = None
    if _wizard_org_id:
        org = db.query(Organisation).filter(Organisation.id == _wizard_org_id).first()
    if org is None:
        # Fallback: find the org for this admin by invite
        if email:
            invite_row = (
                db.query(OrganisationInvite)
                .filter(OrganisationInvite.admin_email == email)
                .order_by(OrganisationInvite.id.desc())
                .first()
            )
            if invite_row:
                org = db.query(Organisation).order_by(Organisation.id.desc()).first()

    if email and org:
        # Link admin user to org
        admin_user = db.query(User).filter(User.email == email).first()
        if admin_user and admin_user.organisation_id is None:
            admin_user.organisation_id = org.id

        # Mark invite accepted
        invite = (
            db.query(OrganisationInvite)
            .filter(OrganisationInvite.admin_email == email)
            .order_by(OrganisationInvite.id.desc())
            .first()
        )
        if invite:
            invite.status = "accepted"

        db.commit()

    return {"success": True}

_CSV_TEMPLATES: Dict[str, Tuple[str, str]] = {
    "employees": (
        "employee_id,full_name,date_of_birth,gender,employment_type,employment_start_date,"
        "role_id,department_id,shift_pattern,manager_id,induction_date,active_status\n"
        "EMP001,Jane Smith,1985-06-15,F,Permanent,2020-01-10,ROLE001,DEPT001,Days,,,Active\n",
        "employees_template.csv",
    ),
    "departments": (
        "department_id,site_id,department_name,manager_id,number_of_teams\n"
        "DEPT001,SITE001,Heavy Assembly,,3\n",
        "departments_template.csv",
    ),
    "working_stations": (
        "station_id,station_name,site_id,department,zone_classification,"
        "primary_hazard_id,staffing_requirement,equipment_list,permit_types_required,access_restrictions\n"
        "STN001,Assembly Line 1,SITE001,Heavy Assembly,Zone A,HAZ001,10,Welding Machine,Hot Work,PPE Required\n",
        "working_stations_template.csv",
    ),
    "roles": (
        "role_id,role_name,job_category,authority_level,permit_authority,safety_signatory\n"
        "ROLE001,Plant Manager,Management,5,Yes,Yes\n",
        "roles_template.csv",
    ),
    "policies": (
        "policy_id,policy_name,category,issue_date,owner,status\n"
        "POL001,Health & Safety Policy,Health & Safety,2024-01-01,HSE Manager,Active\n",
        "policies_template.csv",
    ),
    "permit_types": (
        "permit_type_id,permit_type_name,risk_level,validity_period_hours,concurrent_limit\n"
        "PT001,Hot Work Permit,High,8,3\n",
        "permit_types_template.csv",
    ),
    "hazard_categories": (
        "category_id,category_name,description\n"
        "HCAT001,Chemical,Hazards arising from chemical substances\n",
        "hazard_categories_template.csv",
    ),
    "hazards": (
        "hazard_id,category_id,hazard_name,severity,probability\n"
        "HAZ001,HCAT001,Acid Spill,Serious,Possible\n",
        "hazards_template.csv",
    ),
    "training_programs": (
        "program_id,training_name,duration_hours,frequency,certification,expiry_months\n"
        "TP001,Fire Safety Training,4,Annual,Yes,12\n",
        "training_programs_template.csv",
    ),
    "incidents": (
        "incident_id,report_date,incident_date_time,location_station_id,incident_type,severity,"
        "number_persons_involved,description,immediate_cause,root_cause,hazard_id,permit_active,"
        "control_failure,reported_by,investigation_status,capa_generated,days_away,root_cause_category\n",
        "incidents_template.csv",
    ),
    "near_misses": (
        "id,report_date,event_date_time,location_station_id,description,potential_consequence,"
        "hazard_id,underlying_cause,control_failure,reported_by,capa_escalation\n",
        "near_misses_template.csv",
    ),
    "safety_walks": (
        "id,inspection_date_time,location_station_id,inspector_id,inspection_type,"
        "issues_found,critical_issues,housekeeping_rating,compliance_rating,follow_up_required\n",
        "safety_walks_template.csv",
    ),
    "capa_actions": (
        "id,incident_id,action_type,description,root_cause_addressed,"
        "responsible_person_id,due_date,status,effectiveness_rating\n",
        "capa_actions_template.csv",
    ),
    "shift_schedule": (
        "id,employee_id,shift_date,shift_type,shift_start,shift_end,"
        "actual_hours_worked,station_id,supervisor_id\n",
        "shift_schedule_template.csv",
    ),
    "permits_to_work": (
        "id,permit_type_id,date_issued,time_issued,location_station_id,work_description,"
        "duration_requested_hours,issued_by,approved_by,validity_start,validity_end,"
        "work_start_actual,work_end_actual,number_of_workers,status,deviation_reported,incident_occurred\n",
        "permits_to_work_template.csv",
    ),
}


@router.get("/org-setup/template/{module}")
def org_setup_template_download(module: str):
    from fastapi.responses import Response
    tmpl = _CSV_TEMPLATES.get(module)
    if not tmpl:
        return Response(
            content=f"id,name\n".encode(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={module}_template.csv"},
        )
    content_str, filename = tmpl
    return Response(
        content=content_str.encode(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── AI Chat ───────────────────────────────────────────────────────────────────

# AI Chat is handled by app/controllers/ai.py


# ── Supervisor Endpoints ──────────────────────────────────────────────────────

@router.get("/supervisor/alerts")
def get_supervisor_alerts(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user)
) -> list:
    from datetime import datetime
    # Query incidents from DB for this organisation
    rows = db.execute(
        text("""
            SELECT i.*, ws.station_name as location_name 
            FROM incidents i 
            LEFT JOIN working_stations ws ON i.location_station_id = ws.id 
            WHERE i.organisation_id = :org_id 
            ORDER BY i.id DESC 
            LIMIT 10
        """),
        {"org_id": current_user.org_id}
    ).mappings().all()
    
    alerts = []
    for r in rows:
        severity = r["severity"] or "medium"
        alert_type = "critical" if severity in ["critical", "high"] else "warning"
        
        # Format time ago
        time_ago = "Just now"
        if r["created_at"]:
            diff = datetime.now() - r["created_at"]
            if diff.days > 0:
                time_ago = f"{diff.days}d ago"
            elif diff.seconds >= 3600:
                time_ago = f"{diff.seconds // 3600}h ago"
            elif diff.seconds >= 60:
                time_ago = f"{diff.seconds // 60}m ago"
            else:
                time_ago = "1m ago"
                
        alerts.append({
            "id": str(r["id"]),
            "type": alert_type,
            "message": f"{r['incident_type'].upper()}: {r['description']}",
            "zone": r["location_name"] or "Zone B - Sector 4",
            "time_ago": time_ago,
            "worker_name": "Worker One"
        })
        
    return alerts


@router.get("/supervisor/dashboard")
def get_supervisor_dashboard(
    current_user: CurrentUser = Depends(get_current_user)
) -> dict:
    return {
        "attendance_pct": 94,
        "safety_compliance_pct": 98,
        "active_permits": 5,
        "pending_permits": 2
    }


@router.get("/supervisor/team/shift-status")
def get_supervisor_shift_status(
    current_user: CurrentUser = Depends(get_current_user)
) -> dict:
    return {
        "total": 15,
        "logged_in": 12,
        "pending": 3,
        "is_live": True
    }


@router.get("/supervisor/permits")
def get_supervisor_permits(
    current_user: CurrentUser = Depends(get_current_user)
) -> dict:
    return {
        "items": [
            {
                "id": "1",
                "permit_ref": "PTW-2026-081",
                "permit_type": "Hot Work",
                "title": "Welding in Tank Farm 3",
                "location": "Sector 4 - Tank Farm",
                "requestor": "John Doe",
                "status": "pending",
                "risk_level": "high",
                "validity_start": "2026-07-14T08:00:00",
                "validity_end": "2026-07-14T17:00:00"
            }
        ],
        "total": 1,
        "pending_count": 1,
        "approved_today": 0,
        "risk_flags": 1
    }
