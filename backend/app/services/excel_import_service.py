"""
Parse an uploaded HSE Intelligence Excel workbook and bulk-insert every
sheet into the database in FK-safe order.

All string ID prefixes (ORG001, SITE001, EMP001, STN001 …) are stripped to
plain integers before insertion.

Every sheet's first column is that sheet's OWN row id as declared in the
Excel file (e.g. "EMP007" -> 7). Other sheets reference rows by that same
declared id (e.g. Employees.manager_id = 7 means "the employee whose own
Employees-sheet id is 7"). Declared ids are only unique *within one workbook
upload* — they are NOT the same as the table's auto-increment `id` once more
than one organisation's data lives in the same physical table. `id_maps`
tracks {table_key: {declared_id: actual_db_id}} for the current import so
every FK column can be resolved to the row this same import just inserted,
instead of accidentally colliding with another organisation's real id.
"""

import re
import logging
from io import BytesIO
from typing import Optional, List, Dict

from sqlalchemy.orm import Session
from sqlalchemy import text

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

logger = logging.getLogger(__name__)

IdMaps = Dict[str, Dict[int, int]]


# ── helpers ──────────────────────────────────────────────────────────────────

def _strip_id(val) -> Optional[int]:
    """Strip alphabetic prefix → int, or None."""
    if val is None:
        return None
    m = re.search(r"(\d+)$", str(val).strip())
    return int(m.group(1)) if m else None


def _resolve(id_maps: IdMaps, table_key: str, raw_val) -> Optional[int]:
    """Resolve a sheet-declared id (e.g. 'EMP007') to the db id this import assigned it.

    Falls back to the raw stripped id when this import's own rows don't cover it —
    that's the normal case for standalone single-sheet re-imports (e.g. re-uploading
    just Shift_Schedule against employees created in an earlier import), where the
    sheet's id column is necessarily already a real database id.
    """
    declared_id = _strip_id(raw_val)
    if declared_id is None:
        return None
    resolved = id_maps.get(table_key, {}).get(declared_id)
    if resolved is None:
        logger.debug(
            "%s reference %r (declared id %s) not inserted by this import — "
            "treating it as an existing database id",
            table_key, raw_val, declared_id,
        )
        return declared_id
    return resolved


def _remember(db: Session, id_maps: IdMaps, table_key: str, declared_id: Optional[int]) -> int:
    """Capture LAST_INSERT_ID() and, if this row declared its own id, map it."""
    new_id = int(db.execute(text("SELECT LAST_INSERT_ID()")).scalar())
    if declared_id is not None:
        id_maps.setdefault(table_key, {})[declared_id] = new_id
    return new_id


def _fmt_date(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()[:10]
    return s if s else None


def _fmt_datetime(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if len(s) == 16:
        s += ":00"
    return s if s else None


def _fmt_time(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    # Excel may return a full datetime string; take last HH:MM or HH:MM:SS
    if len(s) > 8:
        s = s[-8:]
    return s if s else None


def _rows(ws, _min_len: int = 25):
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []
    # Pad every data row to _min_len so r[i] never throws IndexError regardless
    # of how many columns the uploaded file actually has.
    return [r + (None,) * max(0, _min_len - len(r)) for r in all_rows[1:]]


def _check_sheet(wb, name: str) -> bool:
    return name in wb.sheetnames


TENANT_TABLES = (
    "sites", "departments", "employees", "incidents", "near_misses",
    "safety_walks", "capa_actions", "permits_to_work", "shift_schedule",
    "working_stations", "roles", "hazard_categories", "hazards",
    "permit_types", "training_programs", "policies",
)


def capture_tenant_table_max_ids(db: Session) -> Dict[str, int]:
    """Snapshot current ids so import linking does not claim old seed/test rows."""
    max_ids: Dict[str, int] = {}
    for tbl in TENANT_TABLES:
        try:
            max_ids[tbl] = int(db.execute(text(f"SELECT COALESCE(MAX(id), 0) FROM `{tbl}`")).scalar() or 0)
        except Exception:
            max_ids[tbl] = 0
    return max_ids


def link_new_rows_to_org(db: Session, org_id: int, before_ids: Dict[str, int]) -> int:
    """Stamp organisation_id only on rows inserted after before_ids was captured."""
    affected = 0
    for tbl in TENANT_TABLES:
        try:
            result = db.execute(
                text(
                    f"UPDATE `{tbl}` "
                    "SET organisation_id = :oid "
                    "WHERE organisation_id IS NULL AND id > :before_id"
                ),
                {"oid": org_id, "before_id": before_ids.get(tbl, 0)},
            )
            affected += result.rowcount
        except Exception:
            pass
    logger.info("Linked %s new import rows to org_id=%s", affected, org_id)
    return affected


def latest_org_id(db: Session) -> Optional[int]:
    org_id = db.execute(text("SELECT MAX(id) FROM organisation")).scalar()
    return int(org_id) if org_id else None


# ── per-table insert functions ────────────────────────────────────────────────

def _insert_organisation(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Organisation"):
        return 0
    count = 0
    for r in _rows(wb["Organisation"]):
        if not any(c for c in r):
            continue
        db.execute(
            text("""INSERT INTO organisation
                 (organisation_name,country,industry_sector,number_of_employees,
                  headquarters_location,parent_company,iso_45001_status,
                  regulatory_authority,establishment_date)
                 VALUES (:n,:c,:i,:ne,:hl,:pc,:iso,:ra,:ed)"""),
            dict(n=r[1], c=r[2], i=r[3], ne=r[4], hl=r[5],
                 pc=r[6], iso=r[7], ra=r[8], ed=_fmt_date(r[9])),
        )
        count += 1
    return count


def _insert_hazard_categories(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Hazard_Categories"):
        return 0
    count = 0
    for r in _rows(wb["Hazard_Categories"]):
        if not any(c for c in r):
            continue
        declared_id = _strip_id(r[0])
        db.execute(
            text("INSERT INTO hazard_categories (category_name,description) VALUES (:n,:d)"),
            dict(n=r[1], d=r[2]),
        )
        _remember(db, id_maps, "hazard_categories", declared_id)
        count += 1
    return count


def _insert_hazards(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Hazards"):
        return 0
    count = 0
    for r in _rows(wb["Hazards"]):
        if not any(c for c in r):
            continue
        declared_id = _strip_id(r[0])
        db.execute(
            text("INSERT INTO hazards (category_id,hazard_name,severity,probability) VALUES (:c,:n,:s,:p)"),
            dict(c=_resolve(id_maps, "hazard_categories", r[1]), n=r[2], s=r[3], p=r[4]),
        )
        _remember(db, id_maps, "hazards", declared_id)
        count += 1
    return count


def _insert_roles(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Roles"):
        return 0
    count = 0
    for r in _rows(wb["Roles"]):
        if not any(c for c in r):
            continue
        declared_id = _strip_id(r[0])
        db.execute(
            text("""INSERT INTO roles (role_name,job_category,authority_level,
                 permit_authority,safety_signatory) VALUES (:n,:jc,:al,:pa,:ss)"""),
            dict(n=r[1], jc=r[2], al=r[3], pa=r[4], ss=r[5]),
        )
        _remember(db, id_maps, "roles", declared_id)
        count += 1
    return count


def _insert_sites(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Sites"):
        return 0
    count = 0
    for r in _rows(wb["Sites"]):
        if not any(c for c in r):
            continue
        declared_id = _strip_id(r[0])
        db.execute(
            text("""INSERT INTO sites
                 (site_name,address,postcode,city,type,operational_status,
                  number_of_working_stations,capacity,primary_products,hazard_classification)
                 VALUES (:sn,:a,:p,:c,:t,:os,:nws,:cap,:pp,:hc)"""),
            dict(sn=r[1], a=r[2], p=r[3], c=r[4], t=r[5], os=r[6],
                 nws=r[7], cap=r[8], pp=r[9], hc=r[10]),
        )
        _remember(db, id_maps, "sites", declared_id)
        count += 1
    return count


def _link_org_to_data(db: Session, wb) -> int:
    """Deprecated no-op.

    Older code globally stamped every NULL organisation_id row, which could
    attach seed/test data to the first tenant that opened setup/dashboard.
    Use link_new_rows_to_org() with a captured pre-import id snapshot instead.
    """
    return 0


def _insert_permit_types(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Permit_Types"):
        return 0
    count = 0
    for r in _rows(wb["Permit_Types"]):
        if not any(c for c in r):
            continue
        declared_id = _strip_id(r[0])
        db.execute(
            text("""INSERT INTO permit_types
                 (permit_type_name,risk_level,validity_period_hours,concurrent_limit)
                 VALUES (:n,:rl,:vph,:cl)"""),
            dict(n=r[1], rl=r[2], vph=r[3], cl=r[4]),
        )
        _remember(db, id_maps, "permit_types", declared_id)
        count += 1
    return count


def _insert_training_programs(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Training_Programs"):
        return 0
    count = 0
    for r in _rows(wb["Training_Programs"]):
        if not any(c for c in r):
            continue
        db.execute(
            text("""INSERT INTO training_programs
                 (training_name,duration_hours,frequency,certification,expiry_months)
                 VALUES (:n,:dh,:f,:cert,:em)"""),
            dict(n=r[1], dh=r[2], f=r[3], cert=r[4], em=r[5]),
        )
        count += 1
    return count


def _insert_policies(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Policies"):
        return 0
    count = 0
    for r in _rows(wb["Policies"]):
        if not any(c for c in r):
            continue
        db.execute(
            text("INSERT INTO policies (policy_name,category,issue_date,owner,status) VALUES (:n,:c,:id,:o,:s)"),
            dict(n=r[1], c=r[2], id=_fmt_date(r[3]), o=r[4], s=r[5]),
        )
        count += 1
    return count


def _insert_departments(db: Session, wb, id_maps: IdMaps) -> int:
    """Insert departments without manager_id (circular FK with employees)."""
    if not _check_sheet(wb, "Departments"):
        return 0
    count = 0
    for r in _rows(wb["Departments"]):
        if not any(c for c in r):
            continue
        declared_id = _strip_id(r[0])
        db.execute(
            text("INSERT INTO departments (site_id,department_name,number_of_teams) VALUES (:sid,:n,:nt)"),
            dict(sid=_resolve(id_maps, "sites", r[1]), n=r[2], nt=r[4]),
        )
        _remember(db, id_maps, "departments", declared_id)
        count += 1
    return count


def _insert_working_stations(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Working_Stations"):
        return 0
    count = 0
    for r in _rows(wb["Working_Stations"]):
        if not any(c for c in r):
            continue
        declared_id = _strip_id(r[0])
        db.execute(
            text("""INSERT INTO working_stations
                 (station_name,site_id,department,zone_classification,
                  primary_hazard_id,staffing_requirement,equipment_list,
                  permit_types_required,access_restrictions)
                 VALUES (:sn,:sid,:d,:zc,:phi,:sr,:el,:ptr,:ar)"""),
            dict(sn=r[1], sid=_resolve(id_maps, "sites", r[2]), d=r[3], zc=r[4],
                 phi=_resolve(id_maps, "hazards", r[5]), sr=r[6], el=r[7], ptr=r[8], ar=r[9]),
        )
        _remember(db, id_maps, "working_stations", declared_id)
        count += 1
    return count


def _insert_employees(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Employees"):
        return 0
    rows = [r for r in _rows(wb["Employees"]) if any(c for c in r)]

    # Pass 1: insert every employee (manager_id deferred — it may forward-reference
    # a row later in this same sheet), building the declared-id -> db-id map.
    pending_manager_updates: List[tuple] = []  # (this row's db_id, manager's declared id)
    for r in rows:
        declared_id = _strip_id(r[0])
        db.execute(
            text("""INSERT INTO employees
                 (full_name,date_of_birth,gender,employment_type,employment_start_date,
                  role_id,department_id,shift_pattern,induction_date,active_status)
                 VALUES (:fn,:dob,:g,:et,:esd,:rid,:did,:sp,:ind,:as_)"""),
            dict(fn=r[1], dob=_fmt_date(r[2]), g=r[3], et=r[4],
                 esd=_fmt_date(r[5]),
                 rid=_resolve(id_maps, "roles", r[6]),
                 did=_resolve(id_maps, "departments", r[7]),
                 sp=r[8],
                 ind=_fmt_date(r[10]),
                 as_=r[11]),
        )
        new_id = _remember(db, id_maps, "employees", declared_id)
        mgr_declared_id = _strip_id(r[9]) if r[9] else None
        if mgr_declared_id is not None:
            pending_manager_updates.append((new_id, mgr_declared_id))

    # Pass 2: now every employee in this sheet has a db id — resolve manager_id.
    # A manager declared in *this* sheet resolves via id_maps; a manager from an
    # earlier import (not part of this sheet) falls back to being treated as an
    # existing database id, same as _resolve() does for cross-table references.
    for emp_db_id, mgr_declared_id in pending_manager_updates:
        mgr_db_id = id_maps.get("employees", {}).get(mgr_declared_id, mgr_declared_id)
        db.execute(
            text("UPDATE employees SET manager_id = :mid WHERE id = :eid"),
            dict(mid=mgr_db_id, eid=emp_db_id),
        )

    return len(rows)


def _update_department_managers(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Departments"):
        return 0
    count = 0
    for r in _rows(wb["Departments"]):
        if not any(c for c in r):
            continue
        dept_id = _resolve(id_maps, "departments", r[0]) if r[0] is not None else None
        mgr_id = _resolve(id_maps, "employees", r[3]) if r[3] else None
        if dept_id and mgr_id:
            db.execute(
                text("UPDATE departments SET manager_id = :mid WHERE id = :did"),
                dict(mid=mgr_id, did=dept_id),
            )
            count += 1
    return count


def _insert_permits_to_work(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Permits_To_Work"):
        return 0
    count = 0
    for r in _rows(wb["Permits_To_Work"]):
        if not any(c for c in r):
            continue
        db.execute(
            text("""INSERT INTO permits_to_work
                 (permit_type_id,date_issued,time_issued,location_station_id,
                  work_description,duration_requested_hours,issued_by,approved_by,
                  validity_start,validity_end,work_start_actual,work_end_actual,
                  number_of_workers,status,deviation_reported,incident_occurred)
                 VALUES (:ptid,:di,:ti,:lsid,:wd,:drh,:ib,:ab,:vs,:ve,:wsa,:wea,:nw,:st,:dr,:io)"""),
            dict(ptid=_resolve(id_maps, "permit_types", r[1]), di=_fmt_date(r[2]), ti=_fmt_time(r[3]),
                 lsid=_resolve(id_maps, "working_stations", r[4]), wd=r[5], drh=r[6],
                 ib=_resolve(id_maps, "employees", r[7]), ab=_resolve(id_maps, "employees", r[8]),
                 vs=_fmt_datetime(r[9]), ve=_fmt_datetime(r[10]),
                 wsa=_fmt_datetime(r[11]), wea=_fmt_datetime(r[12]),
                 nw=r[13], st=r[14], dr=r[15], io=r[16]),
        )
        count += 1
    return count


def _insert_incidents(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Incidents"):
        return 0
    count = 0
    for r in _rows(wb["Incidents"]):
        if not any(c for c in r):
            continue
        declared_id = _strip_id(r[0])
        db.execute(
            text("""INSERT INTO incidents
                 (report_date,incident_date_time,location_station_id,incident_type,
                  severity,number_persons_involved,description,immediate_cause,
                  root_cause,hazard_id,permit_active,control_failure,reported_by,
                  investigation_status,capa_generated,days_away,root_cause_category)
                 VALUES (:rd,:idt,:lsid,:it,:sev,:npi,:desc,:ic,:rc,:hid,:pa,:cf,:rb,:is_,:cg,:da,:rcc)"""),
            dict(rd=_fmt_date(r[1]), idt=_fmt_datetime(r[2]),
                 lsid=_resolve(id_maps, "working_stations", r[3]), it=r[4], sev=r[5], npi=r[6],
                 desc=r[7], ic=r[8], rc=r[9], hid=_resolve(id_maps, "hazards", r[10]),
                 pa=r[11], cf=r[12], rb=_resolve(id_maps, "employees", r[13]),
                 is_=r[14], cg=r[15], da=r[16] or 0, rcc=r[17]),
        )
        _remember(db, id_maps, "incidents", declared_id)
        count += 1
    return count


def _insert_near_misses(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Near_Misses"):
        return 0
    count = 0
    for r in _rows(wb["Near_Misses"]):
        if not any(c for c in r):
            continue
        db.execute(
            text("""INSERT INTO near_misses
                 (report_date,event_date_time,location_station_id,description,
                  potential_consequence,hazard_id,underlying_cause,control_failure,
                  reported_by,capa_escalation)
                 VALUES (:rd,:edt,:lsid,:desc,:pc,:hid,:uc,:cf,:rb,:ce)"""),
            dict(rd=_fmt_date(r[1]), edt=_fmt_datetime(r[2]),
                 lsid=_resolve(id_maps, "working_stations", r[3]), desc=r[4], pc=r[5],
                 hid=_resolve(id_maps, "hazards", r[6]), uc=r[7], cf=r[8],
                 rb=_resolve(id_maps, "employees", r[9]), ce=r[10]),
        )
        count += 1
    return count


def _insert_safety_walks(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Safety_Walks"):
        return 0
    count = 0
    for r in _rows(wb["Safety_Walks"]):
        if not any(c for c in r):
            continue
        db.execute(
            text("""INSERT INTO safety_walks
                 (inspection_date_time,location_station_id,inspector_id,
                  inspection_type,issues_found,critical_issues,
                  housekeeping_rating,compliance_rating,follow_up_required)
                 VALUES (:idt,:lsid,:iid,:it,:if_,:ci,:hr,:cr,:fur)"""),
            dict(idt=_fmt_datetime(r[1]), lsid=_resolve(id_maps, "working_stations", r[2]),
                 iid=_resolve(id_maps, "employees", r[3]), it=r[4], if_=r[5] or 0,
                 ci=r[6] or 0, hr=r[7], cr=r[8], fur=r[9]),
        )
        count += 1
    return count


def _insert_capa_actions(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "CAPA_Actions"):
        return 0
    count = 0
    for r in _rows(wb["CAPA_Actions"]):
        if not any(c for c in r):
            continue
        db.execute(
            text("""INSERT INTO capa_actions
                 (incident_id,action_type,description,root_cause_addressed,
                  responsible_person_id,due_date,status,effectiveness_rating)
                 VALUES (:iid,:at,:desc,:rca,:rpid,:dd,:st,:er)"""),
            dict(iid=_resolve(id_maps, "incidents", r[1]), at=r[2], desc=r[3], rca=r[4],
                 rpid=_resolve(id_maps, "employees", r[5]), dd=_fmt_date(r[6]),
                 st=r[7], er=r[8]),
        )
        count += 1
    return count


def _insert_shift_schedule(db: Session, wb, id_maps: IdMaps) -> int:
    if not _check_sheet(wb, "Shift_Schedule"):
        return 0
    count = 0
    for r in _rows(wb["Shift_Schedule"]):
        if not any(c for c in r):
            continue
        db.execute(
            text("""INSERT INTO shift_schedule
                 (employee_id,shift_date,shift_type,shift_start,shift_end,
                  actual_hours_worked,station_id,supervisor_id)
                 VALUES (:eid,:sd,:st,:ss,:se,:ahw,:stid,:supid)"""),
            dict(eid=_resolve(id_maps, "employees", r[1]), sd=_fmt_date(r[2]),
                 st=r[3], ss=_fmt_time(r[4]), se=_fmt_time(r[5]),
                 ahw=r[6], stid=_resolve(id_maps, "working_stations", r[7]),
                 supid=_resolve(id_maps, "employees", r[8])),
        )
        count += 1
    return count


# ── public entry point ────────────────────────────────────────────────────────

SHEET_STEPS = [
    ("Organisation",          "organisation",       _insert_organisation),
    ("Hazard_Categories",     "hazard_categories",  _insert_hazard_categories),
    ("Hazards",               "hazards",            _insert_hazards),
    ("Roles",                 "roles",              _insert_roles),
    ("Sites",                 "sites",              _insert_sites),
    ("Permit_Types",          "permit_types",       _insert_permit_types),
    ("Training_Programs",     "training_programs",  _insert_training_programs),
    ("Policies",              "policies",           _insert_policies),
    ("Departments",           "departments",        _insert_departments),
    ("Working_Stations",      "working_stations",   _insert_working_stations),
    ("Employees",             "employees",          _insert_employees),
    ("Departments (managers)","dept_managers",      _update_department_managers),
    ("Permits_To_Work",       "permits_to_work",    _insert_permits_to_work),
    ("Incidents",             "incidents",          _insert_incidents),
    ("Near_Misses",           "near_misses",        _insert_near_misses),
    ("Safety_Walks",          "safety_walks",       _insert_safety_walks),
    ("CAPA_Actions",          "capa_actions",       _insert_capa_actions),
    ("Shift_Schedule",        "shift_schedule",     _insert_shift_schedule),
]


def import_excel_stream(file_bytes: bytes, db: Session):
    """Generator version — yields SSE-formatted strings as each sheet is processed.

    Each sheet commits independently so a failure in one sheet cannot roll back
    earlier sheets. FK checks are disabled for the duration of the import so that
    cross-sheet references work regardless of insertion order.
    """
    if openpyxl is None:
        yield _sse({"type": "fatal", "error": "openpyxl is not installed on the server"})
        return

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        yield _sse({"type": "fatal", "error": f"Cannot open Excel file: {exc}"})
        return

    yield _sse({"type": "start", "total": len(SHEET_STEPS)})

    results: Dict[str, int] = {}
    errors: Dict[str, str] = {}
    before_ids = capture_tenant_table_max_ids(db)
    id_maps: IdMaps = {}

    try:
        for idx, (sheet_label, table_key, fn) in enumerate(SHEET_STEPS):
            yield _sse({"type": "processing", "index": idx, "key": table_key, "label": sheet_label})
            try:
                # Must SET within the same transaction as the INSERTs; committing
                # after SET returns the connection to the pool, resetting the flag.
                db.execute(text("SET FOREIGN_KEY_CHECKS=0"))
                count = fn(db, wb, id_maps)
                db.commit()
                results[table_key] = count
                logger.info("Imported %s rows into %s", count, table_key)
                yield _sse({"type": "done", "index": idx, "key": table_key, "label": sheet_label, "count": count})
            except Exception as exc:
                msg = str(exc)
                logger.error("Error importing %s: %s", sheet_label, exc, exc_info=True)
                errors[table_key] = msg
                db.rollback()
                yield _sse({"type": "error", "index": idx, "key": table_key, "label": sheet_label, "error": msg})
    finally:
        try:
            org_id = latest_org_id(db)
            if org_id:
                results["org_link"] = link_new_rows_to_org(db, org_id, before_ids)
            db.execute(text("SET FOREIGN_KEY_CHECKS=1"))
            db.commit()
        except Exception:
            pass

    wb.close()
    yield _sse({
        "type": "complete",
        "results": results,
        "errors": errors,
        "total_rows": sum(results.values()),
        "has_errors": bool(errors),
    })


def _sse(data: dict) -> str:
    import json
    return f"data: {json.dumps(data)}\n\n"


def import_excel(file_bytes: bytes, db: Session) -> dict:
    """Parse the uploaded Excel workbook and insert all sheets into the DB.

    Returns a dict of {table_key: row_count} for each processed sheet.
    Each sheet commits independently; FK checks are disabled for the import.
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot open Excel file: {exc}") from exc

    results: Dict[str, int] = {}
    errors: Dict[str, str] = {}
    before_ids = capture_tenant_table_max_ids(db)
    id_maps: IdMaps = {}

    try:
        for sheet_label, table_key, fn in SHEET_STEPS:
            try:
                db.execute(text("SET FOREIGN_KEY_CHECKS=0"))
                count = fn(db, wb, id_maps)
                db.commit()
                results[table_key] = count
                logger.info("Imported %s rows into %s", count, table_key)
            except Exception as exc:
                logger.error("Error importing %s: %s", sheet_label, exc, exc_info=True)
                errors[table_key] = str(exc)
                db.rollback()
    finally:
        try:
            org_id = latest_org_id(db)
            if org_id:
                results["org_link"] = link_new_rows_to_org(db, org_id, before_ids)
            db.execute(text("SET FOREIGN_KEY_CHECKS=1"))
            db.commit()
        except Exception:
            pass

    wb.close()
    if errors:
        logger.warning("Import completed with errors in sheets: %s", list(errors.keys()))
    return results
