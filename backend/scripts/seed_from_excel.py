"""
Truncates all business-data tables (preserving users / app_roles / _migrations)
and reseeds the database from an HSE Intelligence test-data Excel workbook.

Usage (run from the backend/ directory so the `app` package resolves):
    python scripts/seed_from_excel.py [path_to_xlsx]

If no path is given, defaults to the workbook at the repo root.
"""
import os
import sys
from datetime import date, datetime
from datetime import time as dtime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl  # noqa: E402

from app.config.database import engine  # noqa: E402

DEFAULT_EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "HSE_Intelligence_Test_Data.xlsx",
)
EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL_PATH

# Tables to wipe, in no particular order (FK checks are disabled for the whole run).
TABLES_TO_TRUNCATE = [
    "capa_actions", "checklist_logs", "checklist_submission_items", "checklist_submissions",
    "checklist_templates", "departments", "employees", "hazard_categories", "hazards",
    "incidents", "near_misses", "organisation", "organisation_invite", "permit_types",
    "permits_to_work", "policies", "roles", "safety_walks", "shift_schedule", "sites",
    "training_programs", "working_stations",
]
PRESERVED_TABLES = ["users", "app_roles", "_migrations"]


# ── Parsing helpers ──────────────────────────────────────────────────────────

def s(v):
    """Trim a string cell; blank/None becomes None."""
    if v is None:
        return None
    v = str(v).strip()
    return v or None


def parse_date(v):
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v).strip(), "%Y-%m-%d").date()


def parse_datetime(v):
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v
    txt = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(txt, fmt)
        except ValueError:
            continue
    d = parse_date(txt)
    return datetime.combine(d, dtime.min) if d else None


def parse_time(v):
    if v in (None, ""):
        return None
    if isinstance(v, dtime):
        return v
    txt = str(v).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(txt, fmt).time()
        except ValueError:
            continue
    return None


def yesno(v):
    """Normalize any non-empty value to 'Yes', else None (for enum('Yes','No') columns)."""
    return "Yes" if v not in (None, "") else None


def sheet_rows(wb, name):
    ws = wb[name]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    out = []
    for row in rows_iter:
        if row[0] is None:
            continue
        out.append(dict(zip(header, row)))
    return out


def code_map(rows, id_field):
    """Map each row's Excel ID string -> its 1-based insertion-order integer id."""
    return {r[id_field]: i for i, r in enumerate(rows, start=1)}


def chunked(seq, size=5000):
    for i in range(0, len(seq), size):
        yield seq[i:i + size]


def bulk_insert(cur, sql, params):
    for batch in chunked(params):
        cur.executemany(sql, batch)


def main():
    print(f"Loading workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    data = {name: sheet_rows(wb, name) for name in wb.sheetnames}
    for name, rows in data.items():
        print(f"  {name}: {len(rows)} rows")

    conn = engine.raw_connection()
    cur = conn.cursor()

    try:
        cur.execute("SET FOREIGN_KEY_CHECKS=0")

        print("\nTruncating tables (preserving: " + ", ".join(PRESERVED_TABLES) + ")...")
        for t in TABLES_TO_TRUNCATE:
            cur.execute(f"TRUNCATE TABLE {t}")
        conn.commit()

        # ── Organisation ─────────────────────────────────────────────────
        rows = data["Organisation"]
        bulk_insert(cur, """
            INSERT INTO organisation
                (organisation_name, country, industry_sector, number_of_employees,
                 headquarters_location, parent_company, iso_45001_status,
                 regulatory_authority, establishment_date)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, [(
            s(r["Organisation_Name"]), s(r["Country"]), s(r["Industry_Sector"]),
            r["Number_of_Employees"], s(r["Headquarters_Location"]), s(r["Parent_Company"]),
            s(r["ISO_45001_Status"]), s(r["Regulatory_Authority"]), parse_date(r["Establishment_Date"]),
        ) for r in rows])
        conn.commit()
        print(f"organisation: inserted {len(rows)}")

        # ── Sites ────────────────────────────────────────────────────────
        rows = data["Sites"]
        site_id = code_map(rows, "Site_ID")
        bulk_insert(cur, """
            INSERT INTO sites
                (site_name, address, postcode, city, type, operational_status,
                 number_of_working_stations, capacity, primary_products, hazard_classification)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, [(
            s(r["Site Name"]), s(r["Address"]), s(r["Postcode"]), s(r["City"]), s(r["Type"]),
            s(r["Operational_Status"]), r["Number_of_Working_Stations"], r["Capacity"],
            s(r["Primary_Products"]), s(r["Hazard_Classification"]),
        ) for r in rows])
        conn.commit()
        print(f"sites: inserted {len(rows)}")

        # ── Roles ────────────────────────────────────────────────────────
        rows = data["Roles"]
        role_id = code_map(rows, "Role_ID")
        bulk_insert(cur, """
            INSERT INTO roles (role_name, job_category, authority_level, permit_authority, safety_signatory)
            VALUES (%s,%s,%s,%s,%s)
        """, [(
            s(r["Role_Name"]), s(r["Job_Category"]), r["Authority_Level"],
            s(r["Permit_Authority"]), s(r["Safety_Signatory"]),
        ) for r in rows])
        conn.commit()
        print(f"roles: inserted {len(rows)}")

        # ── Hazard_Categories ────────────────────────────────────────────
        rows = data["Hazard_Categories"]
        hazcat_id = code_map(rows, "Hazard_Category_ID")
        bulk_insert(cur, "INSERT INTO hazard_categories (category_name, description) VALUES (%s,%s)",
                    [(s(r["Category_Name"]), s(r["Description"])) for r in rows])
        conn.commit()
        print(f"hazard_categories: inserted {len(rows)}")

        # ── Hazards ──────────────────────────────────────────────────────
        rows = data["Hazards"]
        hazard_id = code_map(rows, "Hazard_ID")
        bulk_insert(cur, """
            INSERT INTO hazards (category_id, hazard_name, severity, probability)
            VALUES (%s,%s,%s,%s)
        """, [(
            hazcat_id.get(r["Category_ID"]), s(r["Hazard_Name"]), s(r["Severity"]), s(r["Probability"]),
        ) for r in rows])
        conn.commit()
        print(f"hazards: inserted {len(rows)}")

        # ── Working_Stations (needs sites, hazards) ─────────────────────
        rows = data["Working_Stations"]
        station_id = code_map(rows, "Station_ID")
        bulk_insert(cur, """
            INSERT INTO working_stations
                (station_name, site_id, department, zone_classification, primary_hazard_id,
                 staffing_requirement, equipment_list, permit_types_required, access_restrictions)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, [(
            s(r["Station_Name"]), site_id.get(r["Site_ID"]), s(r["Department"]),
            s(r["Zone_Classification"]), hazard_id.get(r["Primary_Hazards"]),
            r["Staffing_Requirement"], s(r["Equipment_List"]), s(r["Permit_Types_Required"]),
            s(r["Access_Restrictions"]),
        ) for r in rows])
        conn.commit()
        print(f"working_stations: inserted {len(rows)}")

        # ── Departments (Site_ID / Manager_ID are plain ints = row position
        #    in Sites / Employees; FK checks are off so insertion order vs.
        #    Employees doesn't matter here) ────────────────────────────────
        rows = data["Departments"]
        dept_id = {f"DEPT{int(r['Department_ID']):03d}": i for i, r in enumerate(rows, start=1)}
        bulk_insert(cur, """
            INSERT INTO departments (site_id, department_name, manager_id, number_of_teams)
            VALUES (%s,%s,%s,%s)
        """, [(
            int(r["Site_ID"]), s(r["Department_Name"]),
            int(r["Manager_ID"]) if r["Manager_ID"] is not None else None,
            r["Number_of_Teams"],
        ) for r in rows])
        conn.commit()
        print(f"departments: inserted {len(rows)}")

        # ── Employees (Job Title = Role code, Department = DEPT code,
        #    Manager_ID = EMP code or blank) ──────────────────────────────
        rows = data["Employees"]
        employee_id = code_map(rows, "Employee ID")
        bulk_insert(cur, """
            INSERT INTO employees
                (full_name, date_of_birth, gender, employment_type, employment_start_date,
                 role_id, department_id, shift_pattern, manager_id, induction_date, active_status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, [(
            s(r["Full Name"]), parse_date(r["Date_of_Birth"]), s(r["Gender"]),
            s(r["Employment_Type"]), parse_date(r["Employment_Start_Date"]),
            role_id.get(r["Job Title"]), dept_id.get(r["Department"]), s(r["Shift_Pattern"]),
            employee_id.get(r["Manager_ID"]), parse_date(r["Induction_Date"]), s(r["Active_Status"]),
        ) for r in rows])
        conn.commit()
        print(f"employees: inserted {len(rows)}")

        # ── Policies (standalone) ───────────────────────────────────────
        rows = data["Policies"]
        bulk_insert(cur, """
            INSERT INTO policies (policy_name, category, issue_date, owner, status)
            VALUES (%s,%s,%s,%s,%s)
        """, [(
            s(r["Policy_Name"]), s(r["Category"]), parse_date(r["Issue_Date"]), s(r["Owner"]), s(r["Status"]),
        ) for r in rows])
        conn.commit()
        print(f"policies: inserted {len(rows)}")

        # ── Permit_Types (standalone) ───────────────────────────────────
        rows = data["Permit_Types"]
        permit_type_id = code_map(rows, "Permit_Type_ID")
        bulk_insert(cur, """
            INSERT INTO permit_types (permit_type_name, risk_level, validity_period_hours, concurrent_limit)
            VALUES (%s,%s,%s,%s)
        """, [(
            s(r["Permit_Type_Name"]), s(r["Risk_Level"]), r["Validity_Period_Hours"], r["Concurrent_Limit"],
        ) for r in rows])
        conn.commit()
        print(f"permit_types: inserted {len(rows)}")

        # ── Training_Programs (standalone) ──────────────────────────────
        rows = data["Training_Programs"]
        bulk_insert(cur, """
            INSERT INTO training_programs (training_name, duration_hours, frequency, certification, expiry_months)
            VALUES (%s,%s,%s,%s,%s)
        """, [(
            s(r["Training_Name"]), r["Duration_Hours"], s(r["Frequency"]), s(r["Certification"]), r["Expiry_Months"],
        ) for r in rows])
        conn.commit()
        print(f"training_programs: inserted {len(rows)}")

        # ── Permits_To_Work (needs permit_types, working_stations, employees) ──
        rows = data["Permits_To_Work"]
        bulk_insert(cur, """
            INSERT INTO permits_to_work
                (permit_type_id, date_issued, time_issued, location_station_id, work_description,
                 duration_requested_hours, issued_by, approved_by, validity_start, validity_end,
                 work_start_actual, work_end_actual, number_of_workers, status,
                 deviation_reported, incident_occurred)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, [(
            permit_type_id.get(r["Permit_Type_ID"]), parse_date(r["Date_Issued"]), parse_time(r["Time_Issued"]),
            station_id.get(r["Location_Station_ID"]), s(r["Work_Description"]), r["Duration_Requested_Hours"],
            employee_id.get(r["Issued_By"]), employee_id.get(r["Approved_By"]),
            parse_datetime(r["Validity_Start"]), parse_datetime(r["Validity_End"]),
            parse_datetime(r["Work_Start_Actual"]), parse_datetime(r["Work_End_Actual"]),
            r["Number_of_Workers"], s(r["Status"]), s(r["Deviation_Reported"]), s(r["Incident_Occurred"]),
        ) for r in rows])
        conn.commit()
        print(f"permits_to_work: inserted {len(rows)}")

        # ── Incidents (needs working_stations, hazards, employees) ─────
        # NOTE: Permit_Active in the source data holds a PTW reference code,
        # not Yes/No, but the DB column is enum('Yes','No') — collapsed to
        # 'Yes' when a permit code is present, else NULL (per confirmed scope).
        rows = data["Incidents"]
        incident_id = code_map(rows, "Incident_ID")
        bulk_insert(cur, """
            INSERT INTO incidents
                (report_date, incident_date_time, location_station_id, incident_type, severity,
                 number_persons_involved, description, immediate_cause, root_cause, hazard_id,
                 permit_active, control_failure, reported_by, investigation_status, capa_generated,
                 days_away, root_cause_category)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, [(
            parse_date(r["Report_Date"]), parse_datetime(r["Incident_DateTime"]),
            station_id.get(r["Location_Station"]), s(r["Incident_Type"]), s(r["Severity"]),
            r["Number_Persons_Involved"], s(r["Description"]), s(r["Immediate_Cause"]), s(r["Root_Cause"]),
            hazard_id.get(r["Hazard_Involved"]), yesno(r["Permit_Active"]), s(r["Control_Failure"]),
            employee_id.get(r["Reported_By"]), s(r["Investigation_Status"]), s(r["CAPA_Generated"]),
            r["Days_Away"], s(r["Root_Cause_Category"]),
        ) for r in rows])
        conn.commit()
        print(f"incidents: inserted {len(rows)}")

        # ── Near_Misses (needs working_stations, hazards, employees) ───
        rows = data["Near_Misses"]
        bulk_insert(cur, """
            INSERT INTO near_misses
                (report_date, event_date_time, location_station_id, description, potential_consequence,
                 hazard_id, underlying_cause, control_failure, reported_by, capa_escalation)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, [(
            parse_date(r["Report_Date"]), parse_datetime(r["Event_DateTime"]),
            station_id.get(r["Location_Station"]), s(r["Description"]), s(r["Potential_Consequence"]),
            hazard_id.get(r["Hazard_Involved"]), s(r["Underlying_Cause"]), s(r["Control_Failure"]),
            employee_id.get(r["Reported_By"]), s(r["CAPA_Escalation"]),
        ) for r in rows])
        conn.commit()
        print(f"near_misses: inserted {len(rows)}")

        # ── Safety_Walks (needs working_stations, employees) ───────────
        rows = data["Safety_Walks"]
        bulk_insert(cur, """
            INSERT INTO safety_walks
                (inspection_date_time, location_station_id, inspector_id, inspection_type,
                 issues_found, critical_issues, housekeeping_rating, compliance_rating, follow_up_required)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, [(
            parse_datetime(r["Inspection_DateTime"]), station_id.get(r["Location_Station"]),
            employee_id.get(r["Inspector"]), s(r["Inspection_Type"]), r["Issues_Found"],
            r["Critical_Issues"], r["Housekeeping_Rating"], r["Compliance_Rating"],
            s(r["Follow_Up_Required"]),
        ) for r in rows])
        conn.commit()
        print(f"safety_walks: inserted {len(rows)}")

        # ── CAPA_Actions (needs incidents, employees) ───────────────────
        rows = data["CAPA_Actions"]
        bulk_insert(cur, """
            INSERT INTO capa_actions
                (incident_id, action_type, description, root_cause_addressed,
                 responsible_person_id, due_date, status, effectiveness_rating)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, [(
            incident_id.get(r["Incident_ID"]), s(r["Action_Type"]), s(r["Description"]),
            s(r["Root_Cause_Addressed"]), employee_id.get(r["Responsible_Person"]),
            parse_date(r["Due_Date"]), s(r["Status"]), r["Effectiveness_Rating"],
        ) for r in rows])
        conn.commit()
        print(f"capa_actions: inserted {len(rows)}")

        # ── Shift_Schedule (needs employees, working_stations) ──────────
        rows = data["Shift_Schedule"]
        bulk_insert(cur, """
            INSERT INTO shift_schedule
                (employee_id, shift_date, shift_type, shift_start, shift_end,
                 actual_hours_worked, station_id, supervisor_id)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, [(
            employee_id.get(r["Employee_ID"]), parse_date(r["Shift_Date"]), s(r["Shift_Type"]),
            parse_time(r["Shift_Start"]), parse_time(r["Shift_End"]), r["Actual_Hours_Worked"],
            station_id.get(r["Station_Assigned"]), employee_id.get(r["Supervisor"]),
        ) for r in rows])
        conn.commit()
        print(f"shift_schedule: inserted {len(rows)}")

        cur.execute("SET FOREIGN_KEY_CHECKS=1")
        conn.commit()

        # ── Post-load sanity checks ──────────────────────────────────────
        print("\nValidating row counts...")
        for table in TABLES_TO_TRUNCATE:
            cur.execute(f"SELECT COUNT(*) FROM {table}")
            print(f"  {table}: {cur.fetchone()[0]}")

        print("\nChecking for orphaned foreign keys...")
        orphan_checks = [
            ("working_stations.site_id", "SELECT COUNT(*) FROM working_stations WHERE site_id IS NOT NULL AND site_id NOT IN (SELECT id FROM sites)"),
            ("working_stations.primary_hazard_id", "SELECT COUNT(*) FROM working_stations WHERE primary_hazard_id IS NOT NULL AND primary_hazard_id NOT IN (SELECT id FROM hazards)"),
            ("departments.site_id", "SELECT COUNT(*) FROM departments WHERE site_id IS NOT NULL AND site_id NOT IN (SELECT id FROM sites)"),
            ("departments.manager_id", "SELECT COUNT(*) FROM departments WHERE manager_id IS NOT NULL AND manager_id NOT IN (SELECT id FROM employees)"),
            ("employees.role_id", "SELECT COUNT(*) FROM employees WHERE role_id IS NOT NULL AND role_id NOT IN (SELECT id FROM roles)"),
            ("employees.department_id", "SELECT COUNT(*) FROM employees WHERE department_id IS NOT NULL AND department_id NOT IN (SELECT id FROM departments)"),
            ("employees.manager_id", "SELECT COUNT(*) FROM employees WHERE manager_id IS NOT NULL AND manager_id NOT IN (SELECT id FROM employees)"),
            ("hazards.category_id", "SELECT COUNT(*) FROM hazards WHERE category_id IS NOT NULL AND category_id NOT IN (SELECT id FROM hazard_categories)"),
            ("permits_to_work.permit_type_id", "SELECT COUNT(*) FROM permits_to_work WHERE permit_type_id IS NOT NULL AND permit_type_id NOT IN (SELECT id FROM permit_types)"),
            ("permits_to_work.location_station_id", "SELECT COUNT(*) FROM permits_to_work WHERE location_station_id IS NOT NULL AND location_station_id NOT IN (SELECT id FROM working_stations)"),
            ("permits_to_work.issued_by", "SELECT COUNT(*) FROM permits_to_work WHERE issued_by IS NOT NULL AND issued_by NOT IN (SELECT id FROM employees)"),
            ("permits_to_work.approved_by", "SELECT COUNT(*) FROM permits_to_work WHERE approved_by IS NOT NULL AND approved_by NOT IN (SELECT id FROM employees)"),
            ("incidents.location_station_id", "SELECT COUNT(*) FROM incidents WHERE location_station_id IS NOT NULL AND location_station_id NOT IN (SELECT id FROM working_stations)"),
            ("incidents.hazard_id", "SELECT COUNT(*) FROM incidents WHERE hazard_id IS NOT NULL AND hazard_id NOT IN (SELECT id FROM hazards)"),
            ("incidents.reported_by", "SELECT COUNT(*) FROM incidents WHERE reported_by IS NOT NULL AND reported_by NOT IN (SELECT id FROM employees)"),
            ("near_misses.location_station_id", "SELECT COUNT(*) FROM near_misses WHERE location_station_id IS NOT NULL AND location_station_id NOT IN (SELECT id FROM working_stations)"),
            ("near_misses.hazard_id", "SELECT COUNT(*) FROM near_misses WHERE hazard_id IS NOT NULL AND hazard_id NOT IN (SELECT id FROM hazards)"),
            ("near_misses.reported_by", "SELECT COUNT(*) FROM near_misses WHERE reported_by IS NOT NULL AND reported_by NOT IN (SELECT id FROM employees)"),
            ("safety_walks.location_station_id", "SELECT COUNT(*) FROM safety_walks WHERE location_station_id IS NOT NULL AND location_station_id NOT IN (SELECT id FROM working_stations)"),
            ("safety_walks.inspector_id", "SELECT COUNT(*) FROM safety_walks WHERE inspector_id IS NOT NULL AND inspector_id NOT IN (SELECT id FROM employees)"),
            ("capa_actions.incident_id", "SELECT COUNT(*) FROM capa_actions WHERE incident_id IS NOT NULL AND incident_id NOT IN (SELECT id FROM incidents)"),
            ("capa_actions.responsible_person_id", "SELECT COUNT(*) FROM capa_actions WHERE responsible_person_id IS NOT NULL AND responsible_person_id NOT IN (SELECT id FROM employees)"),
            ("shift_schedule.employee_id", "SELECT COUNT(*) FROM shift_schedule WHERE employee_id IS NOT NULL AND employee_id NOT IN (SELECT id FROM employees)"),
            ("shift_schedule.station_id", "SELECT COUNT(*) FROM shift_schedule WHERE station_id IS NOT NULL AND station_id NOT IN (SELECT id FROM working_stations)"),
            ("shift_schedule.supervisor_id", "SELECT COUNT(*) FROM shift_schedule WHERE supervisor_id IS NOT NULL AND supervisor_id NOT IN (SELECT id FROM employees)"),
        ]
        any_orphans = False
        for label, query in orphan_checks:
            cur.execute(query)
            count = cur.fetchone()[0]
            if count:
                any_orphans = True
                print(f"  ⚠ {label}: {count} orphaned references")
        if not any_orphans:
            print("  No orphaned foreign keys found.")

        print("\nPreserved tables (untouched):")
        for t in PRESERVED_TABLES:
            cur.execute(f"SELECT COUNT(*) FROM {t}")
            print(f"  {t}: {cur.fetchone()[0]}")

        print("\nDone.")

    except Exception:
        conn.rollback()
        cur.execute("SET FOREIGN_KEY_CHECKS=1")
        conn.commit()
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
