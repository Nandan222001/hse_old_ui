"""
HSE Intelligence Platform - Role-wise Checklist Excel Generator
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DARK_BLUE = "1E293B"; MID_BLUE = "1D4ED8"; LIGHT_BLUE = "DBEAFE"
TEAL = "0D9488"; TEAL_LIGHT = "CCFBF1"; PURPLE = "7C3AED"; PURPLE_LIGHT = "EDE9FE"
GREEN = "16A34A"; GREEN_LIGHT = "DCFCE7"; RED = "DC2626"; RED_LIGHT = "FEE2E2"
AMBER = "D97706"; AMBER_LIGHT = "FEF3C7"; GREY_BG = "F8FAFC"; BORDER_CLR = "E2E8F0"
WHITE = "FFFFFF"; TEXT_DARK = "0F172A"; TEXT_MID = "475569"; BLACK_DARK = "1F1F1F"

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "HSE_Checklist_Roles.xlsx")

def fill(h): return PatternFill(fill_type="solid", fgColor=h)
def font(bold=False, color=TEXT_DARK, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def align(h="left", v="center", wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr(color=BORDER_CLR, sides="all"):
    s = Side(style="thin", color=color); n = Side(style=None)
    if sides == "all": return Border(left=s, right=s, top=s, bottom=s)
    elif sides == "bottom": return Border(bottom=s)
    return Border()
def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def rh(ws, row, h): ws.row_dimensions[row].height = h

def mhdr(ws, row, c1, c2, text, bg=DARK_BLUE, fg=WHITE, size=11, height=30):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.fill = fill(bg); c.font = font(bold=True, color=fg, size=size)
    c.alignment = align(h="center"); rh(ws, row, height)

def sbanner(ws, row, ncols, text, bg, fg=WHITE, size=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(bg); c.font = font(bold=True, color=fg, size=size)
    c.alignment = align(h="left"); rh(ws, row, 22)

def thdr(ws, row, headers, bg, fg=WHITE):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = fill(bg); cell.font = font(bold=True, color=fg, size=9)
        cell.alignment = align(h="center"); cell.border = bdr(color=WHITE)
    rh(ws, row, 20)

def add_dd(ws, col, r1, r2, opts):
    dv = DataValidation(type="list", formula1='"' + ",".join(opts) + '"', allow_blank=True, showDropDown=False)
    dv.sqref = f"{col}{r1}:{col}{r2}"
    ws.add_data_validation(dv)

def build_section(ws, current_row, section, headers):
    sbanner(ws, current_row, 6, f"  {section['title']}", section["bg"])
    current_row += 1
    thdr(ws, current_row, headers, section["bg"])
    current_row += 1
    items = section["items"]
    for i, (item, resp) in enumerate(items):
        rb = WHITE if i % 2 == 0 else GREY_BG
        ws.cell(row=current_row, column=1).fill = fill(rb)
        ws.cell(row=current_row, column=2, value=i+1).fill = fill(rb)
        ws.cell(row=current_row, column=2).font = font(bold=True, color=section["bg"], size=9)
        ws.cell(row=current_row, column=2).alignment = align(h="center")
        ws.cell(row=current_row, column=3, value=item).fill = fill(rb)
        ws.cell(row=current_row, column=3).font = font(color=TEXT_DARK, size=9)
        ws.cell(row=current_row, column=3).alignment = align(h="left")
        ws.cell(row=current_row, column=4, value=resp).fill = fill(rb)
        ws.cell(row=current_row, column=4).font = font(italic=True, color=TEXT_MID, size=8)
        ws.cell(row=current_row, column=4).alignment = align(h="center")
        ws.cell(row=current_row, column=5).fill = fill("FFFBEB")
        ws.cell(row=current_row, column=5).border = bdr(color=AMBER)
        ws.cell(row=current_row, column=5).alignment = align(h="center")
        ws.cell(row=current_row, column=6).fill = fill(rb)
        ws.cell(row=current_row, column=6).border = bdr(color=BORDER_CLR)
        for col in [1, 2, 3, 4]: ws.cell(row=current_row, column=col).border = bdr(color=BORDER_CLR)
        rh(ws, current_row, 30); current_row += 1
    r1_ = current_row - len(items); r2_ = current_row - 1
    if "Pass" in items[0][1]: opts = ["Pass","Fail","N/A"]
    elif "Yes" in items[0][1]: opts = ["Yes","No","N/A"]
    elif "Number" in items[0][1] or "Rating" in items[0][1]: opts = ["Entered","Pending"]
    else: opts = ["Reported","Pending","N/A"]
    add_dd(ws, "E", r1_, r2_, opts)
    rh(ws, current_row, 8); current_row += 2
    return current_row

def sig_block(ws, current_row, labels):
    for l1, l2 in labels:
        ws.cell(row=current_row, column=2, value=l1).font = font(bold=True, size=9)
        ws.cell(row=current_row, column=3).border = bdr(sides="bottom", color=DARK_BLUE)
        ws.cell(row=current_row, column=4, value=l2).font = font(bold=True, size=9)
        ws.cell(row=current_row, column=5).border = bdr(sides="bottom", color=DARK_BLUE)
        rh(ws, current_row, 22); current_row += 1
    return current_row

# ─── COVER ────────────────────────────────────────────────────────────────────
def build_cover(wb):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False; ws.sheet_properties.tabColor = DARK_BLUE
    for c, w in zip(range(1,8),[4,22,28,22,18,16,4]): cw(ws,c,w)
    rh(ws,1,12)
    mhdr(ws,2,1,7,"HSE ROLE-WISE CHECKLIST REFERENCE GUIDE",bg=DARK_BLUE,fg=WHITE,size=16,height=48)
    mhdr(ws,3,1,7,"SafeGuard HSE Intelligence Platform  |  Version 1.0  |  July 2026",bg="334155",fg="CBD5E1",size=9,height=20)
    rh(ws,4,14)
    cards=[
        ("WORKER",TEAL,TEAL_LIGHT,"Field-level daily safety checks","Pre-Shift | Vehicle Pre-Start | Post-Shift | Incident Reporting","Daily (3x/shift)","30 items/day"),
        ("SUPERVISOR",MID_BLUE,LIGHT_BLUE,"Team & site monitoring","Morning Inspection | Mid-Shift | End-of-Shift | Weekly | Incident Investigation","Daily + Weekly","39 items"),
        ("MANAGER",PURPLE,PURPLE_LIGHT,"Compliance & strategic oversight","Daily Review | Weekly Audit | Monthly Compliance | Incident Closure","Daily+Weekly+Monthly","32 items"),
    ]
    r = 5
    for title,hbg,bbg,focus,checks,freq,items in cards:
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
        c=ws.cell(row=r,column=2,value=f"  {title}"); c.fill=fill(hbg); c.font=font(bold=True,color=WHITE,size=13); c.alignment=align(h="left"); rh(ws,r,30)
        for j,(lbl,val) in enumerate([("Focus:",focus),("Checklists:",checks),("Frequency:",freq),("Total Items:",items)]):
            rr=r+1+j
            ws.merge_cells(start_row=rr,start_column=2,end_row=rr,end_column=3)
            ws.merge_cells(start_row=rr,start_column=4,end_row=rr,end_column=6)
            lc=ws.cell(row=rr,column=2,value=lbl); lc.fill=fill(bbg); lc.font=font(bold=True,color=hbg,size=9); lc.alignment=align(h="right")
            vc=ws.cell(row=rr,column=4,value=val); vc.fill=fill(bbg); vc.font=font(color=TEXT_DARK,size=9); vc.alignment=align(h="left")
            rh(ws,rr,22)
        for col in range(2,7): ws.cell(row=r+5,column=col).fill=fill(GREY_BG)
        rh(ws,r+5,6); r+=7
    # Incident highlight
    r2=r+1
    ws.merge_cells(start_row=r2,start_column=2,end_row=r2,end_column=6)
    c=ws.cell(row=r2,column=2,value="INCIDENT REPORTING - Applicable to ALL 3 Roles (Worker > Supervisor > Manager > Admin)")
    c.fill=fill(RED); c.font=font(bold=True,color=WHITE,size=10); c.alignment=align(h="center"); rh(ws,r2,28)
    for j,(role,desc) in enumerate([("Worker","Reports incident immediately - photo + location + severity"),("Supervisor","Acknowledges 30min > Investigates > Raises CAPA"),("Manager","Approves > Regulatory notification > Closes incident"),("Admin","Real-time dashboard - color-coded severity alerts")]):
        rr=r2+1+j
        ws.cell(row=rr,column=2,value=role).fill=fill(RED_LIGHT); ws.cell(row=rr,column=2).font=font(bold=True,color=RED,size=9); ws.cell(row=rr,column=2).alignment=align(h="center")
        ws.merge_cells(start_row=rr,start_column=3,end_row=rr,end_column=6)
        ws.cell(row=rr,column=3,value=desc).fill=fill(RED_LIGHT); ws.cell(row=rr,column=3).font=font(color=TEXT_DARK,size=9); rh(ws,rr,20)
    rh(ws,r2+6,8)
    ws.merge_cells(start_row=r2+7,start_column=2,end_row=r2+7,end_column=6)
    ws.cell(row=r2+7,column=2,value="Navigate using tabs: Worker | Supervisor | Manager | Incident Reporting | Severity Matrix").font=font(italic=True,color=TEXT_MID,size=8)
    ws.cell(row=r2+7,column=2).alignment=align(h="center")

# ─── WORKER ──────────────────────────────────────────────────────────────────
def build_worker(wb):
    ws = wb.create_sheet("Worker Checklist")
    ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=TEAL
    for c,w in zip(range(1,8),[5,6,52,22,16,22,5]): cw(ws,c,w)
    rh(ws,1,10)
    mhdr(ws,2,1,6,"WORKER - HSE Checklist",bg=TEAL,fg=WHITE,size=14,height=40)
    mhdr(ws,3,1,6,"Frequency: Daily (Pre-Shift, During, Post-Shift)  |  Submit via Mobile App",bg="0F766E",fg="CCFBF1",size=8,height=18)
    headers=["","#","Checklist Item","Response Type","Result","Remarks / Evidence"]
    SECTIONS=[
        {"title":"SECTION 1.1 - Pre-Shift / Pre-Task Safety Checklist (Daily - Before starting work)","bg":TEAL,"items":[
            ("PPE Check - Helmet, Gloves, Safety Boots, Hi-Vis Vest, Eye Protection - all present and undamaged","Pass / Fail"),
            ("Work Permit received, reviewed and fully understood before starting task","Yes / No"),
            ("Tools & Equipment condition check - no visible damage, properly calibrated","Pass / Fail"),
            ("Work area hazards identified - slippery surfaces, overhead work, electrical hazards","Pass / Fail"),
            ("Emergency exit location known and route is completely clear","Yes / No"),
            ("Fire Extinguisher accessible and pressure gauge in green zone","Pass / Fail"),
            ("First Aid kit location confirmed and kit is fully stocked","Yes / No"),
            ("Worker feels physically and mentally fit for duty - no dizziness/illness/fatigue","Yes / No"),
            ("Toolbox Talk attended and content understood for today's tasks","Yes / No"),
            ("Housekeeping - work area clean, walkways clear, no trip hazards before starting","Pass / Fail"),
        ]},
        {"title":"SECTION 1.2 - Vehicle / Equipment Pre-Start Check (When operating vehicles or machinery)","bg":"0891B2","items":[
            ("Braking Systems - test service brakes and parking brake for response and firmness","Pass / Fail / N/A"),
            ("Tyres & Wheels - tread depth adequate, correct pressure, wheel nuts secure","Pass / Fail / N/A"),
            ("Lights & Indicators - headlights, taillights, beacons, reverse alarm all functional","Pass / Fail / N/A"),
            ("Fire Extinguisher on vehicle - present, pressure in green zone, tag valid","Pass / Fail / N/A"),
            ("Fluid Levels - engine oil, coolant/water, hydraulic fluid at correct levels","Pass / Fail / N/A"),
            ("Mirrors & Visibility - all mirrors clean, correctly adjusted, no blind spots","Pass / Fail / N/A"),
            ("Seatbelt - functional, latches securely, no fraying or damage","Pass / Fail / N/A"),
            ("Horn - sounds clearly and audibly when tested","Pass / Fail / N/A"),
        ]},
        {"title":"SECTION 1.3 - Post-Shift / End-of-Day Checklist (Before leaving site)","bg":"0D9488","items":[
            ("Work area cleaned - all waste removed, surfaces swept, spills cleaned","Yes / No"),
            ("All tools returned to designated storage and properly secured","Yes / No"),
            ("Any incidents or near-misses occurred during shift? (Report if Yes)","Yes / No + Report"),
            ("Any equipment damage discovered during shift? (Log with photo if Yes)","Yes / No + Photo"),
            ("Hazardous waste disposed as per site waste management procedure","Yes / No"),
            ("Work Permit formally closed with supervisor sign-off","Yes / No"),
        ]},
        {"title":"SECTION 1.4 - INCIDENT REPORTING (Worker) - Immediately when any incident occurs","bg":RED,"items":[
            ("Injury / Accident - self or another worker - report even minor cuts/bruises","Report Now"),
            ("Near Miss - event that could have caused injury but did not (e.g. falling object)","Report Now"),
            ("Unsafe Act - observed a colleague not following safe work procedure","Report Now"),
            ("Unsafe Condition - broken equipment, oil spill, loose wiring, any physical hazard","Report Now"),
            ("Property Damage - any equipment, vehicle, or site asset damaged","Report Now"),
            ("Environmental Spill - any chemical, fuel, or hazardous material on ground/water","Report Now"),
        ]},
    ]
    r=5; rh(ws,4,8); rh(ws,r-1,8)
    for s in SECTIONS: r=build_section(ws,r,s,headers)
    sbanner(ws,r,6,"  Worker Declaration & Signature",DARK_BLUE); r+=1
    sig_block(ws,r,[("Worker Name:","Employee ID:"),("Date:","Time:"),("Signature:","Shift: Morning / Afternoon / Night")])

# ─── SUPERVISOR ───────────────────────────────────────────────────────────────
def build_supervisor(wb):
    ws = wb.create_sheet("Supervisor Checklist")
    ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=MID_BLUE
    for c,w in zip(range(1,8),[5,6,52,22,16,22,5]): cw(ws,c,w)
    rh(ws,1,10)
    mhdr(ws,2,1,6,"SUPERVISOR - HSE Checklist",bg=MID_BLUE,fg=WHITE,size=14,height=40)
    mhdr(ws,3,1,6,"Frequency: Daily (Morning, Mid-Shift, Evening) + Weekly  |  Submit via Mobile App",bg="1E40AF",fg="BFDBFE",size=8,height=18)
    headers=["","#","Checklist Item","Response Type","Result","Remarks / Evidence"]
    SECTIONS=[
        {"title":"SECTION 2.1 - Morning Site Safety Inspection (Daily - Before shift commences)","bg":MID_BLUE,"items":[
            ("All workers on site verified to be wearing correct and complete PPE","Pass / Fail"),
            ("Work Permits issued, valid, and appropriate for all tasks planned today","Pass / Fail"),
            ("Toolbox Talk conducted with full team - attendance recorded (count noted)","Yes / No + Count"),
            ("All hazardous areas properly barricaded with correct signage clearly visible","Pass / Fail"),
            ("Emergency assembly point communicated and known to all team members","Yes / No"),
            ("First Aid kit on site fully stocked and easily accessible","Pass / Fail"),
            ("Fire extinguishers checked - location confirmed, pressure gauge green, valid tag","Pass / Fail"),
            ("Hot work area under proper control (fire watch, permits, suppression ready) - if applicable","Pass / Fail / N/A"),
            ("Overall site housekeeping satisfactory - walkways clear, materials stored safely","Pass / Fail"),
            ("All workers confirmed fit for duty - no signs of fatigue, illness, or substance impairment","Yes / No"),
        ]},
        {"title":"SECTION 2.2 - Mid-Shift Inspection (During work hours - at least once per shift)","bg":"2563EB","items":[
            ("Workers observed following approved safe work procedures and JSA","Pass / Fail"),
            ("No unauthorized personnel present in active or restricted work zones","Yes / No"),
            ("Equipment and machinery being used correctly and within rated capacity","Pass / Fail"),
            ("Any near-miss or incident occurred since morning inspection?","Yes / No + Report"),
            ("Site access roads and pedestrian paths remain clear and safe","Pass / Fail"),
            ("Noise, dust, and chemical exposure verified within permissible limits","Pass / Fail / N/A"),
            ("Waste disposal being carried out correctly per environmental management plan","Pass / Fail"),
        ]},
        {"title":"SECTION 2.3 - End-of-Shift Closeout Checklist (Before releasing workers)","bg":"3B82F6","items":[
            ("All active Work Permits formally closed and signed off by supervisor","Yes / No"),
            ("Full headcount of all workers confirmed - no one unaccounted for","Yes / No + Count"),
            ("All equipment and machinery properly shut down, isolated, and secured","Yes / No"),
            ("Any incidents or near-misses during shift reported to Manager immediately","Yes / No"),
            ("Site properly secured before worker departure - fencing, locks, signage","Yes / No"),
            ("Tomorrow's planned work reviewed for potential hazards - pre-planning done","Yes / No"),
        ]},
        {"title":"SECTION 2.4 - Weekly Team Safety Observation Report","bg":"0EA5E9","items":[
            ("Total safety observations raised by team this week (positive + negative)","Number"),
            ("All CAPA from this week closed within agreed SLA timeframe","Yes / No + %"),
            ("Training compliance rate for all team members - mandatory modules completed","% Score"),
            ("Near-miss reporting culture - team actively reporting without fear of blame","Rating 1-5"),
            ("All toolbox talk records filed and stored for audit trail","Yes / No"),
            ("Weekly safety performance summary submitted to Manager on time","Yes / No"),
        ]},
        {"title":"SECTION 2.5 - INCIDENT REPORTING & INVESTIGATION (Supervisor) - On every reported incident","bg":RED,"items":[
            ("Incident acknowledged within 30 minutes of worker report","Yes / No + Time"),
            ("Scene assessment completed - incident location physically inspected","Yes / No"),
            ("Immediate action taken - work stopped if required, hazard isolated","Yes / No + Action"),
            ("Injured person given first aid / referred to medical facility if needed","Yes / No"),
            ("Incident scene preserved - not disturbed until investigation is complete","Yes / No"),
            ("Root cause analysis completed using 5-Why method","Yes / No"),
            ("Witness statements collected from relevant team members","Yes / No + Count"),
            ("CAPA (Corrective & Preventive Actions) raised with assigned owner and due date","Yes / No"),
            ("Serious incident escalated to Manager immediately - within 1 hour","Yes / No / N/A"),
            ("Incident classified: LTI / MTI / First Aid / Near Miss / Property Damage","Classification"),
        ]},
    ]
    r=5; rh(ws,4,8); rh(ws,r-1,8)
    for s in SECTIONS: r=build_section(ws,r,s,headers)
    sbanner(ws,r,6,"  Supervisor Declaration & Signature",DARK_BLUE); r+=1
    sig_block(ws,r,[("Supervisor Name:","Employee ID:"),("Date:","Shift:"),("Signature:","Team Size:")])

# ─── MANAGER ──────────────────────────────────────────────────────────────────
def build_manager(wb):
    ws = wb.create_sheet("Manager Checklist")
    ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=PURPLE
    for c,w in zip(range(1,8),[5,6,52,22,16,22,5]): cw(ws,c,w)
    rh(ws,1,10)
    mhdr(ws,2,1,6,"MANAGER - HSE Checklist",bg=PURPLE,fg=WHITE,size=14,height=40)
    mhdr(ws,3,1,6,"Frequency: Daily Review + Weekly Audit + Monthly Compliance  |  Admin Dashboard View",bg="6D28D9",fg="DDD6FE",size=8,height=18)
    headers=["","#","Checklist Item","Response Type","Result","Remarks / Evidence"]
    SECTIONS=[
        {"title":"SECTION 3.1 - Daily Management Review Checklist (Every working day)","bg":PURPLE,"items":[
            ("All supervisors have submitted their morning site inspection reports on time","Yes / No"),
            ("All critical incidents from last 24 hours reviewed and immediate actions confirmed","Yes / No + Action"),
            ("Permit-to-Work compliance reviewed - no unauthorized work without valid permit","Pass / Fail"),
            ("KPI dashboard reviewed - incident rate, near-miss count, training compliance %","Yes / No"),
            ("All overdue CAPA (Corrective and Preventive Actions) escalated and actioned","Yes / No"),
            ("Emergency contacts and escalation call tree current and distributed to leadership","Yes / No"),
        ]},
        {"title":"SECTION 3.2 - Weekly Site Management Audit (Every week - comprehensive review)","bg":"6D28D9","items":[
            ("HSE training records up to date for all permanent staff and contractors on site","Pass / Fail"),
            ("Legal compliance documents valid - licenses, certifications, permits not expired","Pass / Fail"),
            ("Risk assessments reviewed and approved for all active high-risk tasks","Pass / Fail"),
            ("All inspection findings actioned within agreed SLA timeframes","Yes / No + % Closed"),
            ("Weekly toolbox talk records maintained, signed, and filed for all teams","Yes / No"),
            ("Near-miss trend analysis completed - patterns identified and addressed proactively","Yes / No"),
            ("Manager personally conducted a safety walkthrough of the full site this week","Yes / No + Date"),
            ("Contractor HSE compliance verified - contractor checklists reviewed and signed","Pass / Fail / N/A"),
            ("HSE resource budget reviewed - PPE stock, equipment, training budget adequate","Yes / No"),
            ("Weekly HSE meeting with all supervisors - minutes recorded and distributed","Yes / No + Minutes"),
        ]},
        {"title":"SECTION 3.3 - Monthly Compliance Checklist (End of every month)","bg":"7C3AED","items":[
            ("Monthly HSE performance report compiled and submitted to senior management/client","Yes / No"),
            ("All audit findings from previous month formally closed - percentage tracked","Yes / No + % Closed"),
            ("All statutory regulatory returns and government filings submitted on time","Yes / No"),
            ("Emergency evacuation drill conducted - participation rate and debrief completed","Yes / No + Date"),
            ("HSE Policy reviewed for relevance and communicated to all staff on site","Yes / No"),
            ("All incident investigation reports finalized, root causes identified, CAPA raised","Yes / No"),
            ("Worker feedback on safety culture collected via survey or toolbox talk forms","Yes / No"),
            ("HSE legal register reviewed for new/updated regulations applicable to operations","Yes / No"),
        ]},
        {"title":"SECTION 3.4 - INCIDENT REPORTING - Manager Review & Closure (On every reported incident)","bg":RED,"items":[
            ("Supervisor's incident investigation report received and reviewed within 24 hours","Yes / No + Date"),
            ("Root cause analysis approved as complete and accurate","Yes / No"),
            ("Regulatory authority notification sent if incident meets statutory threshold","Yes / No / N/A"),
            ("All CAPA actions assigned with owners and due dates tracked in system","Yes / No"),
            ("Lessons learned documented and formally communicated to all site teams","Yes / No"),
            ("Incident formally closed in the HSE system with Manager sign-off","Yes / No + Date"),
            ("KPI impact recorded - LTIFR / TRIFR / Near-Miss rates updated in dashboard","Yes / No"),
            ("Insurance / legal documentation prepared if incident involves liability","Yes / No / N/A"),
        ]},
    ]
    r=5; rh(ws,4,8); rh(ws,r-1,8)
    for s in SECTIONS: r=build_section(ws,r,s,headers)
    sbanner(ws,r,6,"  Manager Declaration & Signature",DARK_BLUE); r+=1
    sig_block(ws,r,[("Manager Name:","Employee ID:"),("Date:","Site / Project:"),("Signature:","Review Period:")])

# ─── INCIDENT REPORTING ───────────────────────────────────────────────────────
def build_incident(wb):
    ws = wb.create_sheet("Incident Reporting")
    ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=RED
    for c,w in zip(range(1,8),[4,14,36,34,20,20,4]): cw(ws,c,w)
    rh(ws,1,10)
    mhdr(ws,2,1,6,"INCIDENT REPORTING - All Roles Reference",bg=RED,fg=WHITE,size=14,height=40)
    mhdr(ws,3,1,6,"Defines incident types, responsibilities, timeframes and escalation flow for Worker, Supervisor & Manager",bg="991B1B",fg="FEE2E2",size=8,height=18)
    r=5
    sbanner(ws,r,6,"  Types of Incidents to Report (Applicable to ALL Roles)",RED); r+=1
    thdr(ws,r,["","#","Incident Type","Description","Who Reports","Severity"],RED); r+=1
    for i,(num,itype,desc,who,sev) in enumerate([
        ("1","Injury / Accident","Any physical harm - cuts, fractures, burns, chemical exposure","Worker (immediately)","Low to Critical"),
        ("2","Near Miss","Event that COULD have caused injury but did not - report without fail","Worker (same day)","Low to High"),
        ("3","Unsafe Act","Colleague not following safe work procedure or violating safety rules","Worker / Supervisor","Low to Medium"),
        ("4","Unsafe Condition","Broken equipment, oil spill, loose wiring, structural damage","Worker / Supervisor","Low to High"),
        ("5","Property Damage","Damage to equipment, vehicles, buildings, or any site asset","Worker / Supervisor","Low to High"),
        ("6","Environmental Spill","Chemical, fuel, or hazardous material release to ground/water/air","Worker / Supervisor","Medium to Critical"),
        ("7","Fire / Explosion","Any uncontrolled fire, explosion, or near-explosion event on site","Supervisor (immediate)","High to Critical"),
        ("8","Electrical Incident","Electrical shock, arc flash, short circuit, or equipment failure","Supervisor (immediate)","High to Critical"),
    ]):
        rb=WHITE if i%2==0 else RED_LIGHT
        for col,val,b in [(1,"",False),(2,num,True),(3,itype,True),(4,desc,False),(5,who,False),(6,sev,True)]:
            c=ws.cell(row=r,column=col,value=val); c.fill=fill(rb)
            c.font=font(bold=b,color=RED if b and col in[2,6] else TEXT_DARK,size=9)
            c.alignment=align(h="center" if col in[2,5,6] else "left"); c.border=bdr()
        rh(ws,r,28); r+=1
    r+=1
    sbanner(ws,r,6,"  Incident Escalation Flow  (Worker > Supervisor > Manager > Admin Dashboard)",DARK_BLUE); r+=1
    for role,hbg,bbg,steps in [
        ("WORKER",TEAL,TEAL_LIGHT,[
            "Step 1: STOP work immediately if there is an ongoing danger",
            "Step 2: Ensure the affected person gets First Aid / Emergency help",
            "Step 3: Open the HSE mobile app > Tap 'Report Incident'",
            "Step 4: Fill in: What happened, Where, When, Who was affected",
            "Step 5: Attach photo or video evidence of the scene",
            "Step 6: Select Severity (Low / Medium / High / Critical)",
            "Step 7: Submit - Supervisor gets instant push notification",
        ]),
        ("SUPERVISOR",MID_BLUE,LIGHT_BLUE,[
            "Step 1: Acknowledge the report within 30 minutes of receipt",
            "Step 2: Go to the scene - physically assess the situation",
            "Step 3: If work is unsafe - STOP work, isolate hazard immediately",
            "Step 4: Ensure injured person receives appropriate medical care",
            "Step 5: Preserve the scene - do NOT move evidence before investigation",
            "Step 6: Collect witness statements from all relevant team members",
            "Step 7: Perform 5-Why Root Cause Analysis in the mobile app",
            "Step 8: Raise CAPA - assign to responsible person with due date",
            "Step 9: If High/Critical - IMMEDIATELY notify Manager (within 1 hour)",
            "Step 10: Update investigation status in app - Manager notified on completion",
        ]),
        ("MANAGER",PURPLE,PURPLE_LIGHT,[
            "Step 1: Receive escalation from Supervisor for High/Critical incidents",
            "Step 2: Review Supervisor's investigation report for completeness",
            "Step 3: Approve or return investigation report for further action",
            "Step 4: Assess if incident meets statutory reporting threshold",
            "Step 5: If yes - notify regulatory authority (OSHA, Factory Inspector, etc.)",
            "Step 6: Ensure all CAPA actions are assigned, tracked, and closed on time",
            "Step 7: Document Lessons Learned and circulate to all site teams",
            "Step 8: Update KPIs - LTIFR, TRIFR, Near-Miss rate in dashboard",
            "Step 9: Formally close incident in the HSE system with Manager sign-off",
            "Step 10: Prepare insurance / legal documentation if liability is involved",
        ]),
    ]:
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
        hc=ws.cell(row=r,column=2,value=f"  {role}"); hc.fill=fill(hbg); hc.font=font(bold=True,color=WHITE,size=10); hc.alignment=align(h="left"); rh(ws,r,24); r+=1
        for step in steps:
            ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
            sc=ws.cell(row=r,column=2,value=f"    {step}"); sc.fill=fill(bbg); sc.font=font(color=TEXT_DARK,size=9); sc.alignment=align(h="left"); sc.border=bdr(); rh(ws,r,20); r+=1
        r+=1
    r+=1
    sbanner(ws,r,6,"  Mandatory Response Timeframes",DARK_BLUE); r+=1
    thdr(ws,r,["","Action","Timeframe","Notes / Conditions","",""],DARK_BLUE); r+=1
    for i,(action,tf,notes) in enumerate([
        ("Worker Reports Incident","Immediately","Within the same hour of occurrence"),
        ("Supervisor Acknowledges Report","Within 30 minutes","After receiving worker's report notification"),
        ("Supervisor Completes Investigation","Within 24 hours","For Low/Medium severity incidents"),
        ("Supervisor Escalates to Manager","Within 1 hour","For High/Critical severity incidents ONLY"),
        ("Manager Approves Investigation","Within 24 hours","After receiving Supervisor's completed report"),
        ("Regulatory Authority Notification","Within 24-48 hours","Subject to local legal requirements"),
        ("CAPA Closure - Low Severity","Within 7 days","All corrective actions completed and verified"),
        ("CAPA Closure - High/Critical","Within 3 days","Immediate corrective measures + long-term fix"),
        ("Incident Formal Closure","Within 72 hours","After all CAPA actions confirmed complete"),
    ]):
        rb=WHITE if i%2==0 else GREY_BG
        ws.cell(row=r,column=2,value=action).fill=fill(rb); ws.cell(row=r,column=2).font=font(bold=True,color=TEXT_DARK,size=9); ws.cell(row=r,column=2).border=bdr()
        ws.cell(row=r,column=3,value=tf).fill=fill(RED_LIGHT); ws.cell(row=r,column=3).font=font(bold=True,color=RED,size=9); ws.cell(row=r,column=3).alignment=align(h="center"); ws.cell(row=r,column=3).border=bdr()
        ws.merge_cells(start_row=r,start_column=4,end_row=r,end_column=6)
        ws.cell(row=r,column=4,value=notes).fill=fill(rb); ws.cell(row=r,column=4).font=font(color=TEXT_MID,size=8.5); ws.cell(row=r,column=4).border=bdr()
        rh(ws,r,22); r+=1

# ─── SEVERITY MATRIX ─────────────────────────────────────────────────────────
def build_severity(wb):
    ws = wb.create_sheet("Severity Matrix")
    ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=AMBER
    for c,w in zip(range(1,7),[4,14,30,32,22,4]): cw(ws,c,w)
    rh(ws,1,10)
    mhdr(ws,2,1,5,"INCIDENT SEVERITY MATRIX & CLASSIFICATION",bg=DARK_BLUE,fg=WHITE,size=13,height=38)
    mhdr(ws,3,1,5,"Use this matrix to classify every reported incident - determines response time and escalation level",bg="334155",fg="CBD5E1",size=8,height=18)
    r=5
    sbanner(ws,r,5,"  Severity Level Classification",DARK_BLUE); r+=1
    thdr(ws,r,["Level","Color Code","Examples","Immediate Response Required","Response Time"],DARK_BLUE); r+=1
    for level,bg,lbg,examples,response,rt in [
        ("LOW (1)","16A34A","DCFCE7","Minor cut, small spill, near miss with no injury, minor tool damage","First aid on site - supervisor informed - CAPA raised within 24hrs","24 hours"),
        ("MEDIUM (2)","D97706","FEF3C7","Sprain/strain, slip needing medical treatment, equipment damage, fuel spill","Medical treatment - supervisor investigates - Manager informed same day","4 hours"),
        ("HIGH (3)","DC2626","FEE2E2","Fracture, severe laceration, chemical exposure, fire, vehicle collision, fall from height","Emergency medical - work STOPPED - Manager notified IMMEDIATELY - site secured","Immediate"),
        ("CRITICAL (4)","1F1F1F","2D2D2D","Fatality, permanent disability, major explosion, large chemical release, collapse","Emergency services called - site locked down - CEO & Regulator notified NOW","Immediate"),
    ]:
        ws.cell(row=r,column=1,value=level).fill=fill(bg); ws.cell(row=r,column=1).font=font(bold=True,color=WHITE,size=9); ws.cell(row=r,column=1).alignment=align(h="center"); ws.cell(row=r,column=1).border=bdr()
        ws.cell(row=r,column=2,value=f"  {level.split('(')[0].strip()}").fill=fill(bg); ws.cell(row=r,column=2).font=font(bold=True,color=WHITE,size=9); ws.cell(row=r,column=2).alignment=align(h="center"); ws.cell(row=r,column=2).border=bdr()
        ws.cell(row=r,column=3,value=examples).fill=fill(lbg); ws.cell(row=r,column=3).font=font(color=TEXT_DARK,size=8.5); ws.cell(row=r,column=3).border=bdr()
        ws.cell(row=r,column=4,value=response).fill=fill(lbg); ws.cell(row=r,column=4).font=font(color=TEXT_DARK,size=8.5); ws.cell(row=r,column=4).border=bdr()
        ws.cell(row=r,column=5,value=rt).fill=fill(bg); ws.cell(row=r,column=5).font=font(bold=True,color=WHITE,size=9); ws.cell(row=r,column=5).alignment=align(h="center"); ws.cell(row=r,column=5).border=bdr()
        rh(ws,r,40); r+=1
    r+=1
    sbanner(ws,r,5,"  Key Rules for All Roles",DARK_BLUE); r+=1
    for icon,title,desc in [
        ("!","Mandatory Reporting","Every incident - no matter how small - MUST be reported. No exceptions, no coverups."),
        ("Photo","Photo Evidence","Always attach a photo or video. Mandatory for all Fail / No responses."),
        ("Time","Never Delay","Report immediately - investigation follows. Delays increase risk and legal liability."),
        ("Safe","No Blame Culture","Reporting is not about punishment. It is about learning and prevention."),
        ("CAPA","CAPA is Mandatory","Every incident must result in at least one Corrective Action (CAPA) being raised."),
        ("Admin","Admin Visibility","Admin Dashboard shows all incidents in real-time with color-coded severity alerts."),
    ]:
        ws.cell(row=r,column=1,value=icon).fill=fill(GREY_BG); ws.cell(row=r,column=1).alignment=align(h="center"); ws.cell(row=r,column=1).border=bdr()
        ws.cell(row=r,column=2,value=title).fill=fill(GREY_BG); ws.cell(row=r,column=2).font=font(bold=True,color=TEXT_DARK,size=9); ws.cell(row=r,column=2).border=bdr()
        ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=5)
        ws.cell(row=r,column=3,value=desc).fill=fill(GREY_BG); ws.cell(row=r,column=3).font=font(color=TEXT_MID,size=8.5); ws.cell(row=r,column=3).border=bdr()
        rh(ws,r,26); r+=1

# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_cover(wb)
    build_worker(wb)
    build_supervisor(wb)
    build_manager(wb)
    build_incident(wb)
    build_severity(wb)
    wb.save(OUTPUT)
    print(f"Excel file generated: {OUTPUT}")

if __name__ == "__main__":
    main()
