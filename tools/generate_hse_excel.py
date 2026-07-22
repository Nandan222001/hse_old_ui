"""
Generate HSE Intelligence Excel workbook with all 18 sheets.
Exact column order matches excel_import_service.py (r[0]=ID, r[1]=first field, etc.)
Run: python generate_hse_excel.py
Output: HSE_Import_Template.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

HEADER_FILL   = PatternFill("solid", fgColor="0B3D91")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=11)
SUBHEAD_FILL  = PatternFill("solid", fgColor="1D4ED8")
SUBHEAD_FONT  = Font(color="FFFFFF", bold=True, size=10)
DATA_FILL_1   = PatternFill("solid", fgColor="EFF6FF")
DATA_FILL_2   = PatternFill("solid", fgColor="FFFFFF")
THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_sheet(ws, headers, col_widths=None):
    """Apply header row styling and set column widths."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.row_dimensions[1].height = 30
        if col_widths and col_idx <= len(col_widths):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx-1]
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

def add_rows(ws, rows):
    """Add data rows with alternating fill."""
    for row_idx, row_data in enumerate(rows, 2):
        fill = DATA_FILL_1 if row_idx % 2 == 0 else DATA_FILL_2
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            cell.border = BORDER
    ws.freeze_panes = "A2"

# ── Sheet 1: Organisation ────────────────────────────────────────────────────
ws = wb.create_sheet("Organisation")
headers = ["ID","Organisation_Name","Country","Industry_Sector","Number_Of_Employees",
           "Headquarters_Location","Parent_Company","ISO_45001_Status",
           "Regulatory_Authority","Establishment_Date"]
style_sheet(ws, headers, [8,30,20,25,22,30,25,20,30,20])
add_rows(ws, [
    ["ORG001","WindTech Industries Ltd","United Kingdom","Manufacturing",850,
     "Sheffield, UK","WindTech Global PLC","Certified","HSE UK","2005-03-15"],
    ["ORG002","AlexCarry Logistics","United Kingdom","Logistics",320,
     "Manchester, UK","","In Progress","HSE UK","2010-07-22"],
])

# ── Sheet 2: Hazard_Categories ───────────────────────────────────────────────
ws = wb.create_sheet("Hazard_Categories")
headers = ["ID","Category_Name","Description"]
style_sheet(ws, headers, [8,30,50])
add_rows(ws, [
    ["HAZ_CAT001","Chemical","Hazards from chemicals, solvents, acids and cleaning agents"],
    ["HAZ_CAT002","Electrical","Electrical shock, arc flash and short circuit hazards"],
    ["HAZ_CAT003","Mechanical","Moving parts, rotating machinery and pinch point hazards"],
    ["HAZ_CAT004","Fire & Explosion","Flammable materials, hot work ignition and pressure vessel hazards"],
    ["HAZ_CAT005","Working at Height","Falls from ladders, scaffolds, rooftops and elevated platforms"],
    ["HAZ_CAT006","Manual Handling","Lifting, carrying and repetitive strain injury hazards"],
    ["HAZ_CAT007","Confined Space","Oxygen deficiency, toxic gas and engulfment in enclosed areas"],
    ["HAZ_CAT008","Noise & Vibration","Hearing damage from excessive noise and hand-arm vibration"],
])

# ── Sheet 3: Hazards ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Hazards")
headers = ["ID","Category_ID","Hazard_Name","Severity","Probability"]
style_sheet(ws, headers, [8,14,40,18,18])
add_rows(ws, [
    ["HAZ001","HAZ_CAT001","Acid spill from storage tank","High","Medium"],
    ["HAZ002","HAZ_CAT002","Live electrical panel exposure","Critical","Low"],
    ["HAZ003","HAZ_CAT003","Unguarded conveyor belt pinch point","High","High"],
    ["HAZ004","HAZ_CAT004","Hot work near flammable store","Critical","Medium"],
    ["HAZ005","HAZ_CAT005","Unsecured scaffold board at 8m height","High","Medium"],
    ["HAZ006","HAZ_CAT006","Repetitive heavy lifting (>25kg)","Medium","High"],
    ["HAZ007","HAZ_CAT007","Low oxygen in storage vessel","Critical","Low"],
    ["HAZ008","HAZ_CAT008","Grinding operation without ear protection","Medium","High"],
])

# ── Sheet 4: Roles ───────────────────────────────────────────────────────────
ws = wb.create_sheet("Roles")
headers = ["ID","Role_Name","Job_Category","Authority_Level","Permit_Authority","Safety_Signatory"]
style_sheet(ws, headers, [8,28,25,20,20,20])
add_rows(ws, [
    ["ROLE001","HSE Manager","Management","Level 3","Yes","Yes"],
    ["ROLE002","Production Supervisor","Operations","Level 2","Yes","Yes"],
    ["ROLE003","Maintenance Technician","Technical","Level 2","Yes","No"],
    ["ROLE004","Chemical Operator","Operations","Level 1","No","No"],
    ["ROLE005","Welder","Technical","Level 1","No","No"],
    ["ROLE006","Forklift Operator","Operations","Level 1","No","No"],
    ["ROLE007","Safety Auditor","Compliance","Level 2","No","Yes"],
    ["ROLE008","Shift Supervisor","Operations","Level 2","Yes","Yes"],
])

# ── Sheet 5: Sites ───────────────────────────────────────────────────────────
ws = wb.create_sheet("Sites")
headers = ["ID","Site_Name","Address","Postcode","City","Type","Operational_Status",
           "Number_Of_Working_Stations","Capacity","Primary_Products","Hazard_Classification"]
style_sheet(ws, headers, [8,28,35,12,18,18,20,24,12,28,25])
add_rows(ws, [
    ["SITE001","North Plant","Unit 12 Steel Road","S1 2AB","Sheffield","Manufacturing","Active",
     12,500,"Wind Turbine Blades","High"],
    ["SITE002","Warehouse A","Logistics Park, Bay 5","M1 3CD","Manchester","Warehouse","Active",
     6,200,"Finished Goods Storage","Medium"],
    ["SITE003","South Assembly","Assembly Drive","S4 7EF","Sheffield","Assembly","Active",
     8,350,"Turbine Assembly","High"],
])

# ── Sheet 6: Permit_Types ────────────────────────────────────────────────────
ws = wb.create_sheet("Permit_Types")
headers = ["ID","Permit_Type_Name","Risk_Level","Validity_Period_Hours","Concurrent_Limit"]
style_sheet(ws, headers, [8,30,18,24,20])
add_rows(ws, [
    ["PT001","Hot Work Permit","High",8,2],
    ["PT002","Confined Space Entry","Critical",4,1],
    ["PT003","Working at Height","High",8,3],
    ["PT004","Electrical Isolation","High",12,2],
    ["PT005","General Work Permit","Medium",24,5],
    ["PT006","Chemical Handling","High",8,2],
])

# ── Sheet 7: Training_Programs ───────────────────────────────────────────────
ws = wb.create_sheet("Training_Programs")
headers = ["ID","Training_Name","Duration_Hours","Frequency","Certification","Expiry_Months"]
style_sheet(ws, headers, [8,35,18,20,18,18])
add_rows(ws, [
    ["TP001","Fire Safety & Evacuation",4,"Annual","Certificate",12],
    ["TP002","Manual Handling & Ergonomics",2,"Annual","Certificate",12],
    ["TP003","Working at Height Safety",8,"Biannual","PASMA/IPAF",24],
    ["TP004","Confined Space Entry",16,"Annual","City & Guilds 6150",12],
    ["TP005","Hot Work & Fire Watch",8,"Biannual","CSWIP",24],
    ["TP006","Chemical Handling & COSHH",4,"Annual","Certificate",12],
    ["TP007","First Aid at Work",24,"3-yearly","HSE Approved FAW",36],
    ["TP008","PPE Inspection & Usage",2,"Annual","In-House",12],
])

# ── Sheet 8: Policies ────────────────────────────────────────────────────────
ws = wb.create_sheet("Policies")
headers = ["ID","Policy_Name","Category","Issue_Date","Owner","Status"]
style_sheet(ws, headers, [8,40,25,18,28,15])
add_rows(ws, [
    ["POL001","Health & Safety Policy Statement","Health & Safety","2024-01-10","HSE Manager","Active"],
    ["POL002","Personal Protective Equipment Policy","PPE","2024-02-15","HSE Manager","Active"],
    ["POL003","Hot Work & Fire Prevention Policy","Fire Safety","2024-03-01","HSE Manager","Active"],
    ["POL004","Working at Height Policy","WAH","2024-01-20","HSE Manager","Active"],
    ["POL005","Chemical Safety & COSHH Policy","Chemical","2024-02-28","HSE Manager","Active"],
    ["POL006","Permit to Work Policy","Permits","2024-01-15","Operations Manager","Active"],
    ["POL007","Incident & Near Miss Reporting Policy","Incident Management","2024-04-01","HSE Manager","Active"],
    ["POL008","Confined Space Entry Policy","Confined Space","2024-03-10","HSE Manager","Active"],
])

# ── Sheet 9: Departments ─────────────────────────────────────────────────────
ws = wb.create_sheet("Departments")
headers = ["ID","Site_ID","Department_Name","Manager_ID","Number_Of_Teams"]
style_sheet(ws, headers, [8,12,30,14,20])
add_rows(ws, [
    ["DEPT001","SITE001","Production","EMP005",4],
    ["DEPT002","SITE001","Maintenance","EMP012",2],
    ["DEPT003","SITE001","HSE","EMP001",1],
    ["DEPT004","SITE002","Logistics","EMP020",3],
    ["DEPT005","SITE003","Assembly","EMP030",3],
    ["DEPT006","SITE001","Quality","EMP015",2],
])

# ── Sheet 10: Working_Stations ───────────────────────────────────────────────
ws = wb.create_sheet("Working_Stations")
headers = ["ID","Station_Name","Site_ID","Department","Zone_Classification",
           "Primary_Hazard_ID","Staffing_Requirement","Equipment_List",
           "Permit_Types_Required","Access_Restrictions"]
style_sheet(ws, headers, [8,28,10,20,22,18,22,35,30,25])
add_rows(ws, [
    ["STN001","Assembly Line 1","SITE001","Production","Controlled Zone",
     "HAZ003",8,"Conveyor, Hydraulic Press","General Work Permit","PPE Mandatory"],
    ["STN002","Chemical Storage","SITE001","HSE","High Hazard Zone",
     "HAZ001",2,"Chemical Cabinets, Spill Kit","Chemical Handling","Trained Personnel Only"],
    ["STN003","Welding Bay","SITE001","Production","Hot Work Zone",
     "HAZ004",4,"MIG Welder, Grinding Machine","Hot Work Permit","FR Clothing Required"],
    ["STN004","Boiler Room","SITE001","Maintenance","Confined Space",
     "HAZ007",2,"Boilers, Pressure Gauges","Confined Space Entry","Permit Required"],
    ["STN005","Roof Access","SITE001","Maintenance","Working at Height",
     "HAZ005",2,"Access Ladder, Anchor Points","Working at Height","Harness Required"],
    ["STN006","Warehouse Bay 1","SITE002","Logistics","Standard Zone",
     "HAZ006",6,"Forklift, Pallet Racking","General Work Permit","PPE Mandatory"],
    ["STN007","Electrical Room","SITE001","Maintenance","Restricted Zone",
     "HAZ002",1,"LV Panels, MCC","Electrical Isolation","Authorized Personnel Only"],
    ["STN008","Grinding Station","SITE003","Assembly","Controlled Zone",
     "HAZ008",3,"Angle Grinders, Bench Grinder","Hot Work Permit","Ear & Eye Protection"],
])

# ── Sheet 11: Employees ──────────────────────────────────────────────────────
ws = wb.create_sheet("Employees")
headers = ["ID","Full_Name","Date_Of_Birth","Gender","Employment_Type",
           "Employment_Start_Date","Role_ID","Department_ID","Shift_Pattern",
           "Manager_ID","Induction_Date","Active_Status"]
style_sheet(ws, headers, [8,28,15,10,20,22,10,14,18,12,16,16])
add_rows(ws, [
    ["EMP001","John Smith","1980-05-14","M","Permanent","2015-03-01","ROLE001","DEPT003","Day","","2015-03-02","Active"],
    ["EMP002","Sarah Jones","1985-08-22","F","Permanent","2016-06-15","ROLE007","DEPT003","Day","EMP001","2016-06-16","Active"],
    ["EMP003","Mike Patel","1978-11-30","M","Permanent","2014-01-10","ROLE002","DEPT001","Morning","EMP001","2014-01-12","Active"],
    ["EMP004","Ravi Kumar","1990-03-18","M","Permanent","2018-09-01","ROLE004","DEPT001","Morning","EMP003","2018-09-03","Active"],
    ["EMP005","James Wright","1975-07-25","M","Permanent","2012-05-20","ROLE002","DEPT001","Day","EMP001","2012-05-21","Active"],
    ["EMP006","Emily Chen","1992-12-05","F","Permanent","2019-04-01","ROLE004","DEPT001","Evening","EMP005","2019-04-02","Active"],
    ["EMP007","David Okafor","1983-09-17","M","Permanent","2013-11-01","ROLE003","DEPT002","Morning","EMP012","2013-11-03","Active"],
    ["EMP008","Lisa Thompson","1988-02-28","F","Permanent","2017-07-15","ROLE003","DEPT002","Day","EMP012","2017-07-16","Active"],
    ["EMP009","Carlos Mendez","1995-06-10","M","Contract","2023-01-09","ROLE005","DEPT001","Morning","EMP003","2023-01-10","Active"],
    ["EMP010","Anna Williams","1987-04-03","F","Permanent","2016-02-20","ROLE006","DEPT004","Day","EMP020","2016-02-22","Active"],
    ["EMP011","Tom Harris","1993-10-15","M","Contract","2022-06-01","ROLE005","DEPT005","Morning","EMP030","2022-06-02","Active"],
    ["EMP012","Patricia Brown","1979-01-20","F","Permanent","2011-08-15","ROLE002","DEPT002","Day","EMP001","2011-08-16","Active"],
    ["EMP013","Kevin Lee","1991-07-08","M","Permanent","2019-10-01","ROLE004","DEPT001","Evening","EMP005","2019-10-03","Active"],
    ["EMP014","Maria Garcia","1986-03-25","F","Permanent","2015-11-01","ROLE006","DEPT004","Day","EMP020","2015-11-02","Active"],
    ["EMP015","Robert Johnson","1977-08-12","M","Permanent","2010-04-01","ROLE001","DEPT006","Day","","2010-04-02","Active"],
    ["EMP016","Aisha Khan","1994-05-30","F","Contract","2023-05-15","ROLE004","DEPT001","Morning","EMP003","2023-05-16","Active"],
    ["EMP017","George Wilson","1982-12-18","M","Permanent","2014-09-01","ROLE003","DEPT002","Evening","EMP012","2014-09-02","Active"],
    ["EMP018","Fatima Al-Hassan","1989-09-22","F","Permanent","2018-03-01","ROLE004","DEPT005","Morning","EMP030","2018-03-02","Active"],
    ["EMP019","Steve Taylor","1976-06-14","M","Permanent","2009-07-01","ROLE002","DEPT005","Day","EMP001","2009-07-02","Active"],
    ["EMP020","Claire Robinson","1984-11-05","F","Permanent","2013-04-15","ROLE002","DEPT004","Day","EMP001","2013-04-16","Active"],
    ["EMP021","Ahmed Hassan","1996-02-11","M","Contract","2024-01-08","ROLE005","DEPT005","Morning","EMP030","2024-01-09","Active"],
    ["EMP022","Priya Singh","1990-08-19","F","Permanent","2017-12-01","ROLE004","DEPT001","Evening","EMP005","2017-12-02","Active"],
    ["EMP023","Paul Davies","1981-04-27","M","Permanent","2012-10-01","ROLE003","DEPT002","Day","EMP012","2012-10-02","Active"],
    ["EMP024","Nadia Omar","1993-07-14","F","Permanent","2020-02-15","ROLE007","DEPT003","Day","EMP001","2020-02-16","Active"],
    ["EMP025","Ben Nguyen","1985-01-09","M","Permanent","2015-06-01","ROLE006","DEPT004","Day","EMP020","2015-06-02","Active"],
    ["EMP030","Daniel Foster","1974-03-08","M","Permanent","2008-09-01","ROLE002","DEPT005","Day","EMP001","2008-09-02","Active"],
])

# ── Sheet 12: Shift_Schedule ─────────────────────────────────────────────────
ws = wb.create_sheet("Shift_Schedule")
headers = ["ID","Employee_ID","Shift_Date","Shift_Type","Shift_Start","Shift_End",
           "Actual_Hours_Worked","Station_ID","Supervisor_ID"]
style_sheet(ws, headers, [8,12,15,18,13,13,22,13,14])
add_rows(ws, [
    ["SS001","EMP003","2025-01-06","Morning","06:00","14:00",8.0,"STN001","EMP001"],
    ["SS002","EMP004","2025-01-06","Morning","06:00","14:00",8.0,"STN001","EMP003"],
    ["SS003","EMP006","2025-01-06","Evening","14:00","22:00",8.0,"STN001","EMP005"],
    ["SS004","EMP007","2025-01-06","Morning","06:00","14:00",8.0,"STN004","EMP012"],
    ["SS005","EMP009","2025-01-06","Morning","06:00","14:00",7.5,"STN003","EMP003"],
    ["SS006","EMP003","2025-01-07","Morning","06:00","14:00",8.0,"STN001","EMP001"],
    ["SS007","EMP004","2025-01-07","Morning","06:00","14:00",8.0,"STN001","EMP003"],
    ["SS008","EMP013","2025-01-07","Evening","14:00","22:00",8.0,"STN001","EMP005"],
    ["SS009","EMP011","2025-01-08","Morning","06:00","14:00",8.0,"STN008","EMP030"],
    ["SS010","EMP018","2025-01-08","Morning","06:00","14:00",7.5,"STN008","EMP030"],
    ["SS011","EMP022","2025-01-08","Evening","14:00","22:00",8.0,"STN002","EMP005"],
    ["SS012","EMP016","2025-01-09","Morning","06:00","14:00",8.0,"STN001","EMP003"],
    ["SS013","EMP021","2025-01-09","Morning","06:00","14:00",8.0,"STN003","EMP003"],
    ["SS014","EMP017","2025-01-10","Day","08:00","16:00",8.0,"STN007","EMP012"],
    ["SS015","EMP023","2025-01-10","Day","08:00","16:00",8.0,"STN004","EMP012"],
    ["SS016","EMP003","2025-02-03","Morning","06:00","14:00",8.0,"STN001","EMP001"],
    ["SS017","EMP004","2025-02-03","Morning","06:00","14:00",8.0,"STN001","EMP003"],
    ["SS018","EMP009","2025-02-03","Morning","06:00","14:00",8.0,"STN003","EMP003"],
    ["SS019","EMP011","2025-03-10","Morning","06:00","14:00",8.0,"STN008","EMP030"],
    ["SS020","EMP021","2025-03-10","Morning","06:00","14:00",8.0,"STN003","EMP003"],
])

# ── Sheet 13: Incidents ──────────────────────────────────────────────────────
ws = wb.create_sheet("Incidents")
headers = ["ID","Report_Date","Incident_Date_Time","Location_Station_ID","Incident_Type",
           "Severity","Number_Persons_Involved","Description","Immediate_Cause","Root_Cause",
           "Hazard_ID","Permit_Active","Control_Failure","Reported_By",
           "Investigation_Status","CAPA_Generated","Days_Away","Root_Cause_Category"]
style_sheet(ws, headers, [8,14,22,22,18,15,24,40,35,35,12,15,18,14,22,18,13,25])
add_rows(ws, [
    ["INC001","2025-01-08","2025-01-08 09:15:00","STN003","Injury","Lost Time",1,
     "Welder suffered burns to right forearm during MIG welding operation. FR sleeve not worn.",
     "Contact with hot workpiece","Inadequate PPE compliance and supervision",
     "HAZ004","No","Yes","EMP009","Completed","Yes",3,"PPE Non-Compliance"],
    ["INC002","2025-01-15","2025-01-15 11:30:00","STN001","Injury","Minor",2,
     "Two workers strained backs lifting heavy component without mechanical aid.",
     "Manual handling without equipment","Inadequate manual handling training refresher",
     "HAZ006","No","No","EMP004","Completed","Yes",0,"Training Gap"],
    ["INC003","2025-02-03","2025-02-03 14:45:00","STN004","Dangerous Occurrence","Significant",0,
     "Boiler pressure relief valve discharged unexpectedly. No injuries. Area evacuated.",
     "Equipment malfunction","Missed scheduled maintenance inspection",
     "HAZ007","Yes","Yes","EMP007","Completed","Yes",0,"Maintenance Failure"],
    ["INC004","2025-02-20","2025-02-20 08:00:00","STN002","Injury","Lost Time",1,
     "Chemical operator sustained chemical burns to hand from acid splash during decanting.",
     "Inadequate PPE — glove hole not detected","Pre-use PPE inspection not conducted",
     "HAZ001","No","Yes","EMP022","In Progress","Yes",5,"PPE Non-Compliance"],
    ["INC005","2025-03-10","2025-03-10 10:20:00","STN008","Injury","Minor",1,
     "Grinder disc shattered. Fragment hit worker eye. Safety glasses worn but cracked.",
     "Defective grinding disc used","Pre-use equipment check not performed",
     "HAZ008","No","Yes","EMP011","Completed","Yes",0,"Equipment Defect"],
    ["INC006","2025-04-02","2025-04-02 13:10:00","STN005","Injury","Lost Time",1,
     "Worker fell from scaffold at 4m height. Harness worn but lanyard not clipped.",
     "Fall arrest system not properly connected","Lack of WAH briefing before task",
     "HAZ005","Yes","Yes","EMP021","In Progress","Yes",8,"Procedural Non-Compliance"],
    ["INC007","2025-04-18","2025-04-18 15:55:00","STN007","Dangerous Occurrence","Critical",0,
     "Arc flash from LV panel during maintenance. No injuries. Panel severely damaged.",
     "Electrical isolation LOTO not applied","Permit to Work not obtained before work",
     "HAZ002","No","Yes","EMP017","Completed","Yes",0,"Permit Violation"],
    ["INC008","2025-05-05","2025-05-05 07:30:00","STN001","Injury","Minor",1,
     "Worker slipped on oil spillage near assembly line. Twisted ankle.",
     "Spillage not cleaned immediately","Housekeeping inspection gap",
     "HAZ003","No","No","EMP013","Completed","No",2,"Housekeeping"],
    ["INC009","2025-06-12","2025-06-12 09:45:00","STN003","Injury","Lost Time",1,
     "Welder inhaled fumes due to ventilation fan failure. Hospitalised for observation.",
     "Local exhaust ventilation failure","Ventilation system not checked pre-shift",
     "HAZ004","Yes","Yes","EMP009","In Progress","Yes",4,"Equipment Defect"],
    ["INC010","2025-07-01","2025-07-01 11:00:00","STN006","Injury","Minor",1,
     "Forklift struck racking at low speed. Driver minor whiplash.",
     "Speeding in warehouse zone","Speed limit signage obstructed",
     "HAZ006","No","No","EMP025","Completed","No",1,"Traffic Management"],
])

# ── Sheet 14: Near_Misses ────────────────────────────────────────────────────
ws = wb.create_sheet("Near_Misses")
headers = ["ID","Report_Date","Event_Date_Time","Location_Station_ID","Description",
           "Potential_Consequence","Hazard_ID","Underlying_Cause","Control_Failure",
           "Reported_By","CAPA_Escalation"]
style_sheet(ws, headers, [8,14,22,22,40,35,12,35,18,14,18])
add_rows(ws, [
    ["NM001","2025-01-12","2025-01-12 08:30:00","STN003",
     "Welding rod left on walkway overnight. Worker nearly tripped in morning.",
     "Trip and fall injury","HAZ004","Poor end-of-shift housekeeping","No","EMP004","Yes"],
    ["NM002","2025-01-20","2025-01-20 14:15:00","STN002",
     "Chemical drum valve left open. Small spill contained by bund. Noticed during inspection.",
     "Chemical exposure and contamination","HAZ001","Inadequate close-out procedure","Yes","EMP002","Yes"],
    ["NM003","2025-02-07","2025-02-07 10:45:00","STN001",
     "Guard rail on conveyor came loose. Spotted during daily safety walk before shift start.",
     "Entanglement and serious injury","HAZ003","Maintenance cycle missed","Yes","EMP001","Yes"],
    ["NM004","2025-02-25","2025-02-25 09:20:00","STN005",
     "Scaffold board cracked and deflected under worker weight. Work immediately stopped.",
     "Fall from height — potentially fatal","HAZ005","Board inspection not carried out","Yes","EMP019","Yes"],
    ["NM005","2025-03-14","2025-03-14 13:30:00","STN007",
     "Engineer found LV panel door open with live terminals exposed in electrical room.",
     "Electrocution","HAZ002","Previous engineer did not re-secure panel","Yes","EMP017","Yes"],
    ["NM006","2025-04-08","2025-04-08 07:50:00","STN004",
     "Gas monitor alarm sounded on entry to boiler room. Workers evacuated. False alarm confirmed.",
     "Oxygen deficiency — loss of consciousness","HAZ007","Monitor calibration overdue","Yes","EMP007","Yes"],
    ["NM007","2025-05-22","2025-05-22 15:10:00","STN006",
     "Forklift nearly struck pedestrian in unmarked crossing zone.",
     "Fatal or serious collision","HAZ006","Pedestrian route not demarcated","Yes","EMP025","Yes"],
    ["NM008","2025-06-18","2025-06-18 11:00:00","STN008",
     "Grinding disc cracked during use. Disc changed immediately. No injury.",
     "Eye or face injury from disc fragment","HAZ008","Disc not inspected before fitting","Yes","EMP011","Yes"],
    ["NM009","2025-07-03","2025-07-03 09:30:00","STN001",
     "Fire extinguisher found with broken seal and low pressure reading during walk.",
     "Inability to fight fire at source","HAZ004","Monthly inspection not completed","No","EMP003","No"],
    ["NM010","2025-07-10","2025-07-10 14:20:00","STN002",
     "Chemical label found missing on drum in storage area. Contents unknown.",
     "Chemical misidentification and exposure","HAZ001","Label fell off, not noticed during receipt","Yes","EMP022","Yes"],
])

# ── Sheet 15: Safety_Walks ───────────────────────────────────────────────────
ws = wb.create_sheet("Safety_Walks")
headers = ["ID","Inspection_Date_Time","Location_Station_ID","Inspector_ID",
           "Inspection_Type","Issues_Found","Critical_Issues",
           "Housekeeping_Rating","Compliance_Rating","Follow_Up_Required"]
style_sheet(ws, headers, [8,22,22,14,25,15,17,22,20,22])
add_rows(ws, [
    ["SW001","2025-01-06 08:00:00","STN001","EMP001","Routine Walk",2,0,4,4,"No"],
    ["SW002","2025-01-06 09:30:00","STN003","EMP001","Pre-Shift Inspection",1,0,3,3,"Yes"],
    ["SW003","2025-01-13 08:00:00","STN001","EMP002","Routine Walk",0,0,5,5,"No"],
    ["SW004","2025-01-13 09:00:00","STN002","EMP001","Chemical Zone Audit",3,1,3,3,"Yes"],
    ["SW005","2025-01-20 08:00:00","STN003","EMP003","Hot Work Pre-Inspection",1,0,4,4,"No"],
    ["SW006","2025-01-27 08:00:00","STN001","EMP001","Weekly Walk",2,0,4,4,"Yes"],
    ["SW007","2025-02-03 08:30:00","STN004","EMP007","Confined Space Inspection",2,1,3,3,"Yes"],
    ["SW008","2025-02-10 09:00:00","STN005","EMP019","WAH Inspection",3,1,3,3,"Yes"],
    ["SW009","2025-02-17 08:00:00","STN001","EMP001","Routine Walk",1,0,4,4,"No"],
    ["SW010","2025-02-24 08:30:00","STN006","EMP020","Warehouse Safety Walk",2,0,4,4,"Yes"],
    ["SW011","2025-03-03 08:00:00","STN001","EMP002","PPE Compliance Audit",4,2,3,3,"Yes"],
    ["SW012","2025-03-10 09:00:00","STN008","EMP030","Grinding Station Walk",2,0,4,3,"Yes"],
    ["SW013","2025-03-17 08:00:00","STN003","EMP001","Hot Work Zone Walk",1,0,4,4,"No"],
    ["SW014","2025-03-24 08:30:00","STN002","EMP002","Chemical Audit",2,1,3,3,"Yes"],
    ["SW015","2025-03-31 08:00:00","STN007","EMP017","Electrical Safety Walk",1,0,4,4,"No"],
    ["SW016","2025-04-07 08:00:00","STN001","EMP001","Monthly Inspection",3,1,3,4,"Yes"],
    ["SW017","2025-04-14 09:00:00","STN005","EMP019","WAH Monthly Check",2,1,3,3,"Yes"],
    ["SW018","2025-04-21 08:00:00","STN004","EMP012","Boiler Room Inspection",1,0,4,4,"No"],
    ["SW019","2025-05-05 08:00:00","STN001","EMP002","Routine Walk",0,0,5,5,"No"],
    ["SW020","2025-05-12 09:00:00","STN006","EMP020","Forklift Zone Audit",2,0,4,4,"Yes"],
    ["SW021","2025-05-19 08:00:00","STN001","EMP001","Weekly Walk",1,0,5,4,"No"],
    ["SW022","2025-06-02 08:00:00","STN003","EMP003","Hot Work Inspection",2,0,4,4,"Yes"],
    ["SW023","2025-06-09 08:30:00","STN002","EMP001","Chemical Zone Walk",1,0,4,4,"No"],
    ["SW024","2025-06-16 08:00:00","STN001","EMP002","Routine Walk",0,0,5,5,"No"],
    ["SW025","2025-06-23 09:00:00","STN008","EMP030","Grinding Safety Check",2,0,4,3,"Yes"],
    ["SW026","2025-06-30 08:00:00","STN001","EMP001","Monthly Walk",3,1,3,4,"Yes"],
    ["SW027","2025-07-07 08:30:00","STN005","EMP019","WAH Weekly Check",1,0,4,4,"No"],
    ["SW028","2025-07-07 09:00:00","STN001","EMP002","PPE Walk",1,0,4,4,"No"],
])

# ── Sheet 16: CAPA_Actions ───────────────────────────────────────────────────
ws = wb.create_sheet("CAPA_Actions")
headers = ["ID","Incident_ID","Action_Type","Description","Root_Cause_Addressed",
           "Responsible_Person_ID","Due_Date","Status","Effectiveness_Rating"]
style_sheet(ws, headers, [8,14,22,45,35,22,14,18,22])
add_rows(ws, [
    ["CAPA001","INC001","Corrective","Mandatory FR sleeve usage enforced at welding bay. PPE inspection checklist updated.",
     "PPE Non-Compliance","EMP001","2025-01-25","Completed",4],
    ["CAPA002","INC001","Preventive","Monthly PPE compliance audit introduced for all hot work zones.",
     "PPE Non-Compliance","EMP002","2025-02-15","Completed",4],
    ["CAPA003","INC002","Corrective","Manual handling risk assessment updated. Team briefing conducted.",
     "Training Gap","EMP003","2025-02-01","Completed",3],
    ["CAPA004","INC002","Preventive","Mechanical lifting aids procured for all components over 15kg.",
     "Training Gap","EMP012","2025-03-01","Completed",5],
    ["CAPA005","INC003","Corrective","Emergency maintenance carried out on boiler relief valve. Valve replaced.",
     "Maintenance Failure","EMP007","2025-02-15","Completed",5],
    ["CAPA006","INC003","Preventive","Preventive maintenance schedule reviewed. Inspection frequency increased to monthly.",
     "Maintenance Failure","EMP012","2025-03-15","Completed",4],
    ["CAPA007","INC004","Corrective","All chemical handling gloves replaced. Pre-use inspection procedure introduced.",
     "PPE Non-Compliance","EMP001","2025-03-10","In Progress",0],
    ["CAPA008","INC004","Preventive","Chemical handling SOP revised. Supervisor sign-off required before task start.",
     "PPE Non-Compliance","EMP022","2025-03-31","Open",0],
    ["CAPA009","INC005","Corrective","All grinding discs removed from service. New batch inspected and approved.",
     "Equipment Defect","EMP030","2025-03-25","Completed",4],
    ["CAPA010","INC005","Preventive","Pre-use equipment check card introduced at grinding station.",
     "Equipment Defect","EMP019","2025-04-10","Completed",4],
    ["CAPA011","INC006","Corrective","Worker retraining on WAH harness connection. Observed competency assessment conducted.",
     "Procedural Non-Compliance","EMP019","2025-04-20","In Progress",0],
    ["CAPA012","INC006","Preventive","WAH toolbox talk mandatory before every job. Buddy system introduced.",
     "Procedural Non-Compliance","EMP001","2025-05-01","Open",0],
    ["CAPA013","INC007","Corrective","LOTO procedure retraining for all electrical maintenance staff.",
     "Permit Violation","EMP017","2025-05-05","Completed",5],
    ["CAPA014","INC007","Preventive","Permit to Work board installed at electrical room entrance. No entry without permit.",
     "Permit Violation","EMP012","2025-05-20","Completed",5],
    ["CAPA015","INC008","Corrective","Spillage cleaned and anti-slip matting installed around assembly line.",
     "Housekeeping","EMP003","2025-05-20","Completed",3],
    ["CAPA016","INC009","Corrective","LEV fan repaired. Pre-shift ventilation check added to hot work checklist.",
     "Equipment Defect","EMP007","2025-07-01","In Progress",0],
    ["CAPA017","INC009","Preventive","Ventilation system added to quarterly maintenance schedule.",
     "Equipment Defect","EMP012","2025-07-31","Open",0],
])

# ── Sheet 17: Permits_To_Work ─────────────────────────────────────────────────
ws = wb.create_sheet("Permits_To_Work")
headers = ["ID","Permit_Type_ID","Date_Issued","Time_Issued","Location_Station_ID",
           "Work_Description","Duration_Requested_Hours","Issued_By","Approved_By",
           "Validity_Start","Validity_End","Work_Start_Actual","Work_End_Actual",
           "Number_Of_Workers","Status","Deviation_Reported","Incident_Occurred"]
style_sheet(ws, headers, [8,14,14,13,22,40,24,12,12,22,22,22,22,18,14,20,20])
add_rows(ws, [
    ["PTW001","PT001","2025-01-08","08:00","STN003","MIG welding repair on assembly jig frame",8,
     "EMP001","EMP005","2025-01-08 08:00:00","2025-01-08 16:00:00","2025-01-08 08:15:00","2025-01-08 15:45:00",
     2,"Closed","No","Yes"],
    ["PTW002","PT002","2025-02-03","07:30","STN004","Boiler internal inspection after pressure incident",4,
     "EMP001","EMP012","2025-02-03 07:30:00","2025-02-03 11:30:00","2025-02-03 07:45:00","2025-02-03 11:15:00",
     2,"Closed","No","No"],
    ["PTW003","PT003","2025-02-25","09:00","STN005","Roof membrane repair on Building A",8,
     "EMP001","EMP005","2025-02-25 09:00:00","2025-02-25 17:00:00","2025-02-25 09:10:00","2025-02-25 16:30:00",
     3,"Closed","Yes","No"],
    ["PTW004","PT004","2025-03-15","08:00","STN007","Replacement of LV panel breaker unit",12,
     "EMP017","EMP012","2025-03-15 08:00:00","2025-03-15 20:00:00","2025-03-15 08:30:00","2025-03-15 17:00:00",
     1,"Closed","No","No"],
    ["PTW005","PT001","2025-04-02","08:00","STN003","Hot cutting of structural steel beams",8,
     "EMP001","EMP019","2025-04-02 08:00:00","2025-04-02 16:00:00","2025-04-02 08:10:00","2025-04-02 15:50:00",
     2,"Closed","No","Yes"],
    ["PTW006","PT006","2025-04-10","09:00","STN002","Chemical drum decanting and relabelling",8,
     "EMP001","EMP005","2025-04-10 09:00:00","2025-04-10 17:00:00","2025-04-10 09:05:00","2025-04-10 16:45:00",
     2,"Closed","No","No"],
    ["PTW007","PT003","2025-05-20","08:00","STN005","HVAC unit installation at roof level",8,
     "EMP001","EMP019","2025-05-20 08:00:00","2025-05-20 16:00:00","2025-05-20 08:15:00","2025-05-20 15:30:00",
     4,"Closed","No","No"],
    ["PTW008","PT005","2025-06-01","08:00","STN001","Conveyor belt replacement — Line 1",12,
     "EMP003","EMP001","2025-06-01 08:00:00","2025-06-01 20:00:00","2025-06-01 08:30:00","2025-06-01 18:00:00",
     5,"Closed","No","No"],
    ["PTW009","PT001","2025-07-05","09:00","STN003","Welding repairs to turbine frame mounting",8,
     "EMP001","EMP005","2025-07-05 09:00:00","2025-07-05 17:00:00","2025-07-05 09:10:00","",
     2,"Active","No","No"],
    ["PTW010","PT002","2025-07-06","07:00","STN004","Annual internal vessel inspection",4,
     "EMP001","EMP012","2025-07-06 07:00:00","2025-07-06 11:00:00","","",
     2,"Active","No","No"],
])

# ── Sheet 18: Equipment_Certifications ───────────────────────────────────────
ws = wb.create_sheet("Equipment_Certifications")
headers = ["ID","Equipment_Name","Equipment_Type","Site_ID","Zone","Serial_Number",
           "Manufacturer","Model","Certification_Type","Certified_By",
           "Issue_Date","Expiry_Date","Next_Inspection_Date","Compliance_Standard"]
style_sheet(ws, headers, [8,35,22,10,20,18,22,18,25,28,14,14,22,25])
add_rows(ws, [
    ["CERT001","MIG Welder Unit 1","Welding Equipment","SITE001","Welding Bay",
     "WLD-2019-001","Lincoln Electric","Power MIG 260","Portable Appliance Test",
     "TUV UK Ltd","2025-01-10","2026-01-10","2025-07-10","BS EN 60974-1"],
    ["CERT002","Overhead Crane — Bay 1","Lifting Equipment","SITE001","Assembly Line 1",
     "CRANE-2015-001","Demag","KBK","Thorough Examination",
     "Lloyds Register","2024-12-01","2025-12-01","2025-06-01","LOLER 1998"],
    ["CERT003","Forklift Truck FLT01","Mobile Plant","SITE002","Warehouse Bay 1",
     "FLT-2020-001","Toyota","8FBN25","Thorough Examination",
     "Briggs Equipment","2025-02-15","2026-02-15","2025-08-15","LOLER 1998"],
    ["CERT004","Pressure Vessel — Boiler 1","Pressure Equipment","SITE001","Boiler Room",
     "BLR-2010-001","Fulton","VTG-50","Written Scheme Inspection",
     "TUV UK Ltd","2025-01-20","2026-01-20","2025-07-20","PSSR 2000"],
    ["CERT005","Electrical Distribution Board","Electrical","SITE001","Electrical Room",
     "EDB-2018-001","Schneider Electric","Prisma","Fixed Wiring Inspection",
     "NICEIC Approved","2023-06-01","2028-06-01","2026-06-01","BS 7671"],
    ["CERT006","Scaffold System — North Plant","Temporary Works","SITE001","North Roof",
     "SCAF-2025-001","SGB Group","Cuplok","Scaffold Inspection",
     "CISRS Inspector","2025-06-15","2025-07-13","2025-07-13","BS EN 12811"],
    ["CERT007","Harness — WAH Kit 01","Personal Protective Equipment","SITE001","WAH Store",
     "HAR-2023-001","Petzl","Vertex","PPE Inspection",
     "In-House Competent Person","2025-01-05","2026-01-05","2025-07-05","EN 361"],
    ["CERT008","Fire Extinguisher Bank — Line 1","Fire Equipment","SITE001","Assembly Line 1",
     "FE-LINE1-001","Firechief","Safeguard CO2 5kg","Annual Service",
     "Chubb Fire & Security","2025-03-01","2026-03-01","2026-03-01","BS 5306-3"],
    ["CERT009","Gas Detector — Boiler Room","Safety Monitoring","SITE001","Boiler Room",
     "GAS-2022-001","Crowcon","Clip SGD","Calibration & Service",
     "Crowcon Service Centre","2025-04-10","2026-04-10","2025-10-10","BS EN 45544"],
    ["CERT010","Angle Grinder 125mm — Grind01","Power Tools","SITE003","Grinding Station",
     "GRD-2024-001","Bosch","GWS 750-125","Portable Appliance Test",
     "In-House PAT Tester","2025-01-12","2026-01-12","2025-07-12","BS EN 60745"],
])

# ── Save workbook ─────────────────────────────────────────────────────────────
output_path = r"c:\Users\Navnath\Desktop\HSE\hse_old_ui\HSE_Import_Template.xlsx"
wb.save(output_path)
print(f"✅ Excel file saved: {output_path}")
print(f"   Total sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    ws = wb[s]
    print(f"   - {s}: {ws.max_row - 1} data rows")
