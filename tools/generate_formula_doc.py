"""
Generate HSE Intelligence — Formula Documentation Excel
All KPI formulas used in the system, organised by page/tab.
Run: python generate_formula_doc.py
Output: HSE_Formula_Documentation.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── Style helpers ─────────────────────────────────────────────────────────────
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

H1_FILL = PatternFill("solid", fgColor="1E3A5F")
H1_FONT = Font(color="FFFFFF", bold=True, size=13)

H2_FILL = PatternFill("solid", fgColor="2E75B6")
H2_FONT = Font(color="FFFFFF", bold=True, size=11)

ALT1_FILL = PatternFill("solid", fgColor="EBF3FB")
ALT2_FILL = PatternFill("solid", fgColor="FFFFFF")

WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
RED_FILL  = PatternFill("solid", fgColor="FCE4D6")

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def make_sheet(title):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = True
    return ws


def header_row(ws, row, cols, widths=None):
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=col)
        c.font = H2_FONT
        c.fill = H2_FILL
        c.alignment = CENTER
        c.border = BORDER
        if widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = 32


def title_row(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = H1_FONT
    c.fill = H1_FILL
    c.alignment = CENTER
    c.border = BORDER
    ws.row_dimensions[row].height = 36


def data_row(ws, row, values, fill=None):
    f = fill if fill else (ALT1_FILL if row % 2 == 0 else ALT2_FILL)
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.fill = f
        c.alignment = WRAP
        c.border = BORDER
    ws.row_dimensions[row].height = 60


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Dashboard / Leading Indicators
# ══════════════════════════════════════════════════════════════════════════════
ws = make_sheet("1. Dashboard")
COLS = ["KPI Name", "Formula", "Inputs Required", "Data Source (DB Table)", "Threshold / Label Logic", "Notes"]
WIDTHS = [28, 55, 38, 30, 38, 40]
title_row(ws, 1, "DASHBOARD — Leading Indicator Formulas", len(COLS))
header_row(ws, 2, COLS, WIDTHS)

dashboard_data = [
    [
        "Predictive Injury Risk Score (%)",
        "Score = (weight_sum / (count × 3)) × 100\n\nwhere weight_sum = Σ severity_weight per incident in last 90 days\n\nSeverity weights:\nCritical/Significant = 3\nHigh/Major = 2\nMedium/Moderate = 1\nAll others = 0.5",
        "incidents table\nincident_date_time (last 90 days)\nseverity column",
        "incidents",
        "< 30% → Low Risk\n30–60% → Medium Risk\n> 60% → High Risk\n\nTrend = current_score - previous_score (prev 90 days)",
        "Anchored on latest incident date in DB, not today. Prevents empty windows when historical data exists."
    ],
    [
        "TRIR\n(Total Recordable Incident Rate)",
        "TRIR = (Recordable Injuries × 200,000) / Man Hours\n\nWhere:\nRecordable = incident_type = 'Injury'\nMan Hours = SUM(actual_hours_worked) from shift_schedule",
        "incidents (type='Injury')\nshift_schedule.actual_hours_worked",
        "incidents\nshift_schedule",
        "Industry standard: < 3.0 = Good\n3–5 = Monitor\n> 5 = Action Required",
        "OSHA standard formula. 200,000 = 100 workers × 2,000 hrs/year."
    ],
    [
        "LTIFR\n(Lost Time Injury Frequency Rate)",
        "LTIFR = (LTI × 1,000,000) / Man Hours\n\nWhere:\nLTI = incidents with type='Injury' AND severity='Lost Time'",
        "incidents (type='Injury', severity='Lost Time')\nshift_schedule.actual_hours_worked",
        "incidents\nshift_schedule",
        "< 1.0 = Good\n1–3 = Monitor\n> 3 = Critical",
        "ILO standard formula. 1,000,000 = per million hours worked."
    ],
    [
        "LTISR\n(Lost Time Injury Severity Rate)",
        "LTISR = (Lost Days × 1,000,000) / Man Hours\n\nWhere:\nLost Days = SUM(days_away) for LTI incidents",
        "incidents.days_away (LTI only)\nshift_schedule.actual_hours_worked",
        "incidents\nshift_schedule",
        "Measures severity, not just frequency. High LTISR with low LTIFR = few but serious injuries.",
        "ILO standard formula."
    ],
    [
        "DART Rate\n(Days Away Restricted or Transferred)",
        "DART = (LTI × 200,000) / Man Hours",
        "incidents (LTI count)\nshift_schedule.actual_hours_worked",
        "incidents\nshift_schedule",
        "OSHA benchmark: < 2.0 = Good",
        "Uses same LTI count as LTIFR but with 200,000 base (OSHA standard)."
    ],
    [
        "FAR\n(Fatal Accident Rate)",
        "FAR = (Fatalities × 100,000,000) / Man Hours",
        "incidents (severity='Fatal')\nshift_schedule.actual_hours_worked",
        "incidents\nshift_schedule",
        "0 = Target. Any value > 0 = Critical.",
        "UK HSE standard. Per 100 million hours worked."
    ],
    [
        "Near Miss Ratio",
        "Ratio = Near Miss Count : Recordable Injuries\n\nDisplayed as 'X.X : 1'",
        "near_misses count\nincidents (type='Injury') count",
        "near_misses\nincidents",
        "Higher ratio = better reporting culture\nTarget: > 10:1 (every injury should have 10+ near misses reported)",
        "Leading indicator of safety culture health."
    ],
    [
        "Safe Days",
        "Safe Days = (latest_data_date - last_LTI_date).days",
        "Latest incident_date_time where type='Injury' AND severity='Lost Time'",
        "incidents",
        "Resets to 0 on any new Lost Time Incident.",
        "Simple days-since-last-LTI counter."
    ],
    [
        "Incident Close-Out Rate (%)",
        "Rate = (Completed Investigations / Total Investigations) × 100",
        "incidents.investigation_status = 'Completed'\nTotal incident count",
        "incidents",
        "< 70% = Needs Attention\n70–90% = Good\n> 90% = Excellent",
        "Tracks investigation completion discipline."
    ],
]

for i, row in enumerate(dashboard_data, 3):
    data_row(ws, i, row)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Contractor Risk Score (Vendors Page)
# ══════════════════════════════════════════════════════════════════════════════
ws = make_sheet("2. Contractor Risk (Vendors)")
COLS = ["Component", "Formula / Logic", "Inputs Required", "Data Source", "Range / Label", "Example (Actual Data)"]
WIDTHS = [28, 55, 38, 28, 32, 40]
title_row(ws, 1, "CONTRACTOR RISK SCORE — Full Formula Breakdown", len(COLS))
header_row(ws, 2, COLS, WIDTHS)

contractor_data = [
    [
        "1. Contractor Incident Rate",
        "Contractor Incident Rate = Contractor Incidents / Contractor Employees",
        "incidents joined to employees\nemployment_type LIKE '%contract%'",
        "incidents\nemployees",
        "Per employee rate",
        "Incidents: 17\nContractors: 13\nRate = 17/13 = 1.31"
    ],
    [
        "2. Organisation Incident Rate",
        "Org Incident Rate = Total Org Incidents / Total Org Employees",
        "All incidents for org_id\nAll employees for org_id",
        "incidents\nemployees",
        "Per employee rate",
        "Total incidents: 270\nTotal employees: 765 (approx)\nRate = 270/765 = 0.35"
    ],
    [
        "3. Relative Risk",
        "Relative Risk = Contractor Incident Rate / Org Incident Rate",
        "Computed from steps 1 & 2",
        "Derived",
        "1.0 = same as org avg\n> 1.0 = higher risk than avg\n< 1.0 = lower risk",
        "1.31 / 0.35 = 3.7\n(contractors 3.7× more incidents than org average)"
    ],
    [
        "4. Incident Penalty",
        "Incident Penalty = min(7.0, Relative Risk × 3.0)\n\nMax penalty for incidents = 7 points",
        "Relative Risk from step 3",
        "Derived",
        "0 = no contractor incidents\n7 = maximum (capped)",
        "min(7.0, 3.7 × 3.0)\n= min(7.0, 11.1)\n= 7.0"
    ],
    [
        "5. Permit Violation Count",
        "Count of permits_to_work WHERE deviation_reported = 'Yes'",
        "permits_to_work.deviation_reported",
        "permits_to_work",
        "Each violation = 0.5 penalty points\nMax 3 points penalty",
        "6 permit violations found"
    ],
    [
        "6. Violation Penalty",
        "Violation Penalty = min(3.0, Violation Count × 0.5)\n\nMax penalty for violations = 3 points",
        "Violation count from step 5",
        "Derived",
        "0 = no violations\n3 = maximum (capped at 6+ violations)",
        "min(3.0, 6 × 0.5)\n= min(3.0, 3.0)\n= 3.0"
    ],
    [
        "7. FINAL: Contractor Risk Score (0–10)",
        "Score = max(0.0, 10 - Incident Penalty - Violation Penalty)\n\nBest score = 10 (no incidents, no violations)\nWorst score = 0 (maximum penalties on both)",
        "Incident Penalty + Violation Penalty",
        "Derived",
        "0–4 = High Risk (Red)\n5–7 = Medium Risk (Amber)\n8–10 = Low Risk (Green)",
        "max(0.0, 10 - 7.0 - 3.0)\n= max(0.0, 0.0)\n= 0.0/10 → HIGH RISK"
    ],
    [
        "8. Client Audit Rule",
        "\"If a violation exists, a 10/10 score is impossible.\"\n\nEnforced by: violation_penalty > 0 whenever violations > 0\n\nMinimum deduction = 0.5 for 1 violation",
        "Any permit with deviation_reported = 'Yes'",
        "permits_to_work",
        "1 violation → max 9.5/10\n2 violations → max 9.0/10\n6+ violations → max 7.0/10 (from violations alone)",
        "With 6 violations: penalty = 3.0\nEven with no incidents: max score = 10 - 0 - 3.0 = 7.0/10"
    ],
]

for i, row in enumerate(contractor_data, 3):
    fill = RED_FILL if i == 9 else (WARN_FILL if i in [4, 6, 7] else None)
    data_row(ws, i, row, fill)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Audit Readiness Score
# ══════════════════════════════════════════════════════════════════════════════
ws = make_sheet("3. Audit Readiness")
COLS = ["Step", "Formula / Logic", "Inputs", "Data Source", "Threshold / Label", "Notes"]
WIDTHS = [28, 55, 38, 28, 32, 40]
title_row(ws, 1, "AUDIT READINESS SCORE — Formula Breakdown", len(COLS))
header_row(ws, 2, COLS, WIDTHS)

audit_data = [
    [
        "1. Identify Walk Window",
        "If user selected a date range: use start_date → end_date\n\nIf no date filter: use last 90 days anchored on the LATEST safety walk date in DB (not today)",
        "safety_walks.inspection_date_time\nUser date filter params",
        "safety_walks",
        "Anchored on data, not current date — prevents 0% when historical data exists",
        "Latest walk in DB = 30 Dec 2025\nWindow = 01 Oct 2025 → 30 Dec 2025"
    ],
    [
        "2. Average Compliance Rating",
        "Avg Compliance = AVG(compliance_rating) for all safety walks within the window",
        "safety_walks.compliance_rating\n(scale: 1–5)",
        "safety_walks",
        "Rating scale: 1 = Very Poor, 5 = Excellent",
        "Example: 28 walks with avg rating = 3.78"
    ],
    [
        "3. Audit Readiness Score (%)",
        "Audit Readiness % = (Avg Compliance Rating / 5) × 100",
        "Average compliance rating from step 2",
        "Derived",
        "≥ 80% = Ready (Green)\n60–79% = Needs Attention (Amber)\n< 60% = Not Ready (Red)",
        "3.78 / 5 × 100 = 75.6% → Needs Attention"
    ],
]

for i, row in enumerate(audit_data, 3):
    data_row(ws, i, row)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Risk Matrix (Risk Page)
# ══════════════════════════════════════════════════════════════════════════════
ws = make_sheet("4. Risk Matrix (Risk Page)")
COLS = ["Component", "Formula / Logic", "Inputs", "Data Source", "Color Mapping", "Notes"]
WIDTHS = [28, 55, 38, 28, 38, 40]
title_row(ws, 1, "RISK MATRIX — Formula & Color Mapping", len(COLS))
header_row(ws, 2, COLS, WIDTHS)

risk_data = [
    [
        "Matrix Population",
        "Each hazard is placed in a matrix cell based on:\nRow = Severity (Catastrophic→Negligible)\nCol = Probability (Frequent→Improbable)\n\nCell count = number of active hazards at that severity × probability",
        "hazards.severity\nhazards.probability",
        "hazards",
        "Red = Catastrophic\nOrange = Urgent\nYellow = Borderline\nGreen = Acceptable",
        "5×5 matrix. Each cell shows count of hazards at that risk level."
    ],
    [
        "Severity Row Mapping",
        "Row 0 (top) = Fatal / Catastrophic\nRow 1 = Significant / Major / Lost Time\nRow 2 = Serious / High\nRow 3 = Moderate / Medium / Low\nRow 4 = Minor / Negligible",
        "hazards.severity text value",
        "hazards",
        "—",
        "Text matching is case-insensitive."
    ],
    [
        "Probability Column Mapping",
        "Col 0 = Frequent / Almost Certain\nCol 1 = Probable / Likely\nCol 2 = Possible / Occasional\nCol 3 = Unlikely / Remote\nCol 4 = Rare / Improbable",
        "hazards.probability text value",
        "hazards",
        "—",
        "Text matching is case-insensitive."
    ],
    [
        "Resolved Hazard Exclusion\n(Client Audit P6)",
        "A hazard is RESOLVED (auto-removed from matrix) when:\n1. ALL incidents linked to that hazard have investigation_status = 'Completed'\nAND\n2. ALL CAPA actions for those incidents have status = 'Completed'\n\nResolved hazards are excluded from the matrix count.",
        "incidents.investigation_status\ncapa_actions.status\nhazards.id",
        "hazards\nincidents\ncapa_actions",
        "Active = counted in matrix\nResolved = excluded automatically",
        "Client audit: closed risks must vanish from matrix AND aging simultaneously."
    ],
    [
        "Cell Color Taxonomy\n(Client Audit P3)",
        "Red = Catastrophic (rows 0, top section)\nOrange = Urgent (medium-high zone)\nYellow = Borderline (medium-low zone)\nGreen = Acceptable (low probability/severity)",
        "Cell position in matrix (row, col)",
        "Frontend only (matrixCells config)",
        "BUG FIXED: Yellow was wrongly mapped to Catastrophic.\nNow: Red=#DC2626, Orange=#EA580C, Yellow=#EAB308, Green=#16A34A",
        "Fixed in audit remediation P3."
    ],
]

for i, row in enumerate(risk_data, 3):
    data_row(ws, i, row)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Risk Aging (Risk Page)
# ══════════════════════════════════════════════════════════════════════════════
ws = make_sheet("5. Risk Aging (Risk Page)")
COLS = ["Component", "Formula / Logic", "Inputs", "Data Source", "Buckets / Labels", "Notes"]
WIDTHS = [28, 55, 38, 28, 38, 40]
title_row(ws, 1, "RISK AGING VISUALIZER — Formula Breakdown", len(COLS))
header_row(ws, 2, COLS, WIDTHS)

aging_data = [
    [
        "Open CAPA Count",
        "All CAPA actions WHERE status != 'Completed'\nOrdered by due_date ASC",
        "capa_actions.status\ncapa_actions.due_date",
        "capa_actions",
        "Any non-Completed status counts as open",
        "Closed/Completed CAPAs are excluded — they represent resolved risks."
    ],
    [
        "Aging Bucket Assignment",
        "days_over = (today - due_date).days\n\n0–30 days overdue → '0-30 Days'\n31–60 days → '31-60 Days'\n61–90 days → '61-90 Days'\n>90 days / no date → '>90 Days'",
        "capa_actions.due_date",
        "capa_actions",
        "0-30 = Early warning\n31-60 = Escalating\n61-90 = High risk\n>90 = Critical",
        "Null due_date → placed in '>90 Days' bucket (worst case assumption)."
    ],
    [
        "Severity within Bucket",
        "Each bucket bar is stacked by severity:\nLow = not yet overdue (future due date)\nMedium = 1–30 days overdue OR no date\nHigh = 31–60 days overdue\nCritical = > 60 days overdue",
        "capa_actions.due_date vs today",
        "capa_actions",
        "Stack colors: Green=Low, Yellow=Medium, Orange=High, Red=Critical",
        "A CAPA can move buckets as time passes."
    ],
    [
        "Recently Closed Count",
        "Count of CAPAs WHERE status = 'Completed' AND due_date >= (today - 7 days)\n\nShown as badge: '✅ X closed this week'",
        "capa_actions.status = 'Completed'\ncapa_actions.due_date (last 7 days)",
        "capa_actions",
        "Positive indicator — shows risks actively being resolved",
        "Client audit P6: closed risks must simultaneously vanish from matrix AND aging."
    ],
    [
        "Control Effectiveness Score (%)",
        "Effectiveness = (Completed CAPAs / Total CAPAs) × 100",
        "capa_actions.status = 'Completed'\nTotal capa_actions count",
        "capa_actions",
        "< 50% = Poor\n50–80% = Moderate\n> 80% = Good",
        "Shown as KPI card on Risk page."
    ],
]

for i, row in enumerate(aging_data, 3):
    data_row(ws, i, row)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Compliance Score (Compliance Page)
# ══════════════════════════════════════════════════════════════════════════════
ws = make_sheet("6. Compliance (Compliance Page)")
COLS = ["KPI", "Formula", "Inputs", "Data Source", "Threshold", "Notes"]
WIDTHS = [28, 55, 38, 28, 32, 40]
title_row(ws, 1, "COMPLIANCE PAGE — Formula Breakdown", len(COLS))
header_row(ws, 2, COLS, WIDTHS)

compliance_data = [
    [
        "PTW Compliance Rate (%)",
        "Rate = (Closed Permits / Total Permits) × 100\n\nClosed = permits_to_work.status = 'Closed'\nActive = excluded from base (still open)\nExpired = non-compliant (counts against score)",
        "permits_to_work.status",
        "permits_to_work",
        "< 60% = Non-Compliant\n60–85% = Needs Improvement\n> 85% = Compliant",
        "Excel KPI spec M4_Assets_Operations formula."
    ],
    [
        "Policy Review Rate (%)",
        "Rate = (Current Policies / Total Policies) × 100\n\nCurrent = policies.status = 'Current'",
        "policies.status",
        "policies",
        "< 70% = Overdue reviews\n> 90% = Good",
        "Tracks how up-to-date safety policies are."
    ],
    [
        "Legal Register Coverage (%)",
        "Coverage = (Distinct Policy Categories / Distinct Hazard Categories) × 100\n\nMax capped at 100%",
        "policies.category (distinct count)\nhazard_categories (distinct count)",
        "policies\nhazard_categories",
        "< 60% = Low\n60–85% = Medium\n> 85% = High",
        "Measures how many hazard types have a corresponding policy."
    ],
    [
        "Overall Compliance Score (%)",
        "Score = (PTW Rate + Legal Register + Audit Readiness) / 3\n\nSimple average of the three sub-scores",
        "PTW Rate\nLegal Register\nAudit Readiness",
        "Derived",
        "< 70% = Needs Improvement\n70–85% = Good\n> 85% = Excellent",
        "Single headline compliance number for leadership reporting."
    ],
    [
        "Findings by Severity",
        "From Compliance-type safety walks:\nCritical = walks with critical_issues ≥ 2\nMajor = walks with critical_issues = 1\nMinor = walks with issues_found ≥ 1 AND critical_issues = 0\nObservation = walks with no issues",
        "safety_walks.critical_issues\nsafety_walks.issues_found\nWhere inspection_type = 'Compliance'",
        "safety_walks",
        "Critical findings → immediate corrective action required",
        "Pie chart on Compliance page."
    ],
]

for i, row in enumerate(compliance_data, 3):
    data_row(ws, i, row)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — Vendors / Contractor Compliance
# ══════════════════════════════════════════════════════════════════════════════
ws = make_sheet("7. Vendors Page")
COLS = ["KPI", "Formula", "Inputs", "Data Source", "Threshold", "Notes"]
WIDTHS = [28, 55, 38, 28, 32, 40]
title_row(ws, 1, "VENDORS PAGE — Formula Breakdown", len(COLS))
header_row(ws, 2, COLS, WIDTHS)

vendor_data = [
    [
        "Contractor Compliance % (Pie)",
        "Compliant = contractors with induction AND no incidents\nNon-Compliant = contractors with incidents\nPending = contractors with no induction date\n\nAll as % of total contractors",
        "employees (employment_type LIKE '%contract%')\nincidents.reported_by\nemployees.induction_date",
        "employees\nincidents",
        "100% Compliant = ideal\nAny Non-Compliant = action required",
        "Uses all-time incidents (not 90-day window) because historical data spans 2024-2025."
    ],
    [
        "Contractor Risk Score (0–10)\n[Vendors Page]",
        "Same formula as Dashboard:\nScore = max(0.0, 10 - Incident Penalty - Violation Penalty)\n\nIncident Penalty = min(7.0, Relative Risk × 3.0)\nViolation Penalty = min(3.0, Permit Violations × 0.5)\nRelative Risk = Contractor Inc Rate / Org Inc Rate",
        "incidents (contractor vs org)\npermits_to_work.deviation_reported = 'Yes'",
        "incidents\nemployees\npermits_to_work",
        "0–4 = High Risk\n5–7 = Medium\n8–10 = Low",
        "CONSISTENT with Dashboard leading indicators."
    ],
    [
        "Permit Violations List",
        "All permits WHERE deviation_reported = 'Yes'\nOrdered by date_issued DESC\nLimit 5 (shown in breaches panel)\n\nSAME query used by both Vendors tab AND Work tab (P2 audit fix)",
        "permits_to_work.deviation_reported = 'Yes'\npermits_to_work.issued_by → employees.full_name",
        "permits_to_work\nemployees",
        "Any violation = requires action",
        "Client audit P2: Vendors and Work tabs must show same violation count. Now both use deviation_reported='Yes' as single source of truth."
    ],
    [
        "Exposure Hours",
        "Hours = SUM(number_of_workers × duration_requested_hours) per month\nFiltered to last 9 months",
        "permits_to_work.number_of_workers\npermits_to_work.duration_requested_hours\npermits_to_work.date_issued",
        "permits_to_work",
        "Threshold = AVG monthly hours × 1.2 (20% above average)",
        "Shown as bar chart with threshold line."
    ],
    [
        "Repeat Breaches",
        "Contractors with incident count > 1\nOrdered by incident count DESC",
        "incidents (joined to contractor employees)\nHAVING count > 1",
        "incidents\nemployees",
        "Any repeat breach = watchlist candidate",
        "Pattern detection for high-risk contractors."
    ],
    [
        "High Risk Contractors",
        "Top 5 contractors by incident count\nRisk score = 50 + (incidents / max_incidents × 49)\nCapped at 99%",
        "incidents.reported_by → employees\nemployment_type LIKE '%contract%'",
        "incidents\nemployees",
        "Risk > 75% = Critical\n50–75% = High",
        "Relative risk within contractor group."
    ],
]

for i, row in enumerate(vendor_data, 3):
    data_row(ws, i, row)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — Work / Permits Page
# ══════════════════════════════════════════════════════════════════════════════
ws = make_sheet("8. Work Page (Permits)")
COLS = ["KPI", "Formula", "Inputs", "Data Source", "Threshold", "Notes"]
WIDTHS = [28, 55, 38, 28, 32, 40]
title_row(ws, 1, "WORK PAGE — Permits Formula Breakdown", len(COLS))
header_row(ws, 2, COLS, WIDTHS)

work_data = [
    [
        "Active Permit Count",
        "COUNT(permits_to_work) WHERE status = 'Active'",
        "permits_to_work.status",
        "permits_to_work",
        "Monitor any permits approaching validity_end",
        "Real-time count of open permits."
    ],
    [
        "PTW Compliance Rate (%)",
        "Rate = (Closed Permits / Total Permits) × 100",
        "permits_to_work.status",
        "permits_to_work",
        "< 60% = Non-Compliant\n> 85% = Compliant",
        "Same formula as Compliance page — single source of truth."
    ],
    [
        "Permit Violations\n(Work Tab — P2 Fix)",
        "COUNT of permits WHERE deviation_reported = 'Yes'\nWith details: PTW-XXXX, station, date\n\nNow SAME as Vendors tab — both use deviation_reported='Yes'",
        "permits_to_work.deviation_reported = 'Yes'\npermits_to_work.location_station_id → working_stations",
        "permits_to_work\nworking_stations",
        "0 = Clean\nAny violation = review required",
        "Client audit P2 fix: Work tab was showing 0 while Vendors showed 5. Now both show same data."
    ],
    [
        "Work Exposure Hours",
        "SUM(number_of_workers × duration_requested_hours)\nFor Active permits only",
        "permits_to_work.number_of_workers\npermits_to_work.duration_requested_hours\nWhere status = 'Active'",
        "permits_to_work",
        "Tracks total worker-hours at risk on active permits",
        "Useful for TRIR normalisation."
    ],
    [
        "Contractor Permit Compliance",
        "Compliant = Closed permits issued by/approved by contractor employees\nNon-Compliant = deviation_reported = 'Yes' permits\n\nRate = Compliant / Total × 100",
        "permits_to_work (contractor filter)\nemployees.employment_type LIKE '%contract%'",
        "permits_to_work\nemployees",
        "Tracks contractor-specific permit compliance separately from permanent staff",
        "Shown in Work tab contractor breakdown."
    ],
    [
        "Permit by Type Status (%)",
        "For each permit type: Active %, Closed %, Expired %\n\nActive = count('Active') / total × 100\nClosed = count('Closed') / total × 100\nExpired = count('Expired') / total × 100",
        "permit_types.permit_type_name\npermits_to_work.status",
        "permits_to_work\npermit_types",
        "High expired % = process breakdown",
        "Horizontal stacked bar chart per permit type."
    ],
]

for i, row in enumerate(work_data, 3):
    data_row(ws, i, row)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 9 — Date Filter Logic
# ══════════════════════════════════════════════════════════════════════════════
ws = make_sheet("9. Date Filter Logic")
COLS = ["KPI / Endpoint", "Default Window (No Filter)", "With Date Filter", "Anchor Strategy", "Affected Fields", "Notes"]
WIDTHS = [28, 40, 40, 38, 35, 40]
title_row(ws, 1, "DATE FILTER — How It Affects Each KPI", len(COLS))
header_row(ws, 2, COLS, WIDTHS)

date_data = [
    [
        "Total Incidents\nNear Misses\nSafety Walks\nCritical Incidents",
        "Anchored on latest incident_date_time in DB (not today)\nPrevents empty windows for historical data",
        "Filtered by user-selected start_date / end_date",
        "max(incident_date_time) from DB",
        "incidents.incident_date_time\nnear_misses.event_date_time\nsafety_walks.inspection_date_time",
        "P1 audit fix: was anchored on today() which caused 0% for 2025 data when accessed in 2026."
    ],
    [
        "Predictive Injury Risk Score\nTRIR / LTIFR / LTISR / DART / FAR",
        "Last 90 days from latest incident date in DB",
        "User date range used as the window",
        "max(incident_date_time) from DB",
        "incidents.incident_date_time",
        "KPIs respond to the selected period on dashboard."
    ],
    [
        "Audit Readiness Score",
        "Last 90 days from latest safety walk date in DB",
        "User date range for walk selection",
        "max(inspection_date_time) from DB",
        "safety_walks.inspection_date_time",
        "Anchored on walk data, not incident data."
    ],
    [
        "Open CAPA Count\nOverdue CAPA\nActive Permits\nTotal Employees\nTotal Sites",
        "ALWAYS org-wide (no date filter)\nThese are point-in-time health metrics",
        "NOT affected by date filter\n(intentional — they reflect current state)",
        "No date anchor — always live count",
        "capa_actions.status\npermits_to_work.status",
        "These metrics show current org status, not historical period data."
    ],
    [
        "CAPA Completion Rate (%)",
        "Org-wide (all CAPAs ever)\nNot date filtered",
        "NOT affected by date filter\n(intentional)",
        "No date anchor",
        "capa_actions.status",
        "Overall health metric, not period metric."
    ],
    [
        "Contractor Risk Score",
        "Uses all-time incident data\n(not rolling window)",
        "Incident penalty uses filtered period\nViolation penalty always org-wide",
        "All-time for historical accuracy",
        "incidents\npermits_to_work.deviation_reported",
        "Historical data (2024-2025) — 90-day rolling window would give 0 incidents in 2026."
    ],
    [
        "Dashboard Presets",
        "Default = 30D (last 30 days from latest data)",
        "7D / 30D / 90D / 1Y / All / Custom",
        "Preset dates calculated as: today - N days → today",
        "All date-filtered endpoints",
        "All preset → uses current calendar date for end.\nAnchor fix only applies when NO date filter is set."
    ],
]

for i, row in enumerate(date_data, 3):
    data_row(ws, i, row)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 10 — Summary / Quick Reference
# ══════════════════════════════════════════════════════════════════════════════
ws = make_sheet("10. Quick Reference")
COLS = ["KPI", "Formula (Short)", "Standard", "Page", "Result (Your Data)"]
WIDTHS = [30, 50, 25, 22, 28]
title_row(ws, 1, "QUICK REFERENCE — All KPI Formulas One-Page Summary", len(COLS))
header_row(ws, 2, COLS, WIDTHS)

summary_data = [
    ["TRIR",               "(Injuries × 200,000) / Man Hours",                    "OSHA",    "Dashboard",  "3.53"],
    ["LTIFR",              "(LTI × 1,000,000) / Man Hours",                       "ILO",     "Dashboard",  "1.61"],
    ["LTISR",              "(Lost Days × 1,000,000) / Man Hours",                 "ILO",     "Dashboard",  "1.61"],
    ["DART Rate",          "(LTI × 200,000) / Man Hours",                         "OSHA",    "Dashboard",  "0.32"],
    ["FAR",                "(Fatalities × 100,000,000) / Man Hours",              "UK HSE",  "Dashboard",  "0"],
    ["Near Miss Ratio",    "Near Misses / Recordable Injuries",                   "HSE",     "Dashboard",  "10.9 : 1"],
    ["Predictive Risk %",  "(Σ severity_weights / count×3) × 100 [90-day]",      "Custom",  "Dashboard",  "58.33%"],
    ["Audit Readiness %",  "(Avg compliance_rating / 5) × 100 [90-day]",         "Custom",  "Dashboard",  "75.6%"],
    ["Contractor Risk /10","10 - incident_penalty - violation_penalty",           "Custom",  "Dashboard / Vendors", "0.0/10"],
    ["Incident Penalty",   "min(7.0, relative_risk × 3.0)",                      "Custom",  "Vendors",    "7.0"],
    ["Violation Penalty",  "min(3.0, permit_violations × 0.5)",                  "Custom",  "Vendors",    "3.0"],
    ["Relative Risk",      "Contractor Inc Rate / Org Inc Rate",                  "Custom",  "Vendors",    "3.7"],
    ["PTW Compliance %",   "(Closed Permits / Total Permits) × 100",              "Custom",  "Compliance / Work", "—"],
    ["Legal Register %",   "(Policy Categories / Hazard Categories) × 100",      "Custom",  "Compliance", "—"],
    ["Overall Compliance","(PTW % + Legal % + Audit %) / 3",                      "Custom",  "Compliance", "—"],
    ["Control Effectiveness","(Completed CAPAs / Total CAPAs) × 100",            "Custom",  "Risk",       "—"],
    ["Safe Days",          "(Latest Data Date - Last LTI Date).days",             "Custom",  "Dashboard",  "699"],
    ["Incident Close-Out %","(Completed Investigations / Total) × 100",          "Custom",  "Dashboard",  "66.04%"],
]

for i, row in enumerate(summary_data, 3):
    f = GOOD_FILL if i % 2 == 0 else ALT2_FILL
    data_row(ws, i, row, f)


# ── Save ──────────────────────────────────────────────────────────────────────
output = r"C:\Users\Navnath\Desktop\HSE\hse_old_ui\HSE_Formula_Documentation.xlsx"
wb.save(output)
print(f"Saved: {output}")
print(f"Sheets: {wb.sheetnames}")
